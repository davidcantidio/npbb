"""Tests for XLSX leads extractor and staging load with normalized actions."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
import sys

import pytest
from openpyxl import Workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select


BACKEND_ROOT = Path(__file__).resolve().parents[1] / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.models.etl_lineage import LineageRef  # noqa: E402
from app.models.etl_registry import IngestionRun, IngestionStatus, Source  # noqa: E402
from app.models.stg_leads import StgLead, StgLeadAction  # noqa: E402
from etl.extract import xlsx_leads  # noqa: E402


def _make_engine():
    """Create isolated in-memory SQLite engine for leads staging tests."""
    return create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )


def _create_required_tables(engine) -> None:  # noqa: ANN001
    """Create minimal registry + lineage + leads staging tables for tests."""
    SQLModel.metadata.create_all(
        engine,
        tables=[
            Source.__table__,
            IngestionRun.__table__,
            LineageRef.__table__,
            StgLead.__table__,
            StgLeadAction.__table__,
        ],
    )


def _write_leads_fixture(path: Path) -> None:
    """Write a minimal XLSX fixture with merged header rows and lead data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    ws["A1"] = "Relatorio Leads"
    ws["A2"] = "Gerado em 2025-12-13"

    ws.merge_cells("A4:D4")
    ws["A4"] = "Lead"
    ws.merge_cells("E4:H4")
    ws["E4"] = "Evento"

    ws["A5"] = "Nome"
    ws["B5"] = "Sobrenome"
    ws["C5"] = "CPF"
    ws["D5"] = "Email"
    ws["E5"] = "Evento"
    ws["F5"] = "DataCriacao"
    ws["G5"] = "Acoes"
    ws["H5"] = "Estado"
    ws["I5"] = "Cidade"
    ws["J5"] = "Sexo"
    ws["K5"] = "Interesses"
    ws["L5"] = "AreaAtuacao"
    ws["M5"] = "CPFPromotor"
    ws["N5"] = "NomePromotor"

    ws["A6"] = "Alice"
    ws["B6"] = "Silva"
    ws["C6"] = "11111111111"
    ws["D6"] = "alice@example.com"
    ws["E6"] = "Festival"
    ws["F6"] = datetime(2025, 12, 13, 15, 0, 0)
    ws["G6"] = "Ativacao QR; Pesquisa | Cupom\nPesquisa, Cupom"
    ws["H6"] = "SP"
    ws["I6"] = "Sao Paulo"
    ws["J6"] = "F"
    ws["K6"] = "esportes"
    ws["L6"] = "marketing"
    ws["M6"] = "99999999999"
    ws["N6"] = "Promotor A"

    # Duplicate basic key (cpf/email/event/data) to validate loader dedupe.
    ws["A7"] = "Alice Dup"
    ws["B7"] = "Silva"
    ws["C7"] = "11111111111"
    ws["D7"] = "alice@example.com"
    ws["E7"] = "Festival"
    ws["F7"] = datetime(2025, 12, 13, 15, 0, 0)
    ws["G7"] = "Pesquisa; Pesquisa"
    ws["H7"] = "SP"
    ws["I7"] = "Sao Paulo"
    ws["J7"] = "F"

    # Second unique lead.
    ws["A8"] = "Bruno"
    ws["B8"] = "Costa"
    ws["C8"] = "22222222222"
    ws["D8"] = "bruno@example.com"
    ws["E8"] = "Festival"
    ws["F8"] = datetime(2025, 12, 13, 16, 30, 0)
    ws["G8"] = "Networking"
    ws["H8"] = "RJ"
    ws["I8"] = "Rio de Janeiro"
    ws["J8"] = "M"
    ws["N8"] = "Promotor B"

    wb.save(path)


def test_parse_actions_is_deterministic_and_deduplicates() -> None:
    """Actions parser should support configured delimiters and dedupe preserving order."""
    parsed = xlsx_leads.parse_actions("A; B | C\nB, C • D")
    assert parsed == ["A", "B", "C", "D"]


def test_extract_leads_xlsx_returns_rows_with_actions_and_lineage_metadata(tmp_path: Path) -> None:
    """Extractor should output canonical lead fields, parsed actions and lineage keys."""
    xlsx_path = tmp_path / "leads.xlsx"
    _write_leads_fixture(xlsx_path)

    rows = list(xlsx_leads.extract_leads_xlsx(xlsx_path))

    assert len(rows) == 3
    first = rows[0]
    assert first["evento"] == "Festival"
    assert first["estado"] == "SP"
    assert first["acoes_list"] == ["Ativacao QR", "Pesquisa", "Cupom"]
    assert first["cpf_hash"]
    assert first["email_hash"]
    assert first["person_key_hash"]
    assert first["cpf_promotor_hash"]
    assert first["__sheet_name"] == "Leads"
    assert first["__header_row"] == 4
    assert first["__header_range"] == "A4:N5"
    assert first["__source_range"] == "A6:N6"


def test_load_leads_xlsx_to_staging_persists_leads_actions_and_lineage(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Loader should persist staging leads/actions with ingestion and lineage references."""
    xlsx_path = tmp_path / "leads.xlsx"
    _write_leads_fixture(xlsx_path)

    engine = _make_engine()
    _create_required_tables(engine)
    monkeypatch.setattr(xlsx_leads, "engine", engine)

    summary = xlsx_leads.load_leads_xlsx_to_staging(
        source_id="src_leads_treze",
        xlsx_path=xlsx_path,
        lineage_policy="required",
        severity_on_missing="partial",
    )

    assert summary["source_id"] == "SRC_LEADS_TREZE"
    assert summary["ingestion_id"] > 0
    assert summary["leads_loaded"] == 2
    assert summary["actions_loaded"] == 4

    with Session(engine) as session:
        run = session.get(IngestionRun, summary["ingestion_id"])
        assert run is not None
        assert run.status == IngestionStatus.SUCCESS

        leads = session.exec(select(StgLead).order_by(StgLead.row_number)).all()
        actions = session.exec(select(StgLeadAction).order_by(StgLeadAction.id)).all()
        assert len(leads) == 2
        assert len(actions) == 4
        assert all(lead.lineage_ref_id > 0 for lead in leads)
        assert all(action.lineage_ref_id > 0 for action in actions)
        assert {action.action_raw for action in actions} == {
            "Ativacao QR",
            "Pesquisa",
            "Cupom",
            "Networking",
        }


def test_load_leads_xlsx_to_staging_marks_failed_when_lineage_missing(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Required lineage policy should fail ingestion when lineage IDs are missing."""
    xlsx_path = tmp_path / "leads.xlsx"
    _write_leads_fixture(xlsx_path)

    engine = _make_engine()
    _create_required_tables(engine)
    monkeypatch.setattr(xlsx_leads, "engine", engine)

    class _DummyLineage:
        id = 0

    monkeypatch.setattr(xlsx_leads, "create_lineage_ref", lambda *args, **kwargs: _DummyLineage())

    with pytest.raises(ValueError, match="exige linhagem"):
        xlsx_leads.load_leads_xlsx_to_staging(
            source_id="SRC_LEADS_MISSING_LINEAGE",
            xlsx_path=xlsx_path,
            lineage_policy="required",
            severity_on_missing="failed",
        )

    with Session(engine) as session:
        runs = session.exec(select(IngestionRun).order_by(IngestionRun.id.desc())).all()
        assert runs
        assert runs[0].status == IngestionStatus.FAILED
