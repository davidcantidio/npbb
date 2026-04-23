"""File extraction helpers for backend ETL preview."""

from __future__ import annotations

import csv
import hashlib
import io
import re
from typing import Any

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlmodel import Session, select

from app.models.models import ImportAlias
from app.services.imports.alias_service import upsert_alias
from app.observability.prometheus_leads_import import record_import_upload_rejection
from app.services.imports.file_reader import ImportFileError, inspect_upload
from core.leads_etl.transform.column_normalize import normalize_column_name
from etl.extract.xlsx_utils import HeaderNotFound, build_columns_with_metadata, find_header_row

from .contracts import (
    EtlCpfColumnRequired,
    EtlFieldAliasSelection,
    EtlHeaderColumn,
    EtlHeaderRequired,
)
from .exceptions import EtlImportContractError


HEADER_ALIAS_DOMAIN = "lead_import_etl_header"
REQUIRED_HEADER_FIELDS = ("cpf",)
_SUPPORTED_HEADER_FIELDS = {"cpf", "email"}
_SPACE_RE = re.compile(r"\s+")
_DEFAULT_COLUMN_ALIASES: dict[str, str] = {
    "cpf": "cpf",
    "email": "email",
    "e_mail": "email",
}

ETL_DEFAULT_MAX_SCAN_ROWS = 40
ETL_MAX_SCAN_ROWS_CAP = 500

ExtractRowsResult = tuple[list[dict[str, Any]], dict[str, Any]] | EtlHeaderRequired | EtlCpfColumnRequired


def clamp_etl_max_scan_rows(value: int | None) -> int:
    """Clamp operator-supplied scan window (API validates; this is the single backend guard)."""
    if value is None:
        return ETL_DEFAULT_MAX_SCAN_ROWS
    return max(1, min(int(value), ETL_MAX_SCAN_ROWS_CAP))


def read_upload_bytes(file: UploadFile, *, max_bytes: int) -> tuple[str, str, bytes]:
    try:
        filename, ext, _size = inspect_upload(file, max_bytes=max_bytes)
    except ImportFileError as exc:
        record_import_upload_rejection(exc, filename_hint=getattr(file, "filename", None))
        raise
    file.file.seek(0)
    payload = file.file.read()
    file.file.seek(0)
    return filename, ext, payload


def compute_file_fingerprint(filename: str, payload: bytes) -> str:
    digest = hashlib.sha256()
    digest.update(filename.encode("utf-8"))
    digest.update(b":")
    digest.update(payload)
    return digest.hexdigest()


def list_xlsx_sheet_titles(payload: bytes) -> list[str]:
    """Return workbook sheet titles without loading full cell data (read-only)."""
    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _clean_cell_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    return _SPACE_RE.sub(" ", text)


def _decode_csv_payload(payload: bytes) -> str:
    try:
        return payload.decode("utf-8-sig")
    except UnicodeDecodeError:
        return payload.decode("latin-1")


def _csv_max_field_count(sample_text: str, delimiter: str, *, max_probe_lines: int = 30) -> int:
    reader = csv.reader(io.StringIO(sample_text), delimiter=delimiter)
    best = 0
    for i, row in enumerate(reader):
        if i >= max_probe_lines:
            break
        best = max(best, len(row))
    return best


def _detect_csv_delimiter_robust(sample_text: str) -> str:
    """Pick CSV delimiter: tab-heavy first line, then Sniffer, then comma vs semicolon (tie -> comma)."""
    if not sample_text:
        return ","
    lines = sample_text.splitlines()
    first = lines[0] if lines else ""
    tab_n = first.count("\t")
    comma_n = first.count(",")
    semi_n = first.count(";")
    if tab_n > max(comma_n, semi_n):
        return "\t"

    probe = sample_text[:8192] if len(sample_text) > 8192 else sample_text
    try:
        dialect = csv.Sniffer().sniff(probe, delimiters=";\t,")
        d = dialect.delimiter
        if d in ("\t", ";", ",") and _csv_max_field_count(probe, d) >= 2:
            return d
    except csv.Error:
        pass

    if semi_n == comma_n:
        return ","
    return ";" if semi_n > comma_n else ","


def _load_csv_worksheet(payload: bytes) -> Worksheet:
    text = _decode_csv_payload(payload)
    delimiter = _detect_csv_delimiter_robust(text[:8192])
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)

    workbook = Workbook()
    ws = workbook.active
    ws.title = "CSV"
    try:
        for row in reader:
            ws.append([cell.strip() for cell in row])
    except csv.Error as exc:
        raise EtlImportContractError("CSV invalido para importacao ETL.") from exc
    return ws


def _header_columns(ws: Worksheet, header_row: int) -> tuple[EtlHeaderColumn, ...]:
    columns: list[EtlHeaderColumn] = []
    for col_idx in range(1, ws.max_column + 1):
        value = _clean_cell_text(ws.cell(row=header_row, column=col_idx).value)
        if not value:
            continue
        columns.append(
            EtlHeaderColumn(
                column_index=col_idx,
                column_letter=get_column_letter(col_idx),
                source_value=value,
            )
        )
    return tuple(columns)


def _columns_from_header_not_found(not_found: HeaderNotFound) -> tuple[EtlHeaderColumn, ...]:
    return tuple(
        EtlHeaderColumn(
            column_index=column.column_index,
            column_letter=column.column_letter,
            source_value=column.source_value,
        )
        for column in not_found.columns
    )


def _load_persisted_header_aliases(session: Session) -> dict[str, list[str]]:
    aliases: dict[str, list[str]] = {}
    rows = session.exec(
        select(ImportAlias).where(ImportAlias.domain == HEADER_ALIAS_DOMAIN)
    ).all()
    for row in rows:
        field_name = str(row.field_name or "").strip().lower()
        source_value = str(row.source_value or "").strip()
        if field_name not in _SUPPORTED_HEADER_FIELDS or not source_value:
            continue
        aliases.setdefault(field_name, []).append(source_value)
    return aliases


def _resolve_manual_alias_values(
    ws: Worksheet,
    *,
    header_row: int | None,
    field_aliases: dict[str, EtlFieldAliasSelection],
) -> dict[str, str]:
    resolved: dict[str, str] = {}
    for raw_field, selection in field_aliases.items():
        field = raw_field.strip().lower()
        if field not in _SUPPORTED_HEADER_FIELDS:
            continue
        source_value = (selection.source_value or "").strip()
        if not source_value and selection.column_index is not None:
            if header_row is None:
                raise EtlImportContractError("header_row e obrigatorio para alias por column_index.")
            if selection.column_index < 1 or selection.column_index > ws.max_column:
                raise EtlImportContractError("column_index de alias fora dos limites da planilha.")
            source_value = _clean_cell_text(ws.cell(row=header_row, column=selection.column_index).value)
        if not source_value:
            raise EtlImportContractError("Alias de campo deve informar source_value ou column_index valido.")
        resolved[field] = source_value
    return resolved


def _build_term_aliases(
    persisted_aliases: dict[str, list[str]],
    manual_alias_values: dict[str, str],
) -> dict[str, tuple[str, ...]]:
    aliases: dict[str, tuple[str, ...]] = {}
    for field_name, required_term in (("cpf", "CPF"), ("email", "Email")):
        values = [*persisted_aliases.get(field_name, [])]
        manual = manual_alias_values.get(field_name)
        if manual:
            values.append(manual)
        aliases[required_term] = tuple(dict.fromkeys(value for value in values if value.strip()))
    return aliases


def _build_column_alias_map(
    persisted_aliases: dict[str, list[str]],
    manual_alias_values: dict[str, str],
) -> dict[str, str]:
    alias_map = dict(_DEFAULT_COLUMN_ALIASES)
    for field_name, values in persisted_aliases.items():
        for value in values:
            normalized = normalize_column_name(value)
            if normalized:
                alias_map[normalized] = field_name
    for field_name, value in manual_alias_values.items():
        normalized = normalize_column_name(value)
        if normalized:
            alias_map[normalized] = field_name
    return alias_map


def _persist_manual_aliases(session: Session, manual_alias_values: dict[str, str]) -> None:
    for field_name, source_value in manual_alias_values.items():
        if field_name not in REQUIRED_HEADER_FIELDS:
            continue
        upsert_alias(
            session,
            domain=HEADER_ALIAS_DOMAIN,
            field_name=field_name,
            source_value=source_value,
            canonical_value=field_name,
            canonical_ref_id=None,
        )


def _select_xlsx_worksheet(workbook, sheet_name: str | None):
    """Resolve worksheet: first sheet when sheet_name omitted/blank; exact title match otherwise."""
    if sheet_name is None or not str(sheet_name).strip():
        return workbook.worksheets[0]
    name = str(sheet_name).strip()
    if name not in workbook.sheetnames:
        avail = ", ".join(repr(t) for t in workbook.sheetnames)
        raise EtlImportContractError(f"Aba XLSX inexistente: {name!r}. Abas disponiveis: {avail}")
    return workbook[name]


def _extract_worksheet_rows(
    ws: Worksheet,
    *,
    db: Session,
    header_row: int | None = None,
    field_aliases: dict[str, EtlFieldAliasSelection] | None = None,
    max_scan_rows: int = ETL_DEFAULT_MAX_SCAN_ROWS,
    promote_merged_header: bool = False,
    available_sheets: tuple[str, ...] = (),
    active_sheet: str | None = None,
) -> ExtractRowsResult:
    # Merged cells: xlsx_utils uses anchor value for empty cells in a merge (no silent column shift here).
    if header_row is not None and header_row < 1:
        raise EtlImportContractError("header_row deve ser 1-indexed e positivo.")

    scan_cap = min(max_scan_rows, ws.max_row) if ws.max_row else 0
    sheet_ctx = available_sheets
    active = active_sheet or ws.title

    persisted_aliases = _load_persisted_header_aliases(db)
    manual_alias_values = _resolve_manual_alias_values(
        ws,
        header_row=header_row,
        field_aliases=field_aliases or {},
    )
    term_aliases = _build_term_aliases(persisted_aliases, manual_alias_values)
    found_header_row = find_header_row(
        ws,
        required_terms=["CPF"],
        term_aliases=term_aliases,
        forced_row=header_row,
        soft_fail=True,
        promote_merged_header=promote_merged_header,
        max_scan_rows=max_scan_rows,
        min_non_empty_cells=1,
    )

    if isinstance(found_header_row, HeaderNotFound):
        if header_row is None:
            return EtlHeaderRequired(
                status="header_required",
                message="Nao foi possivel identificar automaticamente a linha do cabecalho com CPF.",
                max_row=ws.max_row,
                scanned_rows=found_header_row.scanned_rows,
                available_sheets=sheet_ctx,
                active_sheet=active,
            )
        return EtlCpfColumnRequired(
            status="cpf_column_required",
            message="A linha indicada nao contem uma coluna de CPF reconhecida.",
            header_row=header_row,
            columns=_columns_from_header_not_found(found_header_row),
            available_sheets=sheet_ctx,
            active_sheet=active,
        )

    resolved_header_row = int(found_header_row)
    column_aliases = _build_column_alias_map(persisted_aliases, manual_alias_values)
    columns_result = build_columns_with_metadata(
        ws,
        header_row=resolved_header_row,
        header_depth=1,
        aliases=column_aliases,
    )
    if "cpf" not in columns_result.columns:
        if header_row is None and not manual_alias_values:
            return EtlHeaderRequired(
                status="header_required",
                message="Nao foi possivel identificar automaticamente a linha do cabecalho com CPF.",
                max_row=ws.max_row,
                scanned_rows=scan_cap,
                available_sheets=sheet_ctx,
                active_sheet=active,
            )
        return EtlCpfColumnRequired(
            status="cpf_column_required",
            message="A linha indicada nao contem uma coluna de CPF reconhecida.",
            header_row=resolved_header_row,
            columns=_header_columns(ws, resolved_header_row),
            available_sheets=sheet_ctx,
            active_sheet=active,
        )

    _persist_manual_aliases(db, manual_alias_values)
    data_start_row = resolved_header_row + columns_result.lineage.header_depth
    rows: list[dict[str, Any]] = []

    for row_idx in range(data_start_row, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(columns_result.columns) + 1)]
        if all(value is None or not str(value).strip() for value in values):
            continue
        rows.append(
            {
                "__row_number": row_idx,
                "__sheet_name": ws.title,
                **{
                    columns_result.columns[index]: values[index]
                    for index in range(len(columns_result.columns))
                },
            }
        )

    meta_sheets = list(sheet_ctx)
    return rows, {
        "sheet_name": ws.title,
        "available_sheets": meta_sheets,
        "header_row": columns_result.lineage.header_row,
        "header_range": columns_result.lineage.header_range,
        "used_range": columns_result.lineage.used_range,
        "applied_field_aliases": manual_alias_values,
    }


def extract_xlsx_rows(
    payload: bytes,
    *,
    db: Session,
    header_row: int | None = None,
    field_aliases: dict[str, EtlFieldAliasSelection] | None = None,
    sheet_name: str | None = None,
    max_scan_rows: int = ETL_DEFAULT_MAX_SCAN_ROWS,
    promote_merged_header: bool = False,
) -> ExtractRowsResult:
    workbook = load_workbook(io.BytesIO(payload), read_only=False, data_only=True)  # Normal mode has ws.merged_cells.
    titles = tuple(workbook.sheetnames)
    ws = _select_xlsx_worksheet(workbook, sheet_name)
    return _extract_worksheet_rows(
        ws,
        db=db,
        header_row=header_row,
        field_aliases=field_aliases,
        max_scan_rows=max_scan_rows,
        promote_merged_header=promote_merged_header,
        available_sheets=titles,
        active_sheet=ws.title,
    )


def extract_csv_rows(
    payload: bytes,
    *,
    db: Session,
    header_row: int | None = None,
    field_aliases: dict[str, EtlFieldAliasSelection] | None = None,
    max_scan_rows: int = ETL_DEFAULT_MAX_SCAN_ROWS,
    promote_merged_header: bool = False,
) -> ExtractRowsResult:
    ws = _load_csv_worksheet(payload)
    return _extract_worksheet_rows(
        ws,
        db=db,
        header_row=header_row,
        field_aliases=field_aliases,
        max_scan_rows=max_scan_rows,
        promote_merged_header=promote_merged_header,
        available_sheets=(),
        active_sheet=ws.title,
    )
