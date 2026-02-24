"""Extractor for leads XLSX exports (Festival de Esportes).

Writes staging artifacts:
- CSV with canonical columns (PII hashed by default)
- JSONL with best-effort raw row dicts (PII removed by default)
- Evidence JSON envelope
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

from npbb.etl.extract.common import (
    EvidenceEnvelope,
    hash_pii,
    snapshot_file,
    utc_now_iso,
    write_csv,
    write_json,
    write_jsonl,
)
from npbb.etl.extract.xlsx_helpers import HeaderSpec, build_simple_header_spec, infer_last_data_row, to_iso


PII_KEYS = {"Nome", "Sobrenome", "RG", "CPF", "Email", "Telefone", "DataNascimento", "Cep"}


def _row_to_raw_dict(hdr: HeaderSpec, values: List[Any]) -> Dict[str, Any]:
    d: Dict[str, Any] = {}
    for i, h in enumerate(hdr.headers):
        d[h] = to_iso(values[i] if i < len(values) else None)
    return d


def extract_leads_xlsx(
    *,
    source_id: str,
    xlsx_path: Path,
    out_dir: Path,
    sheet_name: str | None = None,
    include_pii: bool = False,
) -> Dict[str, Any]:
    started = utc_now_iso()
    notes: List[str] = []

    snap = snapshot_file(xlsx_path)

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    hdr = build_simple_header_spec(ws, required_terms=["CPF", "Email", "Evento"])
    data_start = hdr.header_row + 1
    last_row = infer_last_data_row(ws, start_row=data_start, max_col=hdr.max_col)

    canonical_rows: List[Dict[str, Any]] = []
    raw_rows: List[Dict[str, Any]] = []

    for r in range(data_start, last_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, hdr.max_col + 1)]
        if all(v is None or str(v).strip() == "" for v in values):
            continue

        raw = _row_to_raw_dict(hdr, values)

        cpf = str(raw.get("CPF") or "").strip()
        email = str(raw.get("Email") or "").strip()

        canonical: Dict[str, Any] = {
            "source_id": source_id,
            "sheet_name": ws.title,
            "row_number": r,
            "evento": raw.get("Evento"),
            "cpf_hash": hash_pii(cpf) if cpf else "",
            "email_hash": hash_pii(email) if email else "",
            "sexo": raw.get("Sexo"),
            "estado": raw.get("Estado"),
            "cidade": raw.get("Cidade"),
            "data_criacao": raw.get("DataCriacao"),
            "acoes": raw.get("Acoes"),
            "interesses": raw.get("Interesses"),
            "area_atuacao": raw.get("AreaAtuacao"),
            "cpf_promotor": raw.get("CPFPromotor"),
            "nome_promotor": raw.get("NomePromotor"),
        }

        if include_pii:
            canonical.update(
                {
                    "cpf_raw": cpf,
                    "email_raw": email,
                    "nome": raw.get("Nome"),
                    "sobrenome": raw.get("Sobrenome"),
                    "telefone": raw.get("Telefone"),
                    "cep": raw.get("Cep"),
                    "data_nascimento": raw.get("DataNascimento"),
                    "rg": raw.get("RG"),
                }
            )

        canonical_rows.append(canonical)

        if include_pii:
            raw_rows.append(raw)
        else:
            scrubbed = {k: v for k, v in raw.items() if k not in PII_KEYS}
            scrubbed["cpf_hash"] = canonical["cpf_hash"]
            scrubbed["email_hash"] = canonical["email_hash"]
            raw_rows.append(scrubbed)

    finished = utc_now_iso()

    staging_csv = out_dir / f"stg_leads_festival__{source_id}.csv"
    staging_jsonl = out_dir / f"stg_leads_festival__{source_id}.raw.jsonl"
    evidence_json = out_dir / f"evidence__stg_leads_festival__{source_id}.json"

    base_fields = [
        "source_id",
        "sheet_name",
        "row_number",
        "evento",
        "cpf_hash",
        "email_hash",
        "sexo",
        "estado",
        "cidade",
        "data_criacao",
        "acoes",
        "interesses",
        "area_atuacao",
        "cpf_promotor",
        "nome_promotor",
    ]
    pii_fields = [
        "cpf_raw",
        "email_raw",
        "nome",
        "sobrenome",
        "telefone",
        "cep",
        "data_nascimento",
        "rg",
    ]
    fields = base_fields + (pii_fields if include_pii else [])

    write_csv(staging_csv, canonical_rows, fieldnames=fields)
    write_jsonl(staging_jsonl, raw_rows)

    env = EvidenceEnvelope(
        extractor="extract_xlsx_leads_festival",
        source_id=source_id,
        source_path=str(xlsx_path),
        started_at=started,
        finished_at=finished,
        status="OK",
        notes=notes,
        stats={
            "file_snapshot": {"sha256": snap.sha256, "size_bytes": snap.size_bytes, "mtime_utc": snap.mtime_utc},
            "sheet": ws.title,
            "header_row": hdr.header_row,
            "max_col": hdr.max_col,
            "rows_extracted": len(canonical_rows),
            "include_pii": include_pii,
            "outputs": {
                "csv": str(staging_csv),
                "jsonl": str(staging_jsonl),
            },
            "header": hdr.headers,
        },
    )
    write_json(evidence_json, env.to_dict())

    return {"csv": staging_csv, "jsonl": staging_jsonl, "evidence": evidence_json}


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source-id", required=True)
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out-dir", required=True)
    ap.add_argument("--sheet", default="")
    ap.add_argument("--include-pii", action="store_true")
    args = ap.parse_args(argv)

    extract_leads_xlsx(
        source_id=args.source_id,
        xlsx_path=Path(args.xlsx),
        out_dir=Path(args.out_dir),
        sheet_name=args.sheet or None,
        include_pii=bool(args.include_pii),
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
