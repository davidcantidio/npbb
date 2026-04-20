"""Rotas de conversoes de lead."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import os
import re
import tracemalloc
from datetime import date, datetime
from pathlib import Path
from collections.abc import Iterator

import httpx

from fastapi import APIRouter, Depends, File, Form, Query, UploadFile, status
from fastapi import Response as FastAPIResponse
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy.orm import aliased
from sqlalchemy.orm import load_only
from sqlalchemy import and_, case, func
from sqlalchemy import or_
from sqlalchemy.sql import text
from sqlmodel import Session, select
from lead_pipeline.constants import TIPO_EVENTO_PADRAO

from core.leads_etl.models import coerce_lead_field
from app.core.auth import get_current_user
from app.db.database import get_session, set_internal_service_db_context
from app.models.lead_batch import BatchStage, LeadBatch, PipelineStatus
from app.models.models import (
    Ativacao,
    Evento,
    Lead,
    LeadAlias,
    LeadAliasTipo,
    LeadConversao,
    LeadEvento,
    TipoEvento,
    Usuario,
    now_utc,
)
from app.modules.leads_publicidade.application.leads_import_usecases import (
    importar_leads_usecase,
    preview_import_sample_usecase,
    validar_mapeamento_usecase,
)
from app.modules.leads_publicidade.application.leads_import_etl_usecases import (
    commit_leads_with_etl,
    import_leads_with_etl,
)
from app.modules.leads_publicidade.application.etl_import.exceptions import (
    EtlImportContractError,
    EtlImportJobConflictError,
    EtlImportJobNotFoundError,
    EtlImportValidationError,
    EtlPreviewSessionConflictError,
    EtlPreviewSessionNotFoundError,
)
from app.modules.leads_publicidade.application.etl_import.job_service import (
    create_preview_job,
    execute_commit_job,
    execute_preview_job,
    queue_commit_job,
    queue_preview_reprocess,
    read_job,
    read_job_errors,
)
from app.modules.leads_publicidade.application.etl_import.extract import ETL_MAX_SCAN_ROWS_CAP
from app.modules.leads_publicidade.application.etl_import.rejection_reasons import aggregate_rejection_reason_counts
from app.modules.leads_publicidade.application.etl_import.persistence import (
    build_dedupe_key,
    find_existing_lead,
    merge_lead,
    persist_lead_batch,
)
from app.observability.prometheus_leads_import import inc_gold_reclaimed_stale, record_import_upload_rejection
from app.services.imports.file_reader import ImportFileError, inspect_upload, read_raw_file_preview
from app.services.imports.payload_storage import PayloadNotFoundError, read_batch_payload
from app.services.lead_batch_intake_service import (
    execute_lead_batch_intake,
    parse_lead_batch_intake_manifest,
)
from app.services.lead_mapping import (
    BatchMappingBlockedError,
    mapear_batch,
    mapear_batches,
    suggest_batch_column_mapping,
    suggest_column_mapping,
)
from app.services.lead_pipeline_service import (
    executar_pipeline_gold_em_thread,  # noqa: F401 - compatibilidade com harness legado de testes
    is_gold_pipeline_progress_stale,
    load_batch_without_bronze,
    load_batch_without_bronze_for_update,
    pipeline_stale_after_seconds,
    queue_pipeline_batch,
)
from app.services.evento_activation_import import (
    ACTIVATION_IMPORT_REQUIRES_AGENCY_CODE,
    get_activation_import_block_reason,
    get_activation_import_block_reason_from_agencia_id,
)
from app.schemas.lead_batch import (
    BatchColumnGroupRead,
    BatchColumnOccurrenceRead,
    ColunasBatchRequest,
    ColunasBatchResponse,
    ColunasResponse,
    ColumnSuggestionRead,
    ExecutarPipelineResponse,
    LeadBatchImportHintRead,
    LeadBatchIntakeResponse,
    LeadBatchPreviewResponse,
    LeadBatchRead,
    LeadReferenceEventoRead,
    MapearBatchRequest,
    MapearBatchResponse,
    MapearLotesBatchItemResponse,
    MapearLotesBatchRequest,
    MapearLotesBatchResponse,
)
from app.schemas.lead_conversao import LeadConversaoCreate, LeadConversaoRead
from app.schemas.lead_import import LeadImportMapping
from app.schemas.lead_import_etl import (
    ImportEtlCommitRequest,
    ImportEtlCpfColumnRequiredResponse,
    ImportEtlHeaderRequiredResponse,
    ImportEtlJobCommitRequest,
    ImportEtlJobErrorsResponse,
    ImportEtlJobQueuedResponse,
    ImportEtlJobRead,
    ImportEtlJobReprocessRequest,
    ImportEtlPreviewResponse,
    ImportEtlPreviewResponseUnion,
    ImportEtlResult,
)
from app.schemas.lead_list import LeadListItemRead, LeadListQuery, LeadListResponse
from app.schemas.landing_public import LandingSubmitRequest, LandingSubmitResponse
from app.schemas.reconhecimento import LeadRecognitionRead
from app.services.leads_export import generate_gold_export
from app.services.landing_page_submission import get_public_lead_success_message, submit_public_lead
from app.services.reconhecimento import set_lead_recognition_cookie, validar_token
from app.utils.cpf import validate_and_normalize_cpf
from app.utils.http_errors import raise_http_error
from app.utils.text_normalize import normalize_text
from app.utils.fuzzy_match import best_match

router = APIRouter(prefix="/leads", tags=["leads"])
logger = logging.getLogger(__name__)

ALLOWED_IMPORT_EXTENSIONS = {".csv", ".xlsx"}
BATCH_SIZE = 500
DEFAULT_IMPORT_MAX_BYTES = 50 * 1024 * 1024
DEFAULT_BATCH_SUMMARY_LIMIT = 200
DEFAULT_LOG_MEMORY = "0"

CEP_CACHE: dict[str, dict[str, str]] = {}
CEP_CACHE_MAX = 500
ARQUIVO_SHA256_PATTERN = re.compile(r"^[0-9a-f]{64}$")


def _should_execute_import_jobs_inline() -> bool:
    raw = os.getenv("LEAD_IMPORT_JOB_INLINE_EXECUTION", "").strip().lower()
    return os.getenv("TESTING", "").strip().lower() == "true" or raw in {"1", "true", "yes", "on"}


def _fetch_cep_data(cep: str) -> dict[str, str] | None:
    if not cep or len(cep) != 8:
        return None
    cached = CEP_CACHE.get(cep)
    if cached is not None:
        return cached or None
    try:
        resp = httpx.get(f"https://viacep.com.br/ws/{cep}/json/", timeout=2.0)
        if resp.status_code != 200:
            CEP_CACHE[cep] = {}
            return None
        data = resp.json()
        if data.get("erro"):
            CEP_CACHE[cep] = {}
            return None
        result = {
            "endereco_rua": (data.get("logradouro") or "").strip(),
            "bairro": (data.get("bairro") or "").strip(),
            "cidade": (data.get("localidade") or "").strip(),
            "estado": (data.get("uf") or "").strip(),
        }
        CEP_CACHE[cep] = result
        if len(CEP_CACHE) > CEP_CACHE_MAX:
            CEP_CACHE.clear()
        return result
    except Exception:
        return None


def _find_existing_lead(session: Session, payload: dict[str, object]) -> Lead | None:
    return find_existing_lead(session, payload)


def _dedupe_key(payload: dict[str, object]) -> str | None:
    return build_dedupe_key(payload)


def _merge_lead(existing: Lead, payload: dict[str, object]) -> None:
    merge_lead(existing, payload)


def _process_batch(
    session: Session,
    batch: list[tuple[dict[str, object], int] | None],
) -> tuple[int, int, int, bool]:
    return persist_lead_batch(session, batch)


def _get_env_int(name: str, default: int) -> int:
    raw = os.getenv(name)
    if raw is None or raw == "":
        return default
    try:
        return int(raw)
    except ValueError:
        return default


MAX_IMPORT_FILE_BYTES = _get_env_int("LEADS_IMPORT_MAX_BYTES", DEFAULT_IMPORT_MAX_BYTES)
BATCH_SUMMARY_LIMIT = _get_env_int("LEADS_IMPORT_BATCH_SUMMARY_LIMIT", DEFAULT_BATCH_SUMMARY_LIMIT)
LOG_MEMORY = os.getenv("LEADS_IMPORT_LOG_MEMORY", DEFAULT_LOG_MEMORY).strip().lower() in {"1", "true", "yes"}


def _log_memory_usage(stage: str, batch_index: int | None = None) -> None:
    if not LOG_MEMORY:
        return
    if not tracemalloc.is_tracing():
        tracemalloc.start()
    current, peak = tracemalloc.get_traced_memory()
    logger.info(
        "Lead import memory stage=%s batch=%s current_mb=%.2f peak_mb=%.2f",
        stage,
        batch_index,
        current / (1024 * 1024),
        peak / (1024 * 1024),
    )


def _get_lead_or_404(*, session: Session, lead_id: int) -> Lead:
    lead = session.get(Lead, lead_id)
    if not lead:
        raise_http_error(
            status.HTTP_404_NOT_FOUND,
            code="LEAD_NOT_FOUND",
            message="Lead nao encontrado",
        )
    return lead


def _get_public_submit_context(
    session: Session,
    *,
    payload: LandingSubmitRequest,
) -> tuple[Evento, Ativacao | None, str | None]:
    set_internal_service_db_context(session)
    ativacao: Ativacao | None = None

    if payload.ativacao_id is not None:
        ativacao = session.get(Ativacao, payload.ativacao_id)
        if ativacao is None:
            raise_http_error(
                status.HTTP_404_NOT_FOUND,
                code="ATIVACAO_NOT_FOUND",
                message="Ativacao nao encontrada",
                field="ativacao_id",
            )

        if payload.event_id is not None and payload.event_id != ativacao.evento_id:
            raise_http_error(
                status.HTTP_400_BAD_REQUEST,
                code="EVENTO_ATIVACAO_MISMATCH",
                message="event_id diverge da ativacao informada",
                field="event_id",
            )
        if not payload.cpf:
            raise_http_error(
                status.HTTP_400_BAD_REQUEST,
                code="CPF_REQUIRED",
                message="cpf obrigatorio quando ativacao_id for informado",
                field="cpf",
            )
        try:
            cpf_for_conversion = validate_and_normalize_cpf(payload.cpf)
        except ValueError:
            raise_http_error(
                status.HTTP_400_BAD_REQUEST,
                code="CPF_INVALID",
                message="CPF invalido",
                field="cpf",
            )

        evento = session.get(Evento, ativacao.evento_id)
        if evento is None:
            raise_http_error(
                status.HTTP_404_NOT_FOUND,
                code="EVENTO_NOT_FOUND",
                message="Evento nao encontrado",
                field="event_id",
            )
        return evento, ativacao, cpf_for_conversion

    if payload.event_id is None:
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="EVENT_ID_REQUIRED",
            message="event_id obrigatorio quando ativacao_id nao for informado",
            field="event_id",
        )

    evento = session.get(Evento, payload.event_id)
    if evento is None:
        raise_http_error(
            status.HTTP_404_NOT_FOUND,
            code="EVENTO_NOT_FOUND",
            message="Evento nao encontrado",
            field="event_id",
        )
    return evento, None, None


@router.post("", response_model=LandingSubmitResponse, status_code=status.HTTP_201_CREATED)
@router.post("/", response_model=LandingSubmitResponse, status_code=status.HTTP_201_CREATED, include_in_schema=False)
def criar_lead_publico(
    payload: LandingSubmitRequest,
    response: FastAPIResponse,
    session: Session = Depends(get_session),
):
    evento, ativacao, cpf_for_conversion = _get_public_submit_context(
        session,
        payload=payload,
    )
    result = submit_public_lead(
        session,
        evento=evento,
        ativacao=ativacao,
        payload=payload,
        success_message=get_public_lead_success_message(session, evento=evento),
        cpf_for_conversion=cpf_for_conversion,
    )
    if result.token_reconhecimento:
        set_lead_recognition_cookie(response, result.token_reconhecimento)
    return result


@router.get("/reconhecer", response_model=LeadRecognitionRead)
def reconhecer_lead_publico(
    token: str,
    evento_id: int,
    session: Session = Depends(get_session),
):
    set_internal_service_db_context(session)
    result = validar_token(session, token=token, evento_id=evento_id)
    return LeadRecognitionRead(
        lead_reconhecido=result is not None,
        lead_id=result.lead_id if result else None,
    )


def _format_data_csv_dia_meia_noite(d: date | None) -> str | None:
    if d is None:
        return None
    return f"{d.isoformat()} 00:00:00"


def _format_data_evento_export(
    evento_data: date | None,
    data_conversao: datetime | None,
) -> str | None:
    if evento_data is not None:
        return _format_data_csv_dia_meia_noite(evento_data)
    if data_conversao is not None:
        return _format_data_csv_dia_meia_noite(data_conversao.date())
    return None


def _ano_evento_export(evento_data: date | None, data_conversao: datetime | None) -> int | None:
    if evento_data is not None:
        return int(evento_data.year)
    if data_conversao is not None:
        return int(data_conversao.year)
    return None


def _ref_date_for_age(evento_data: date | None, data_conversao: datetime | None) -> date | None:
    if evento_data is not None:
        return evento_data
    if data_conversao is not None:
        return data_conversao.date()
    return None


def _idade_na_data_ref(data_nascimento: date | None, ref: date | None) -> int | None:
    if data_nascimento is None or ref is None:
        return None
    age = ref.year - data_nascimento.year
    if (ref.month, ref.day) < (data_nascimento.month, data_nascimento.day):
        age -= 1
    return max(0, age)


def _faixa_etaria_export_data2(idade: int | None) -> str | None:
    if idade is None:
        return None
    if idade < 18:
        return "0-17"
    if idade <= 40:
        return "18-40"
    return "40+"


def _format_local_evento_export(cidade: str | None, estado: str | None) -> str | None:
    cidade_value = (cidade or "").strip()
    estado_value = (estado or "").strip()
    if not cidade_value and not estado_value:
        return None
    if cidade_value and estado_value:
        return f"{cidade_value}-{estado_value}"
    return cidade_value or estado_value


def _lead_list_effective_event_date_expr(evento_origem):
    """Data do evento usada na listagem (conversao mais recente com evento ou evento de origem do lote)."""
    data_conv = func.coalesce(LeadConversao.data_conversao_evento, LeadConversao.created_at)
    conversao_branch = func.coalesce(Evento.data_inicio_prevista, func.date(data_conv))
    return case(
        (LeadConversao.evento_id.isnot(None), conversao_branch),
        else_=evento_origem.data_inicio_prevista,
    )


def _lead_list_filters(params: LeadListQuery, origem_evento_id_col=None, *, evento_origem=None) -> list:
    filters: list = []
    if params.data_inicio is not None or params.data_fim is not None:
        if evento_origem is None:
            raise ValueError("evento_origem is required when filtering by data_inicio/data_fim")
        eff = _lead_list_effective_event_date_expr(evento_origem)
        if params.data_inicio is not None:
            filters.append(eff >= params.data_inicio)
        if params.data_fim is not None:
            filters.append(eff <= params.data_fim)
    if params.evento_id is not None:
        if origem_evento_id_col is None:
            filters.append(LeadConversao.evento_id == params.evento_id)
        else:
            filters.append(
                or_(
                    LeadConversao.evento_id == params.evento_id,
                    origem_evento_id_col == params.evento_id,
                )
            )
    return filters


def _configure_listar_leads_statement_timeout(session: Session, params: LeadListQuery) -> None:
    if not params.long_running:
        return
    bind = session.get_bind()
    dialect_name = getattr(getattr(bind, "dialect", None), "name", "")
    if dialect_name != "postgresql":
        return
    timeout_ms = max(int(os.getenv("LEADS_EXPORT_STATEMENT_TIMEOUT_MS", "900000")), 0)
    session.exec(text(f"SET LOCAL statement_timeout = {timeout_ms}"))


@router.get("", response_model=LeadListResponse)
@router.get("/", response_model=LeadListResponse)
def listar_leads(
    params: LeadListQuery = Depends(),
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    _configure_listar_leads_statement_timeout(session, params)
    offset = (params.page - 1) * params.page_size
    evento_origem = aliased(Evento)
    tipo_evento_origem = aliased(TipoEvento)

    latest_conversao_subquery = (
        select(
            LeadConversao.lead_id.label("lead_id"),
            func.max(LeadConversao.id).label("latest_conversao_id"),
        )
        .group_by(LeadConversao.lead_id)
        .subquery()
    )

    list_filters = _lead_list_filters(params, LeadBatch.evento_id, evento_origem=evento_origem)

    base_from = (
        select(
            Lead.id,
            Lead.nome,
            Lead.sobrenome,
            Lead.email,
            Lead.cpf,
            Lead.telefone,
            Lead.evento_nome,
            Lead.cidade,
            Lead.estado,
            Lead.data_compra,
            Lead.data_criacao,
            Lead.data_nascimento,
            LeadConversao.tipo,
            LeadConversao.data_conversao_evento,
            LeadConversao.created_at,
            LeadConversao.evento_id,
            Evento.nome.label("evento_convertido_nome"),
            Evento.data_inicio_prevista.label("evento_data_inicio_prevista"),
            Evento.data_fim_prevista.label("evento_data_fim_prevista"),
            Evento.cidade.label("evento_cidade"),
            Evento.estado.label("evento_estado"),
            TipoEvento.nome.label("tipo_evento_nome"),
            LeadBatch.evento_id.label("evento_origem_id"),
            evento_origem.data_inicio_prevista.label("evento_origem_data_inicio_prevista"),
            evento_origem.data_fim_prevista.label("evento_origem_data_fim_prevista"),
            evento_origem.cidade.label("evento_origem_cidade"),
            evento_origem.estado.label("evento_origem_estado"),
            evento_origem.tipo_id.label("evento_origem_tipo_id"),
            tipo_evento_origem.nome.label("tipo_evento_origem_nome"),
            Lead.rg,
            Lead.genero,
            Lead.is_cliente_bb,
            Lead.is_cliente_estilo,
            Lead.endereco_rua,
            Lead.endereco_numero,
            Lead.complemento,
            Lead.bairro,
            Lead.cep,
            Lead.fonte_origem,
            LeadBatch.origem_lote,
            LeadBatch.plataforma_origem,
        )
        .select_from(Lead)
        .outerjoin(latest_conversao_subquery, latest_conversao_subquery.c.lead_id == Lead.id)
        .outerjoin(LeadConversao, LeadConversao.id == latest_conversao_subquery.c.latest_conversao_id)
        .outerjoin(Evento, Evento.id == LeadConversao.evento_id)
        .outerjoin(TipoEvento, TipoEvento.id == Evento.tipo_id)
        .outerjoin(LeadBatch, LeadBatch.id == Lead.batch_id)
        .outerjoin(evento_origem, evento_origem.id == LeadBatch.evento_id)
        .outerjoin(tipo_evento_origem, tipo_evento_origem.id == evento_origem.tipo_id)
    )
    if list_filters:
        base_from = base_from.where(and_(*list_filters))

    count_stmt = (
        select(func.count(Lead.id))
        .select_from(Lead)
        .outerjoin(latest_conversao_subquery, latest_conversao_subquery.c.lead_id == Lead.id)
        .outerjoin(LeadConversao, LeadConversao.id == latest_conversao_subquery.c.latest_conversao_id)
        .outerjoin(Evento, Evento.id == LeadConversao.evento_id)
        .outerjoin(LeadBatch, LeadBatch.id == Lead.batch_id)
        .outerjoin(evento_origem, evento_origem.id == LeadBatch.evento_id)
    )
    if list_filters:
        count_stmt = count_stmt.where(and_(*list_filters))
    total = int(session.exec(count_stmt).one() or 0)

    rows = session.exec(
        base_from.order_by(Lead.data_criacao.desc(), Lead.id.desc()).offset(offset).limit(params.page_size)
    ).all()

    items: list[LeadListItemRead] = []
    for (
        lead_id,
        nome,
        sobrenome,
        email,
        cpf,
        telefone,
        evento_nome,
        cidade,
        estado,
        data_compra,
        data_criacao,
        data_nascimento,
        tipo_conversao,
        data_conversao_evento,
        data_conversao_criada,
        evento_convertido_id,
        evento_convertido_nome,
        evento_data_inicio_prevista,
        evento_data_fim_prevista,
        evento_cidade,
        evento_estado,
        tipo_evento_nome,
        evento_origem_id,
        evento_origem_data_inicio_prevista,
        evento_origem_data_fim_prevista,
        evento_origem_cidade,
        evento_origem_estado,
        evento_origem_tipo_id,
        tipo_evento_origem_nome,
        rg,
        genero,
        is_cliente_bb,
        is_cliente_estilo,
        endereco_rua,
        endereco_numero,
        complemento,
        bairro,
        cep,
        fonte_origem,
        origem_lote,
        plataforma_origem,
    ) in rows:
        tipo_conversao_value = None
        if tipo_conversao is not None:
            tipo_conversao_value = (
                tipo_conversao.value if hasattr(tipo_conversao, "value") else str(tipo_conversao)
            )
        nome_completo = " ".join(
            part.strip()
            for part in [nome or "", sobrenome or ""]
            if part and part.strip()
        ) or None
        data_conv = data_conversao_evento or data_conversao_criada
        has_evento_convertido = evento_convertido_id is not None
        evento_data_export = (
            evento_data_inicio_prevista if has_evento_convertido else evento_origem_data_inicio_prevista
        )
        evento_fim_export = (
            evento_data_fim_prevista if has_evento_convertido else evento_origem_data_fim_prevista
        )
        local_evento = _format_local_evento_export(
            evento_cidade if has_evento_convertido else evento_origem_cidade,
            evento_estado if has_evento_convertido else evento_origem_estado,
        )
        tipo_evento_export = tipo_evento_nome
        if not has_evento_convertido:
            if tipo_evento_origem_nome:
                tipo_evento_export = tipo_evento_origem_nome
            elif evento_origem_id is not None and evento_origem_tipo_id is None:
                tipo_evento_export = TIPO_EVENTO_PADRAO
        ref_date = _ref_date_for_age(evento_data_export, data_conv if has_evento_convertido else None)
        idade_val = _idade_na_data_ref(data_nascimento, ref_date)
        faixa_val = _faixa_etaria_export_data2(idade_val)
        origem_export = _format_origem_lead_export(fonte_origem, origem_lote, plataforma_origem)
        items.append(
            LeadListItemRead(
                id=int(lead_id),
                nome=nome_completo,
                sobrenome=sobrenome,
                email=email,
                cpf=cpf,
                telefone=telefone,
                evento_nome=evento_nome,
                cidade=cidade,
                estado=estado,
                data_compra=data_compra,
                data_criacao=data_criacao,
                evento_convertido_id=int(evento_convertido_id) if evento_convertido_id is not None else None,
                evento_convertido_nome=evento_convertido_nome,
                tipo_conversao=tipo_conversao_value,
                data_conversao=data_conv,
                rg=rg,
                genero=genero,
                is_cliente_bb=is_cliente_bb,
                is_cliente_estilo=is_cliente_estilo,
                logradouro=endereco_rua,
                numero=endereco_numero,
                complemento=complemento,
                bairro=bairro,
                cep=cep,
                data_nascimento=data_nascimento,
                data_evento=_format_data_evento_export(
                    evento_data_export,
                    data_conv if has_evento_convertido else None,
                ),
                soma_de_ano_evento=_ano_evento_export(
                    evento_data_export,
                    data_conv if has_evento_convertido else None,
                ),
                tipo_evento=tipo_evento_export,
                local_evento=local_evento,
                faixa_etaria=faixa_val,
                soma_de_idade=idade_val,
                origem=origem_export or None,
                evento_inicio_prevista=evento_data_export,
                evento_fim_prevista=evento_fim_export,
            )
        )

    return LeadListResponse(
        page=params.page,
        page_size=params.page_size,
        total=total,
        items=items,
    )


LEADS_LIST_EXPORT_PAGE_SIZE = 100
LEADS_LIST_EXPORT_HEADERS = [
    "nome",
    "cpf",
    "data_nascimento",
    "email",
    "telefone",
    "origem",
    "evento",
    "tipo_evento",
    "local",
    "data_evento",
]


def _format_origem_lead_export(
    fonte_origem: str | None,
    origem_lote: str | None,
    plataforma_origem: str | None,
) -> str:
    """Rotulo para CSV: Proponente | Ativação (prioriza metadados do lote, depois fonte do lead)."""
    for raw in (origem_lote, fonte_origem, plataforma_origem):
        if not raw:
            continue
        s = raw.strip().lower()
        if s in ("ativacao", "ativação") or "ativacao" in s or "ativação" in s:
            return "Ativação"
        if s == "proponente" or "proponente" in s:
            return "Proponente"
    return ""


def _format_leads_list_export_date(value: date | datetime | str | None) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        head = value.split(" ", 1)[0]
        try:
            parsed = date.fromisoformat(head)
        except ValueError:
            return value
        return f"{parsed.month}/{parsed.day}/{parsed.year}"
    if isinstance(value, datetime):
        value = value.date()
    return f"{value.month}/{value.day}/{value.year}"


def _resolve_leads_list_export_event(item: LeadListItemRead) -> str:
    return item.evento_convertido_nome or item.evento_nome or ""


def _build_leads_list_export_cells(item: LeadListItemRead) -> list[str]:
    return [
        item.nome or "",
        item.cpf or "",
        _format_leads_list_export_date(item.data_nascimento),
        item.email or "",
        item.telefone or "",
        item.origem or "",
        _resolve_leads_list_export_event(item),
        item.tipo_evento or "",
        item.local_evento or "",
        _format_leads_list_export_date(item.data_evento),
    ]


def _iter_leads_list_export_csv(
    session: Session,
    current_user: Usuario,
    first_page: LeadListResponse,
    *,
    data_inicio: date | None = None,
    data_fim: date | None = None,
    evento_id: int | None = None,
):
    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer, delimiter=",", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")

    csv_buffer.write("\ufeff")
    writer.writerow(LEADS_LIST_EXPORT_HEADERS)
    for item in first_page.items:
        writer.writerow(_build_leads_list_export_cells(item))
    yield csv_buffer.getvalue().encode("utf-8")
    csv_buffer.seek(0)
    csv_buffer.truncate(0)

    page = 2
    while True:
        response = FastAPIResponse()
        result = listar_leads(
            LeadListQuery(
                page=page,
                page_size=LEADS_LIST_EXPORT_PAGE_SIZE,
                long_running=True,
                data_inicio=data_inicio,
                data_fim=data_fim,
                evento_id=evento_id,
            ),
            session=session,
            current_user=current_user,
        )

        for item in result.items:
            writer.writerow(_build_leads_list_export_cells(item))

        chunk = csv_buffer.getvalue()
        if chunk:
            yield chunk.encode("utf-8")
        csv_buffer.seek(0)
        csv_buffer.truncate(0)

        if len(result.items) < LEADS_LIST_EXPORT_PAGE_SIZE:
            break
        page += 1


@router.get("/export/csv")
@router.get("/export/csv/")
def exportar_leads_csv(
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
    data_inicio: date | None = Query(None),
    data_fim: date | None = Query(None),
    evento_id: int | None = Query(None, ge=1),
):
    first_page = listar_leads(
        LeadListQuery(
            page=1,
            page_size=LEADS_LIST_EXPORT_PAGE_SIZE,
            long_running=True,
            data_inicio=data_inicio,
            data_fim=data_fim,
            evento_id=evento_id,
        ),
        session=session,
        current_user=current_user,
    )
    if first_page.total == 0:
        return FastAPIResponse(status_code=204)

    filename = f"leads-{date.today().isoformat()}.csv"
    return StreamingResponse(
        _iter_leads_list_export_csv(
            session,
            current_user,
            first_page,
            data_inicio=data_inicio,
            data_fim=data_fim,
            evento_id=evento_id,
        ),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{lead_id}/conversoes", response_model=LeadConversaoRead, status_code=status.HTTP_201_CREATED)
@router.post("/{lead_id}/conversoes/", response_model=LeadConversaoRead, status_code=status.HTTP_201_CREATED)
def criar_conversao(
    lead_id: int,
    payload: LeadConversaoCreate,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    lead = _get_lead_or_404(session=session, lead_id=lead_id)

    conversao = LeadConversao(
        lead_id=lead.id,
        tipo=payload.tipo,
        acao_nome=payload.acao_nome,
        fonte_origem=payload.fonte_origem,
        evento_id=payload.evento_id,
        data_conversao_evento=payload.data_conversao_evento,
        created_at=now_utc(),
    )
    session.add(conversao)
    session.commit()
    session.refresh(conversao)
    return LeadConversaoRead.model_validate(conversao, from_attributes=True)


@router.get("/{lead_id}/conversoes", response_model=list[LeadConversaoRead])
@router.get("/{lead_id}/conversoes/", response_model=list[LeadConversaoRead])
def listar_conversoes(
    lead_id: int,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    _get_lead_or_404(session=session, lead_id=lead_id)
    conversoes = session.exec(
        select(LeadConversao).where(LeadConversao.lead_id == lead_id).order_by(LeadConversao.created_at.desc())
    ).all()
    return [LeadConversaoRead.model_validate(item, from_attributes=True) for item in conversoes]


@router.post("/import/upload")
@router.post("/import/upload/")
def validar_upload_import(
    file: UploadFile = File(...),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    try:
        filename, _ext, size = inspect_upload(file, max_bytes=MAX_IMPORT_FILE_BYTES)
    except ImportFileError as err:
        record_import_upload_rejection(err, filename_hint=file.filename)
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code=err.code,
            message=err.message,
            field=err.field,
            extra=err.extra,
        )
    return {"filename": filename, "size_bytes": size}


@router.get("/referencias/eventos", response_model=list[LeadReferenceEventoRead])
@router.get("/referencias/eventos/", response_model=list[LeadReferenceEventoRead])
def listar_referencia_eventos(
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    count_stmt = (
        select(LeadEvento.evento_id, func.count(func.distinct(LeadEvento.lead_id)))
        .group_by(LeadEvento.evento_id)
    )
    lead_counts: dict[int, int] = {}
    for evento_id, cnt in session.exec(count_stmt).all():
        if evento_id is not None:
            lead_counts[int(evento_id)] = int(cnt or 0)

    stmt = (
        select(Evento.id, Evento.nome, Evento.data_inicio_prevista, Evento.agencia_id)
        .order_by(Evento.data_inicio_prevista.desc().nulls_last(), Evento.nome)
    )
    eventos = session.exec(stmt).all()
    return [
        {
            "id": int(eid),
            "nome": nome,
            "data_inicio_prevista": str(data_inicio) if data_inicio else None,
            "agencia_id": int(agencia_id) if agencia_id is not None else None,
            "supports_activation_import": bool(agencia_id is not None),
            "activation_import_block_reason": get_activation_import_block_reason_from_agencia_id(agencia_id),
            "leads_count": lead_counts.get(int(eid), 0),
        }
        for eid, nome, data_inicio, agencia_id in eventos
        if eid is not None and nome
    ]


@router.get("/referencias/cidades")
@router.get("/referencias/cidades/")
def listar_referencia_cidades(
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    rows = session.exec(select(Evento.cidade).where(Evento.cidade.is_not(None)).distinct()).all()
    return sorted({str(c).strip() for c in rows if c})


@router.get("/referencias/estados")
@router.get("/referencias/estados/")
def listar_referencia_estados(
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    rows = session.exec(select(Evento.estado).where(Evento.estado.is_not(None)).distinct()).all()
    return sorted({str(c).strip() for c in rows if c})


@router.get("/referencias/generos")
@router.get("/referencias/generos/")
def listar_referencia_generos(
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    rows = session.exec(select(Lead.genero).where(Lead.genero.is_not(None)).distinct()).all()
    return sorted({str(c).strip() for c in rows if c})


REFERENCE_THRESHOLDS = {
    LeadAliasTipo.EVENTO: 0.82,
    LeadAliasTipo.CIDADE: 0.85,
    LeadAliasTipo.ESTADO: 0.9,
    LeadAliasTipo.GENERO: 0.85,
}


@router.get("/referencias/suggest")
@router.get("/referencias/suggest/")
def sugerir_referencia(
    tipo: LeadAliasTipo,
    valor: str,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    threshold = REFERENCE_THRESHOLDS.get(tipo, 0.85)
    if tipo == LeadAliasTipo.EVENTO:
        candidates = [row[0] for row in session.exec(select(Evento.nome)).all() if row and row[0]]
    elif tipo == LeadAliasTipo.CIDADE:
        candidates = [row[0] for row in session.exec(select(Evento.cidade)).all() if row and row[0]]
    elif tipo == LeadAliasTipo.ESTADO:
        candidates = [row[0] for row in session.exec(select(Evento.estado)).all() if row and row[0]]
    else:
        candidates = [row[0] for row in session.exec(select(Lead.genero)).all() if row and row[0]]

    suggestion, score = best_match(valor, candidates, threshold=threshold)
    return {"suggested": suggestion, "score": score, "threshold": threshold}


def _normalize_alias_value(value: str) -> str:
    return normalize_text(value)


def _map_etl_import_error(exc: Exception) -> None:
    if isinstance(exc, EtlImportJobNotFoundError):
        raise_http_error(
            status.HTTP_404_NOT_FOUND,
            code="ETL_JOB_NOT_FOUND",
            message="job_id invalido ou expirado",
            field="job_id",
        )
    if isinstance(exc, EtlImportJobConflictError):
        raise_http_error(
            status.HTTP_409_CONFLICT,
            code="ETL_JOB_CONFLICT",
            message=str(exc),
            field="job_id",
        )
    if isinstance(exc, EtlPreviewSessionNotFoundError):
        raise_http_error(
            status.HTTP_404_NOT_FOUND,
            code="ETL_SESSION_NOT_FOUND",
            message="session_token invalido ou expirado",
            field="session_token",
        )
    if isinstance(exc, EtlPreviewSessionConflictError):
        raise_http_error(
            status.HTTP_409_CONFLICT,
            code="ETL_SESSION_CONFLICT",
            message=str(exc),
            field="session_token",
        )
    if isinstance(exc, EtlImportValidationError):
        raise_http_error(
            status.HTTP_409_CONFLICT,
            code="ETL_COMMIT_BLOCKED",
            message=str(exc),
        )
    if isinstance(exc, EtlImportContractError):
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="ETL_INVALID_INPUT",
            message=str(exc),
        )
    if isinstance(exc, ImportFileError):
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code=exc.code,
            message=exc.message,
            field=exc.field,
            extra=exc.extra,
        )
    if isinstance(exc, ValueError):
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="ETL_INVALID_INPUT",
            message=str(exc),
        )
    raise exc


def _serialize_etl_preview_response(
    result,
) -> ImportEtlPreviewResponse | ImportEtlHeaderRequiredResponse | ImportEtlCpfColumnRequiredResponse:
    if getattr(result, "status", None) == "header_required":
        return ImportEtlHeaderRequiredResponse.model_validate(
            {
                "status": result.status,
                "message": result.message,
                "max_row": result.max_row,
                "scanned_rows": result.scanned_rows,
                "required_fields": list(result.required_fields),
                "available_sheets": list(getattr(result, "available_sheets", ()) or ()),
                "active_sheet": getattr(result, "active_sheet", None),
            }
        )
    if getattr(result, "status", None) == "cpf_column_required":
        return ImportEtlCpfColumnRequiredResponse.model_validate(
            {
                "status": result.status,
                "message": result.message,
                "header_row": result.header_row,
                "columns": [column.__dict__ for column in result.columns],
                "required_fields": list(result.required_fields),
                "available_sheets": list(getattr(result, "available_sheets", ()) or ()),
                "active_sheet": getattr(result, "active_sheet", None),
            }
        )
    return ImportEtlPreviewResponse.model_validate(
        {
            "status": "previewed",
            "session_token": result.session_token,
            "total_rows": result.total_rows,
            "valid_rows": result.valid_rows,
            "invalid_rows": result.invalid_rows,
            "dq_report": [item.__dict__ for item in result.dq_report],
            "rejection_reason_counts": aggregate_rejection_reason_counts(result.rejected_rows),
            "sheet_name": getattr(result, "sheet_name", None),
            "available_sheets": list(getattr(result, "available_sheets", ()) or ()),
        }
    )


def _serialize_etl_commit_response(result) -> ImportEtlResult:
    return ImportEtlResult.model_validate(
        {
            "session_token": result.session_token,
            "total_rows": result.total_rows,
            "valid_rows": result.valid_rows,
            "invalid_rows": result.invalid_rows,
            "created": result.created,
            "updated": result.updated,
            "skipped": result.skipped,
            "errors": result.errors,
            "strict": result.strict,
            "status": result.status,
            "duplicate_rows": result.duplicate_rows,
            "staged_rows": result.staged_rows,
            "ingestion_strategy": result.ingestion_strategy,
            "merge_strategy": result.merge_strategy,
            "dq_report": [item.__dict__ for item in result.dq_report],
            "persistence_failures": [
                {"row_number": row_number, "reason": reason}
                for row_number, reason in result.persistence_failures
            ],
        }
    )


@router.get("/aliases")
@router.get("/aliases/")
def buscar_alias(
    tipo: LeadAliasTipo,
    valor_origem: str,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    normalized = _normalize_alias_value(valor_origem)
    alias = session.exec(
        select(LeadAlias).where(
            LeadAlias.tipo == tipo,
            LeadAlias.valor_normalizado == normalized,
        )
    ).first()
    if not alias:
        return None
    return {
        "id": alias.id,
        "tipo": alias.tipo,
        "valor_origem": alias.valor_origem,
        "valor_normalizado": alias.valor_normalizado,
        "canonical_value": alias.canonical_value,
        "evento_id": alias.evento_id,
    }


@router.post("/aliases", status_code=status.HTTP_201_CREATED)
@router.post("/aliases/", status_code=status.HTTP_201_CREATED)
def criar_alias(
    payload: dict,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    try:
        tipo = LeadAliasTipo(payload.get("tipo"))
    except Exception:
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="INVALID_ALIAS_TYPE",
            message="Tipo de alias invalido",
            field="tipo",
        )
    valor_origem = (payload.get("valor_origem") or "").strip()
    if not valor_origem:
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="INVALID_ALIAS_VALUE",
            message="valor_origem obrigatorio",
            field="valor_origem",
        )
    canonical_value = (payload.get("canonical_value") or "").strip() or None
    evento_id = payload.get("evento_id")
    normalized = _normalize_alias_value(valor_origem)

    existing = session.exec(
        select(LeadAlias).where(
            LeadAlias.tipo == tipo,
            LeadAlias.valor_normalizado == normalized,
        )
    ).first()
    if existing:
        updated = False
        if canonical_value and existing.canonical_value != canonical_value:
            existing.canonical_value = canonical_value
            updated = True
        if evento_id and existing.evento_id != evento_id:
            existing.evento_id = evento_id
            updated = True
        if updated:
            session.add(existing)
            session.commit()
            session.refresh(existing)
        return {
            "id": existing.id,
            "tipo": existing.tipo,
            "valor_origem": existing.valor_origem,
            "valor_normalizado": existing.valor_normalizado,
            "canonical_value": existing.canonical_value,
            "evento_id": existing.evento_id,
        }

    alias = LeadAlias(
        tipo=tipo,
        valor_origem=valor_origem,
        valor_normalizado=normalized,
        canonical_value=canonical_value,
        evento_id=evento_id,
        created_at=now_utc(),
    )
    session.add(alias)
    session.commit()
    session.refresh(alias)
    return {
        "id": alias.id,
        "tipo": alias.tipo,
        "valor_origem": alias.valor_origem,
        "valor_normalizado": alias.valor_normalizado,
        "canonical_value": alias.canonical_value,
        "evento_id": alias.evento_id,
    }


def _detect_csv_delimiter(sample_text: str) -> str:
    first_line = sample_text.splitlines()[0] if sample_text else ""
    comma_count = first_line.count(",")
    semi_count = first_line.count(";")
    return ";" if semi_count > comma_count else ","


def _score_row_tabular(row: list[str]) -> int:
    non_empty = [cell for cell in row if cell.strip()]
    if not row:
        return 0
    fill_ratio = len(non_empty) / max(len(row), 1)
    if len(non_empty) >= 2 and fill_ratio >= 0.5:
        return int(fill_ratio * 100)
    return 0


def _detect_data_start_index(rows: list[list[str]]) -> int:
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows[:50]):
        score = _score_row_tabular(row)
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx


def _read_csv_sample(file: UploadFile, max_rows: int) -> dict:
    file.file.seek(0)
    sample_bytes = file.file.read(4096)
    try:
        sample_text = sample_bytes.decode("utf-8-sig", errors="ignore")
    except Exception:
        sample_text = ""
    delimiter = _detect_csv_delimiter(sample_text)

    file.file.seek(0)
    text_stream = io.TextIOWrapper(file.file, encoding="utf-8-sig", errors="ignore", newline="")
    reader = csv.reader(text_stream, delimiter=delimiter)
    rows: list[list[str]] = []
    try:
        for row in reader:
            if row is None:
                continue
            rows.append([cell.strip() for cell in row])
            if len(rows) >= max_rows + 1:
                break
    finally:
        text_stream.detach()

    start_index = _detect_data_start_index(rows)
    headers = rows[start_index] if rows else []
    data_rows = rows[start_index + 1 :] if len(rows) > start_index + 1 else []
    return {
        "headers": headers,
        "rows": data_rows,
        "delimiter": delimiter,
        "start_index": start_index,
    }


def _read_xlsx_sample(file: UploadFile, max_rows: int) -> dict:
    file.file.seek(0)
    wb = load_workbook(file.file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows: list[list[str]] = []
    for row in ws.iter_rows(max_row=max_rows + 1, values_only=True):
        rows.append([("" if v is None else str(v)).strip() for v in row])
        if len(rows) >= max_rows + 1:
            break
    start_index = _detect_data_start_index(rows)
    headers = rows[start_index] if rows else []
    data_rows = rows[start_index + 1 :] if len(rows) > start_index + 1 else []
    return {
        "headers": headers,
        "rows": data_rows,
        "delimiter": None,
        "start_index": start_index,
    }


def _read_xlsx_rows(file: UploadFile) -> list[list[str]]:
    file.file.seek(0)
    wb = load_workbook(file.file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([("" if v is None else str(v)).strip() for v in row])
    return rows


def _iter_csv_data_rows(file: UploadFile, delimiter: str, start_index: int) -> Iterator[list[str]]:
    file.file.seek(0)
    text_stream = io.TextIOWrapper(file.file, encoding="utf-8-sig", errors="ignore", newline="")
    reader = csv.reader(text_stream, delimiter=delimiter)
    try:
        for idx, row in enumerate(reader):
            if idx <= start_index:
                continue
            yield [cell.strip() for cell in row]
    finally:
        text_stream.detach()


def _iter_xlsx_data_rows(file: UploadFile, start_index: int) -> Iterator[list[str]]:
    file.file.seek(0)
    wb = load_workbook(file.file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx <= start_index:
            continue
        yield [("" if v is None else str(v)).strip() for v in row]


def _ensure_mapping_has_essential(mappings: list[LeadImportMapping]) -> None:
    mapped_fields = {m.campo for m in mappings if m.campo}
    if "cpf" not in mapped_fields:
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="MAPPING_MISSING_ESSENTIAL",
            message="O mapeamento deve incluir a coluna CPF.",
        )


def _column_samples(rows: list[list[str]], col_index: int, max_samples: int = 5) -> list[str]:
    samples: list[str] = []
    for row in rows:
        if col_index >= len(row):
            continue
        value = row[col_index].strip()
        if not value:
            continue
        samples.append(value)
        if len(samples) >= max_samples:
            break
    return samples


def _score_email(values: list[str]) -> float:
    if not values:
        return 0.0
    pattern = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
    hits = sum(1 for v in values if pattern.match(v.lower()))
    return hits / len(values)


def _score_cpf(values: list[str]) -> float:
    if not values:
        return 0.0
    pattern = re.compile(r"^\d{11}$")
    hits = 0
    for v in values:
        digits = "".join(ch for ch in v if ch.isdigit())
        if pattern.match(digits):
            hits += 1
    return hits / len(values)


def _score_phone_br(values: list[str]) -> float:
    if not values:
        return 0.0
    # DDDs validos (Brasil)
    valid_ddds = {
        "11", "12", "13", "14", "15", "16", "17", "18", "19",
        "21", "22", "24", "27", "28",
        "31", "32", "33", "34", "35", "37", "38",
        "41", "42", "43", "44", "45", "46",
        "47", "48", "49",
        "51", "53", "54", "55",
        "61", "62", "63", "64", "65", "66", "67", "68", "69",
        "71", "73", "74", "75", "77",
        "79",
        "81", "82", "83", "84", "85", "86", "87", "88", "89",
        "91", "92", "93", "94", "95", "96", "97", "98", "99",
    }
    hits = 0
    for v in values:
        digits = "".join(ch for ch in v if ch.isdigit())
        if digits.startswith("55"):
            digits = digits[2:]
        if len(digits) != 11:
            continue
        ddd = digits[:2]
        numero = digits[2:]
        if ddd not in valid_ddds:
            continue
        if not numero.startswith("9"):
            continue
        hits += 1
    return hits / len(values)


def _score_cep(values: list[str]) -> float:
    if not values:
        return 0.0
    hits = 0
    for v in values:
        digits = "".join(ch for ch in v if ch.isdigit())
        if len(digits) == 8:
            hits += 1
    return hits / len(values)


def _score_uf(values: list[str]) -> float:
    if not values:
        return 0.0
    ufs = {
        "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA",
        "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN",
        "RS", "RO", "RR", "SC", "SP", "SE", "TO",
    }
    hits = sum(1 for v in values if v.strip().upper() in ufs)
    return hits / len(values)


def _score_date(values: list[str]) -> float:
    if not values:
        return 0.0
    patterns = [
        re.compile(r"^\d{2}/\d{2}/\d{4}$"),
        re.compile(r"^\d{4}-\d{2}-\d{2}$"),
    ]
    hits = sum(1 for v in values if any(p.match(v.strip()) for p in patterns))
    return hits / len(values)


def _score_number(values: list[str]) -> float:
    if not values:
        return 0.0
    hits = 0
    for v in values:
        text = v.strip().replace(".", "").replace(",", ".")
        try:
            float(text)
            hits += 1
        except ValueError:
            continue
    return hits / len(values)


def _score_datetime(values: list[str]) -> float:
    if not values:
        return 0.0
    patterns = [
        re.compile(r"^\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}(:\d{2})?$"),
        re.compile(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}(:\d{2})?$"),
    ]
    hits = sum(1 for v in values if any(p.match(v.strip()) for p in patterns))
    return hits / len(values)


def _score_address(values: list[str]) -> float:
    if not values:
        return 0.0
    keywords = ("rua", "avenida", "av", "travessa", "alameda", "rodovia", "praca", "praça")
    hits = 0
    for v in values:
        text = v.strip().lower()
        if any(k in text for k in keywords):
            hits += 1
    return hits / len(values)


def _score_genero(values: list[str]) -> float:
    if not values:
        return 0.0
    vocab = {
        "masculino",
        "feminino",
        "homem",
        "mulher",
        "m",
        "f",
    }
    hits = 0
    for v in values:
        text = v.strip().lower()
        if text in vocab:
            hits += 1
    return hits / len(values)


def _apply_sample_penalty(score: float, sample_size: int) -> float:
    if sample_size <= 0:
        return 0.0
    if sample_size < 3:
        return score * (sample_size / 3)
    return score


def _infer_column_mapping(header: str, samples: list[str]) -> LeadImportMapping:
    header_lc = header.lower().strip()
    sample_size = len(samples)
    thresholds = {
        "email": 0.8,
        "cpf": 0.8,
        "telefone": 0.85,
        "cep": 0.85,
        "uf": 0.85,
        "datetime": 0.85,
        "data": 0.75,
        "numero": 0.85,
        "endereco": 0.7,
        "genero": 0.7,
    }
    scores = {
        "email": _apply_sample_penalty(_score_email(samples), sample_size),
        "cpf": _apply_sample_penalty(_score_cpf(samples), sample_size),
        "telefone": _apply_sample_penalty(_score_phone_br(samples), sample_size),
        "cep": _apply_sample_penalty(_score_cep(samples), sample_size),
        "uf": _apply_sample_penalty(_score_uf(samples), sample_size),
        "data": _apply_sample_penalty(_score_date(samples), sample_size),
        "datetime": _apply_sample_penalty(_score_datetime(samples), sample_size),
        "numero": _apply_sample_penalty(_score_number(samples), sample_size),
        "endereco": _apply_sample_penalty(_score_address(samples), sample_size),
        "genero": _apply_sample_penalty(_score_genero(samples), sample_size),
    }

    campo = None
    confianca = None
    allow_auto = sample_size >= 3

    if allow_auto and scores["email"] >= thresholds["email"]:
        campo = "email"
        confianca = scores["email"]
    elif allow_auto and scores["cpf"] >= thresholds["cpf"]:
        campo = "cpf"
        confianca = scores["cpf"]
    elif allow_auto and scores["telefone"] >= thresholds["telefone"]:
        campo = "telefone"
        confianca = scores["telefone"]
    elif allow_auto and scores["cep"] >= thresholds["cep"]:
        campo = "cep"
        confianca = scores["cep"]
    elif allow_auto and scores["uf"] >= thresholds["uf"]:
        campo = "estado"
        confianca = scores["uf"]
    elif allow_auto and scores["datetime"] >= thresholds["datetime"]:
        campo = "data_compra"
        confianca = scores["datetime"]
    elif "email" in header_lc:
        campo = "email"
        confianca = max(scores["email"], 0.6)
    elif "cpf" in header_lc:
        campo = "cpf"
        confianca = max(scores["cpf"], 0.6)
    elif "telefone" in header_lc or "fone" in header_lc or "cel" in header_lc:
        campo = "telefone"
        confianca = max(scores["telefone"], 0.6)
    elif "cep" in header_lc:
        campo = "cep"
        confianca = max(scores["cep"], 0.6)
    elif "uf" in header_lc or "estado" in header_lc:
        campo = "estado"
        confianca = max(scores["uf"], 0.6)
    elif "compra" in header_lc and "hora" in header_lc:
        campo = "data_compra"
        confianca = max(scores["datetime"], 0.6)
    elif "endereco" in header_lc or "rua" in header_lc or header_lc.startswith("av"):
        campo = "endereco_rua"
        confianca = max(scores["endereco"], 0.55)
    elif "genero" in header_lc or "gênero" in header_lc:
        campo = "genero"
        confianca = max(scores["genero"], 0.55)
    elif "data" in header_lc or "nasc" in header_lc:
        campo = "data_nascimento"
        confianca = max(scores["data"], 0.5)
    elif allow_auto and scores["data"] >= thresholds["data"]:
        campo = "data_nascimento"
        confianca = scores["data"]
    elif allow_auto and scores["numero"] >= thresholds["numero"]:
        campo = "ingresso_qtd"
        confianca = scores["numero"]
    elif allow_auto and scores["endereco"] >= thresholds["endereco"]:
        campo = "endereco_rua"
        confianca = scores["endereco"]
    elif allow_auto and scores["genero"] >= thresholds["genero"]:
        campo = "genero"
        confianca = scores["genero"]

    if confianca is not None:
        confianca = min(confianca, 0.9)
    return LeadImportMapping(coluna=header or "", campo=campo, confianca=confianca)


 


@router.post("/import/sample")
@router.post("/import/sample/")
def preview_import_sample(
    file: UploadFile = File(...),
    sample_rows: int = 10,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    return preview_import_sample_usecase(
        file=file,
        sample_rows=sample_rows,
        session=session,
        current_user=current_user,
        allowed_import_extensions=ALLOWED_IMPORT_EXTENSIONS,
        max_import_file_bytes=MAX_IMPORT_FILE_BYTES,
        read_xlsx_sample=_read_xlsx_sample,
        read_csv_sample=_read_csv_sample,
        column_samples=_column_samples,
        infer_column_mapping=_infer_column_mapping,
        normalize_alias_value=_normalize_alias_value,
        raise_http_error=raise_http_error,
    )


@router.post("/import/validate")
@router.post("/import/validate/")
def validar_mapeamento(
    mappings: list[LeadImportMapping],
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    return validar_mapeamento_usecase(
        mappings=mappings,
        ensure_mapping_has_essential=_ensure_mapping_has_essential,
    )


@router.post("/import")
@router.post("/import/")
def importar_leads(
    file: UploadFile = File(...),
    mappings_json: str = Form(...),
    fonte_origem: str | None = Form(None),
    enriquecer_cep: bool = Form(False),
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    return importar_leads_usecase(
        file=file,
        mappings_json=mappings_json,
        fonte_origem=fonte_origem,
        enriquecer_cep=enriquecer_cep,
        session=session,
        current_user=current_user,
        allowed_import_extensions=ALLOWED_IMPORT_EXTENSIONS,
        read_xlsx_sample=_read_xlsx_sample,
        read_csv_sample=_read_csv_sample,
        iter_xlsx_data_rows=_iter_xlsx_data_rows,
        iter_csv_data_rows=_iter_csv_data_rows,
        ensure_mapping_has_essential=_ensure_mapping_has_essential,
        dedupe_key=_dedupe_key,
        process_batch=_process_batch,
        fetch_cep_data=_fetch_cep_data,
        coerce_field=coerce_lead_field,
        log_memory_usage=_log_memory_usage,
        batch_size=BATCH_SIZE,
        batch_summary_limit=BATCH_SUMMARY_LIMIT,
        raise_http_error=raise_http_error,
    )


@router.post("/import/preview")
@router.post("/import/preview/")
def preview_import(
    file: UploadFile = File(...),
    sample_rows: int = 10,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    return preview_import_sample(
        file=file,
        sample_rows=sample_rows,
        session=session,
        current_user=current_user,
    )


@router.post("/import/etl/jobs", response_model=ImportEtlJobQueuedResponse, status_code=status.HTTP_202_ACCEPTED)
@router.post("/import/etl/jobs/", response_model=ImportEtlJobQueuedResponse, status_code=status.HTTP_202_ACCEPTED)
def criar_job_import_etl(
    file: UploadFile = File(...),
    evento_id: int = Form(...),
    strict: bool = Form(False),
    header_row: int | None = Form(None),
    field_aliases_json: str | None = Form(None),
    sheet_name: str | None = Form(None),
    max_scan_rows: int | None = Form(None),
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    if max_scan_rows is not None and (max_scan_rows < 1 or max_scan_rows > ETL_MAX_SCAN_ROWS_CAP):
        raise_http_error(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            code="ETL_INVALID_INPUT",
            message=f"max_scan_rows deve estar entre 1 e {ETL_MAX_SCAN_ROWS_CAP}.",
            field="max_scan_rows",
        )
    try:
        field_aliases = json.loads(field_aliases_json) if field_aliases_json and field_aliases_json.strip() else {}
        if field_aliases and not isinstance(field_aliases, dict):
            raise EtlImportContractError("field_aliases_json deve ser um objeto JSON.")
        job = create_preview_job(
            session=session,
            requested_by=int(current_user.id),
            file=file,
            evento_id=evento_id,
            strict=strict,
            header_row=header_row,
            field_aliases=field_aliases,
            sheet_name=sheet_name,
            max_scan_rows=max_scan_rows,
        )
        if _should_execute_import_jobs_inline():
            execute_preview_job(job.job_id)
    except Exception as exc:
        _map_etl_import_error(exc)
    return ImportEtlJobQueuedResponse(job_id=job.job_id, status="queued")


@router.get("/import/etl/jobs/{job_id}", response_model=ImportEtlJobRead)
@router.get("/import/etl/jobs/{job_id}/", response_model=ImportEtlJobRead)
def obter_job_import_etl(
    job_id: str,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    try:
        return read_job(session, job_id, requested_by=int(current_user.id))
    except Exception as exc:
        _map_etl_import_error(exc)


@router.get("/import/etl/jobs/{job_id}/errors", response_model=ImportEtlJobErrorsResponse)
@router.get("/import/etl/jobs/{job_id}/errors/", response_model=ImportEtlJobErrorsResponse)
def obter_erros_job_import_etl(
    job_id: str,
    phase: str = Query("all"),
    limit: int = Query(200, ge=1, le=500),
    offset: int = Query(0, ge=0),
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    try:
        return read_job_errors(
            session,
            job_id=job_id,
            requested_by=int(current_user.id),
            phase=phase,
            limit=limit,
            offset=offset,
        )
    except Exception as exc:
        _map_etl_import_error(exc)


@router.post(
    "/import/etl/jobs/{job_id}/reprocess-preview",
    response_model=ImportEtlJobQueuedResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
@router.post(
    "/import/etl/jobs/{job_id}/reprocess-preview/",
    response_model=ImportEtlJobQueuedResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
def reprocessar_job_preview_import_etl(
    job_id: str,
    payload: ImportEtlJobReprocessRequest,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    if payload.max_scan_rows is not None and (
        payload.max_scan_rows < 1 or payload.max_scan_rows > ETL_MAX_SCAN_ROWS_CAP
    ):
        raise_http_error(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            code="ETL_INVALID_INPUT",
            message=f"max_scan_rows deve estar entre 1 e {ETL_MAX_SCAN_ROWS_CAP}.",
            field="max_scan_rows",
        )
    try:
        job = queue_preview_reprocess(
            session=session,
            job_id=job_id,
            requested_by=int(current_user.id),
            header_row=payload.header_row,
            field_aliases={key: value.model_dump(exclude_none=True) for key, value in payload.field_aliases.items()},
            sheet_name=payload.sheet_name,
            max_scan_rows=payload.max_scan_rows,
        )
        if _should_execute_import_jobs_inline():
            execute_preview_job(job.job_id)
    except Exception as exc:
        _map_etl_import_error(exc)
    return ImportEtlJobQueuedResponse(job_id=job.job_id, status="queued")


@router.post(
    "/import/etl/jobs/{job_id}/commit",
    response_model=ImportEtlJobQueuedResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
@router.post(
    "/import/etl/jobs/{job_id}/commit/",
    response_model=ImportEtlJobQueuedResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
def confirmar_job_import_etl(
    job_id: str,
    payload: ImportEtlJobCommitRequest,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    try:
        job = queue_commit_job(
            session=session,
            job_id=job_id,
            requested_by=int(current_user.id),
            force_warnings=payload.force_warnings,
        )
        if _should_execute_import_jobs_inline():
            execute_commit_job(job.job_id)
    except Exception as exc:
        _map_etl_import_error(exc)
    return ImportEtlJobQueuedResponse(job_id=job.job_id, status="commit_queued")


@router.post("/import/etl/preview", response_model=ImportEtlPreviewResponseUnion)
@router.post("/import/etl/preview/", response_model=ImportEtlPreviewResponseUnion)
async def preview_import_etl(
    file: UploadFile = File(...),
    evento_id: int = Form(...),
    strict: bool = Form(False),
    header_row: int | None = Form(None),
    field_aliases_json: str | None = Form(None),
    sheet_name: str | None = Form(None),
    max_scan_rows: int | None = Form(None),
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    if max_scan_rows is not None and (max_scan_rows < 1 or max_scan_rows > ETL_MAX_SCAN_ROWS_CAP):
        raise_http_error(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            code="ETL_INVALID_INPUT",
            message=f"max_scan_rows deve estar entre 1 e {ETL_MAX_SCAN_ROWS_CAP}.",
            field="max_scan_rows",
        )
    try:
        result = await import_leads_with_etl(
            file=file,
            evento_id=evento_id,
            db=session,
            strict=strict,
            requested_by=int(current_user.id),
            header_row=header_row,
            field_aliases_json=field_aliases_json,
            sheet_name=sheet_name,
            max_scan_rows=max_scan_rows,
        )
    except Exception as exc:
        _map_etl_import_error(exc)
    return _serialize_etl_preview_response(result)


@router.post("/import/etl/commit", response_model=ImportEtlResult)
@router.post("/import/etl/commit/", response_model=ImportEtlResult)
async def commit_import_etl(
    payload: ImportEtlCommitRequest,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    try:
        result = await commit_leads_with_etl(
            session_token=payload.session_token,
            evento_id=payload.evento_id,
            db=session,
            force_warnings=payload.force_warnings,
        )
    except Exception as exc:
        _map_etl_import_error(exc)
    return _serialize_etl_commit_response(result)


# ---------------------------------------------------------------------------
# Lead Batch (Bronze) endpoints
# ---------------------------------------------------------------------------


@router.get("/export/gold")
@router.get("/export/gold/")
def exportar_leads_gold(
    evento_id: int | None = None,
    formato: str = "xlsx",
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    """Exports Gold-stage leads as .xlsx or .csv.

    Returns HTTP 204 when no leads are found for the selected filters.
    """
    _ = current_user
    resultado = generate_gold_export(db=session, evento_id=evento_id, formato=formato)

    if resultado is None:
        return FastAPIResponse(status_code=204)

    file_bytes, filename = resultado

    if formato == "csv":
        media_type = "text/csv; charset=utf-8"
    else:
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/batches/intake", response_model=LeadBatchIntakeResponse, status_code=status.HTTP_201_CREATED)
@router.post("/batches/intake/", response_model=LeadBatchIntakeResponse, status_code=status.HTTP_201_CREATED)
def intake_batches(
    manifest_json: str = Form(...),
    files: list[UploadFile] = File(...),
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    if not current_user.id:
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="INVALID_USER",
            message="Usuario autenticado invalido para upload",
        )

    manifest = parse_lead_batch_intake_manifest(manifest_json)
    return execute_lead_batch_intake(
        session=session,
        user_id=int(current_user.id),
        manifest=manifest,
        files=files,
        max_import_file_bytes=MAX_IMPORT_FILE_BYTES,
    )


@router.post("/batches", response_model=LeadBatchRead, status_code=status.HTTP_201_CREATED)
@router.post("/batches/", response_model=LeadBatchRead, status_code=status.HTTP_201_CREATED)
def criar_batch(
    file: UploadFile = File(...),
    plataforma_origem: str = Form(...),
    data_envio: str = Form(...),
    evento_id: int | None = Form(default=None),
    origem_lote: str = Form(default="proponente"),
    tipo_lead_proponente: str | None = Form(default=None),
    ativacao_id: int | None = Form(default=None),
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    if not current_user.id:
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="INVALID_USER",
            message="Usuario autenticado invalido para upload",
        )
    manifest = parse_lead_batch_intake_manifest(
        json.dumps(
            {
                "items": [
                    {
                        "client_row_id": "single-upload",
                        "plataforma_origem": plataforma_origem,
                        "data_envio": data_envio,
                        "evento_id": evento_id,
                        "origem_lote": origem_lote,
                        "tipo_lead_proponente": tipo_lead_proponente,
                        "ativacao_id": ativacao_id,
                        "file_name": file.filename,
                    }
                ]
            },
            ensure_ascii=False,
        )
    )
    try:
        intake = execute_lead_batch_intake(
            session=session,
            user_id=int(current_user.id),
            manifest=manifest,
            files=[file],
            max_import_file_bytes=MAX_IMPORT_FILE_BYTES,
        )
    except Exception as exc:
        if isinstance(exc, ImportFileError):
            record_import_upload_rejection(exc, filename_hint=file.filename)
        raise
    return intake.items[0].batch


@router.get("/batches/import-hint", response_model=LeadBatchImportHintRead)
@router.get("/batches/import-hint/", response_model=LeadBatchImportHintRead)
def obter_hint_importacao_batch(
    arquivo_sha256: str | None = Query(default=None),
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    if not current_user.id:
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="INVALID_USER",
            message="Usuario autenticado invalido para consulta de hint",
        )

    normalized_hash = (arquivo_sha256 or "").strip().lower()
    if not ARQUIVO_SHA256_PATTERN.fullmatch(normalized_hash):
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="INVALID_ARQUIVO_SHA256",
            message="arquivo_sha256 deve ser hexadecimal com 64 caracteres",
            field="arquivo_sha256",
        )

    stmt = (
        select(LeadBatch)
        .options(
            load_only(
                LeadBatch.id,
                LeadBatch.arquivo_sha256,
                LeadBatch.plataforma_origem,
                LeadBatch.data_envio,
                LeadBatch.origem_lote,
                LeadBatch.tipo_lead_proponente,
                LeadBatch.evento_id,
                LeadBatch.ativacao_id,
                LeadBatch.created_at,
            )
        )
        .where(LeadBatch.enviado_por == int(current_user.id))
        .where(LeadBatch.arquivo_sha256 == normalized_hash)
        .order_by(LeadBatch.created_at.desc(), LeadBatch.id.desc())
        .limit(1)
    )
    source_batch = session.exec(stmt).first()
    if source_batch is None:
        return FastAPIResponse(status_code=status.HTTP_204_NO_CONTENT)

    logger.info(
        "lead_batch.import_hint.matched user_id=%r source_batch_id=%r arquivo_sha256=%r",
        current_user.id,
        source_batch.id,
        normalized_hash,
    )
    return LeadBatchImportHintRead(
        arquivo_sha256=normalized_hash,
        source_batch_id=int(source_batch.id),
        plataforma_origem=source_batch.plataforma_origem,
        data_envio=source_batch.data_envio,
        origem_lote=source_batch.origem_lote,
        tipo_lead_proponente=source_batch.tipo_lead_proponente,
        evento_id=source_batch.evento_id,
        ativacao_id=source_batch.ativacao_id,
        confidence="exact_hash_match",
        source_created_at=source_batch.created_at,
    )


@router.get("/batches/{batch_id}/arquivo")
@router.get("/batches/{batch_id}/arquivo/")
def download_arquivo_batch(
    batch_id: int,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    batch = session.get(LeadBatch, batch_id)
    if not batch:
        raise_http_error(
            status.HTTP_404_NOT_FOUND,
            code="BATCH_NOT_FOUND",
            message="Lote nao encontrado",
        )
    try:
        payload = read_batch_payload(batch)
    except PayloadNotFoundError:
        payload = None
    if not payload:
        raise_http_error(
            status.HTTP_404_NOT_FOUND,
            code="FILE_NOT_FOUND",
            message="Arquivo nao encontrado no lote",
        )
    filename = batch.nome_arquivo_original or "arquivo"
    ext = Path(filename).suffix.lower()
    media_type = "text/csv" if ext == ".csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return StreamingResponse(
        io.BytesIO(payload),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/batches/{batch_id}/preview", response_model=LeadBatchPreviewResponse)
@router.get("/batches/{batch_id}/preview/", response_model=LeadBatchPreviewResponse)
def preview_batch(
    batch_id: int,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    batch = session.get(LeadBatch, batch_id)
    if not batch:
        raise_http_error(
            status.HTTP_404_NOT_FOUND,
            code="BATCH_NOT_FOUND",
            message="Lote nao encontrado",
        )
    try:
        payload = read_batch_payload(batch)
    except PayloadNotFoundError:
        payload = None
    if not payload:
        raise_http_error(
            status.HTTP_404_NOT_FOUND,
            code="FILE_NOT_FOUND",
            message="Arquivo nao encontrado no lote",
        )
    filename = batch.nome_arquivo_original or ""
    try:
        preview = read_raw_file_preview(payload, filename=filename, sample_rows=3)
        result = {
            "headers": preview.headers,
            "rows": preview.rows,
            "total_rows": len(preview.rows),
        }
    except Exception:
        raise_http_error(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            code="PREVIEW_PARSE_ERROR",
            message="Nao foi possivel ler o arquivo do lote para gerar o preview.",
        )
    return LeadBatchPreviewResponse(**result)


# ---------------------------------------------------------------------------
# F2 — Silver mapping endpoints
# ---------------------------------------------------------------------------


@router.post("/batches/colunas", response_model=ColunasBatchResponse)
@router.post("/batches/colunas/", response_model=ColunasBatchResponse)
def sugerir_mapeamento_colunas_batch(
    payload: ColunasBatchRequest,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user

    try:
        preview = suggest_batch_column_mapping(payload.batch_ids, db=session)
    except ValueError as exc:
        message = str(exc)
        raise_http_error(
            status.HTTP_404_NOT_FOUND if "nao encontrado" in message else status.HTTP_400_BAD_REQUEST,
            code="BATCH_NOT_FOUND" if "nao encontrado" in message else "INVALID_BATCH_MAPPING_REQUEST",
            message=message,
        )

    return ColunasBatchResponse(
        batch_ids=preview.batch_ids,
        primary_batch_id=preview.primary_batch_id,
        aggregation_rule=preview.aggregation_rule,
        colunas=[
            BatchColumnGroupRead(
                chave_agregada=group.chave_agregada,
                nome_exibicao=group.nome_exibicao,
                variantes=group.variantes,
                aparece_em_arquivos=group.aparece_em_arquivos,
                ocorrencias=[
                    BatchColumnOccurrenceRead(
                        batch_id=occurrence.batch_id,
                        file_name=occurrence.file_name,
                        coluna_original=occurrence.coluna_original,
                        amostras=occurrence.amostras,
                        campo_sugerido=occurrence.campo_sugerido,
                        confianca=occurrence.confianca,
                        evento_id=occurrence.evento_id,
                        plataforma_origem=occurrence.plataforma_origem,
                    )
                    for occurrence in group.ocorrencias
                ],
                campo_sugerido=group.campo_sugerido,
                confianca=group.confianca,
                warnings=group.warnings,
            )
            for group in preview.colunas
        ],
        warnings=preview.warnings,
        blockers=preview.blockers,
        blocked_batch_ids=preview.blocked_batch_ids,
    )


@router.post("/batches/mapear", response_model=MapearLotesBatchResponse, status_code=status.HTTP_200_OK)
@router.post("/batches/mapear/", response_model=MapearLotesBatchResponse, status_code=status.HTTP_200_OK)
def confirmar_mapeamento_batch(
    payload: MapearLotesBatchRequest,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    if not payload.mapeamento:
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="EMPTY_MAPPING",
            message="mapeamento nao pode ser vazio",
            field="mapeamento",
        )

    if not current_user.id:
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="INVALID_USER",
            message="Usuario autenticado invalido",
        )

    try:
        result = mapear_batches(
            batch_ids=payload.batch_ids,
            mapeamento=payload.mapeamento,
            user_id=int(current_user.id),
            db=session,
        )
    except BatchMappingBlockedError as exc:
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="BATCH_MAPPING_BLOCKED",
            message="O mapeamento unificado deste batch foi bloqueado.",
            extra={"blockers": exc.blockers},
        )
    except ValueError as exc:
        message = str(exc)
        if message == "mapeamento nao pode ser vazio":
            raise_http_error(
                status.HTTP_400_BAD_REQUEST,
                code="EMPTY_MAPPING",
                message=message,
                field="mapeamento",
            )
        raise_http_error(
            status.HTTP_404_NOT_FOUND if "nao encontrado" in message else status.HTTP_400_BAD_REQUEST,
            code="BATCH_NOT_FOUND" if "nao encontrado" in message else "INVALID_BATCH_MAPPING_REQUEST",
            message=message,
        )

    return MapearLotesBatchResponse(
        batch_ids=[item.batch_id for item in result.results],
        primary_batch_id=result.primary_batch_id,
        total_silver_count=result.total_silver_count,
        results=[
            MapearLotesBatchItemResponse(
                batch_id=item.batch_id,
                silver_count=item.silver_count,
                stage=item.stage,
            )
            for item in result.results
        ],
        stage="silver",
    )


@router.get("/batches/{batch_id}/colunas", response_model=ColunasResponse)
@router.get("/batches/{batch_id}/colunas/", response_model=ColunasResponse)
def sugerir_mapeamento_colunas(
    batch_id: int,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    batch = session.get(LeadBatch, batch_id)
    if not batch:
        raise_http_error(
            status.HTTP_404_NOT_FOUND,
            code="BATCH_NOT_FOUND",
            message="Lote nao encontrado",
        )

    try:
        suggestions = suggest_column_mapping(batch_id=batch_id, db=session)
    except ValueError as exc:
        raise_http_error(
            status.HTTP_404_NOT_FOUND,
            code="BATCH_NOT_FOUND",
            message=str(exc),
        )

    return ColunasResponse(
        batch_id=batch_id,
        colunas=[
            ColumnSuggestionRead(
                coluna_original=s.coluna_original,
                campo_sugerido=s.campo_sugerido,
                confianca=s.confianca,
            )
            for s in suggestions
        ],
    )


@router.post("/batches/{batch_id}/mapear", response_model=MapearBatchResponse, status_code=status.HTTP_200_OK)
@router.post("/batches/{batch_id}/mapear/", response_model=MapearBatchResponse, status_code=status.HTTP_200_OK)
def confirmar_mapeamento(
    batch_id: int,
    payload: MapearBatchRequest,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    batch = session.get(LeadBatch, batch_id)
    if not batch:
        raise_http_error(
            status.HTTP_404_NOT_FOUND,
            code="BATCH_NOT_FOUND",
            message="Lote nao encontrado",
        )

    if not payload.mapeamento:
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="EMPTY_MAPPING",
            message="mapeamento nao pode ser vazio",
            field="mapeamento",
        )

    if not current_user.id:
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="INVALID_USER",
            message="Usuario autenticado invalido",
        )

    try:
        result = mapear_batch(
            batch_id=batch_id,
            evento_id=payload.evento_id,
            mapeamento=payload.mapeamento,
            user_id=int(current_user.id),
            db=session,
        )
    except ValueError as exc:
        raise_http_error(
            status.HTTP_404_NOT_FOUND,
            code="BATCH_NOT_FOUND",
            message=str(exc),
        )

    return MapearBatchResponse(
        batch_id=result.batch_id,
        silver_count=result.silver_count,
        stage=result.stage,
    )


# ---------------------------------------------------------------------------
# F3 — Gold pipeline endpoints
# ---------------------------------------------------------------------------


@router.get("/batches/{batch_id}", response_model=LeadBatchRead)
@router.get("/batches/{batch_id}/", response_model=LeadBatchRead)
def get_batch(
    batch_id: int,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    batch = load_batch_without_bronze(session, batch_id)
    if not batch:
        raise_http_error(
            status.HTTP_404_NOT_FOUND,
            code="BATCH_NOT_FOUND",
            message="Lote nao encontrado",
        )
    base = LeadBatchRead.model_validate(batch, from_attributes=True)
    return base.model_copy(
        update={
            "gold_pipeline_stale_after_seconds": pipeline_stale_after_seconds(),
            "gold_pipeline_progress_is_stale": is_gold_pipeline_progress_stale(batch),
        }
    )


@router.post(
    "/batches/{batch_id}/executar-pipeline",
    response_model=ExecutarPipelineResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
@router.post(
    "/batches/{batch_id}/executar-pipeline/",
    response_model=ExecutarPipelineResponse,
    status_code=status.HTTP_202_ACCEPTED,
    include_in_schema=False,
)
async def disparar_pipeline(
    batch_id: int,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    user_id = getattr(current_user, "id", None)
    batch = load_batch_without_bronze_for_update(session, batch_id)
    if not batch:
        logger.warning(
            "lead_gold_pipeline.dispatch.rejected batch_id=%r user_id=%r reason=%r",
            batch_id,
            user_id,
            "BATCH_NOT_FOUND",
        )
        raise_http_error(
            status.HTTP_404_NOT_FOUND,
            code="BATCH_NOT_FOUND",
            message="Lote nao encontrado",
        )

    if batch.stage != BatchStage.SILVER:
        logger.warning(
            "lead_gold_pipeline.dispatch.rejected batch_id=%r user_id=%r reason=%r stage=%r pipeline_status=%r evento_id=%r",
            batch_id,
            user_id,
            "INVALID_STAGE",
            batch.stage,
            batch.pipeline_status,
            batch.evento_id,
        )
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="INVALID_STAGE",
            message="Pipeline Gold so pode ser disparado em lotes com stage=silver",
            extra={"stage_atual": batch.stage},
        )

    reclaimed_stale_lock = False
    if batch.pipeline_progress is not None:
        if batch.pipeline_status == PipelineStatus.PENDING and is_gold_pipeline_progress_stale(batch):
            reclaimed_stale_lock = True
            logger.warning(
                "lead_gold_pipeline.dispatch.reclaimed_stale_lock batch_id=%r user_id=%r stage=%r pipeline_status=%r evento_id=%r",
                batch_id,
                user_id,
                batch.stage,
                batch.pipeline_status,
                batch.evento_id,
            )
            inc_gold_reclaimed_stale()
        else:
            logger.warning(
                "lead_gold_pipeline.dispatch.rejected batch_id=%r user_id=%r reason=%r stage=%r pipeline_status=%r evento_id=%r",
                batch_id,
                user_id,
                "PIPELINE_ALREADY_RUNNING",
                batch.stage,
                batch.pipeline_status,
                batch.evento_id,
            )
            raise_http_error(
                status.HTTP_409_CONFLICT,
                code="PIPELINE_ALREADY_RUNNING",
                message="Pipeline Gold ja esta em execucao para este lote",
            )

    queue_pipeline_batch(batch)
    queued_stage = batch.stage
    queued_pipeline_status = batch.pipeline_status
    queued_evento_id = batch.evento_id
    session.add(batch)
    session.commit()

    logger.info(
        "lead_gold_pipeline.dispatch.accepted batch_id=%r user_id=%r stage=%r pipeline_status=%r evento_id=%r reclaimed_stale=%r",
        batch_id,
        user_id,
        queued_stage,
        queued_pipeline_status,
        queued_evento_id,
        reclaimed_stale_lock,
    )
    return ExecutarPipelineResponse(
        batch_id=batch_id,
        status="queued",
        reclaimed_stale_lock=reclaimed_stale_lock,
    )
