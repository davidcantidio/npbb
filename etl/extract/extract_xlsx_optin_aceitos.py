"""Extractor for Eventim Opt-In Aceitos XLSX exports.

Writes staging artifacts:
- CSV with canonical columns (PII hashed by default)
- JSONL with best-effort raw row dicts (PII removed by default)
- Evidence JSON envelope
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

from npbb.etl.extract.common import (
    EvidenceEnvelope,
    hash_pii,
    snapshot_file,
    utc_now_iso,
    write_csv,
    write_json,
    write_jsonl,
)
from npbb.etl.extract.xlsx_helpers import HeaderSpec, build_optin_header_spec, infer_last_data_row, to_iso


PII_KEYS = {
    "Nome",
    "Sobrenome",
    "Email",
    "CPF",
    "Endereco Rua",
    "Endereco Numero",
    "Bairro",
    "CEP",
    "Telefone",
    "Dt Nascimento",
}


def _row_to_raw_dict(hdr: HeaderSpec, values: List[Any]) -> Dict[str, Any]:
    d: Dict[str, Any] = {}
    for i, h in enumerate(hdr.headers):
        d[h] = to_iso(values[i] if i < len(values) else None)
    return d


def extract_optin_xlsx(
    *,
    source_id: str,
    xlsx_path: Path,
    out_dir: Path,
    sheet_name: str | None = None,
    include_pii: bool = False,
) -> Dict[str, Any]:
    started = utc_now_iso()
    notes: List[str] = []

    snap = snapshot_file(xlsx_path)

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    hdr = build_optin_header_spec(ws)
    data_start = hdr.header_row + 1
    last_row = infer_last_data_row(ws, start_row=data_start, max_col=hdr.max_col)

    canonical_rows: List[Dict[str, Any]] = []
    raw_rows: List[Dict[str, Any]] = []

    for r in range(data_start, last_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, hdr.max_col + 1)]
        if all(v is None or str(v).strip() == "" for v in values):
            continue

        raw = _row_to_raw_dict(hdr, values)

        cpf = str(raw.get("CPF") or "").strip()
        email = str(raw.get("Email") or "").strip()

        canonical: Dict[str, Any] = {
            "source_id": source_id,
            "sheet_name": ws.title,
            "row_number": r,
            "evento": raw.get("Evento"),
            "sessao_start_at": raw.get("Sessao"),
            "dt_hr_compra": raw.get("Dt Hr Compra"),
            "opt_in": raw.get("Opt In"),
            "opt_in_id": raw.get("Opt In ID"),
            "opt_in_status": raw.get("Opt-in Status"),
            "canal_venda": raw.get("Canal Venda"),
            "metodo_entrega": raw.get("Metodo Entrega"),
            "ingresso": raw.get("INGRESSO") or raw.get("Ingresso"),
            "qtd_ingresso": raw.get("QTD INGRESSO") or raw.get("Qtd Ingresso"),
            "cpf_hash": hash_pii(cpf) if cpf else "",
            "email_hash": hash_pii(email) if email else "",
        }

        if include_pii:
            canonical.update(
                {
                    "cpf_raw": cpf,
                    "email_raw": email,
                    "nome": raw.get("Nome"),
                    "sobrenome": raw.get("Sobrenome"),
                    "cidade": raw.get("CIDADE") or raw.get("Cidade"),
                    "estado": raw.get("Estado"),
                    "cep": raw.get("CEP"),
                    "telefone": raw.get("TELEFONE") or raw.get("Telefone"),
                    "genero": raw.get("Genero"),
                    "dt_nascimento": raw.get("Dt Nascimento"),
                    "endereco_rua": raw.get("Endereco Rua"),
                    "endereco_numero": raw.get("Endereco Numero"),
                    "bairro": raw.get("Bairro"),
                }
            )

        canonical_rows.append(canonical)

        if include_pii:
            raw_rows.append(raw)
        else:
            scrubbed = {k: v for k, v in raw.items() if k not in PII_KEYS}
            scrubbed["cpf_hash"] = canonical["cpf_hash"]
            scrubbed["email_hash"] = canonical["email_hash"]
            raw_rows.append(scrubbed)

    finished = utc_now_iso()

    staging_csv = out_dir / f"stg_optin_aceitos__{source_id}.csv"
    staging_jsonl = out_dir / f"stg_optin_aceitos__{source_id}.raw.jsonl"
    evidence_json = out_dir / f"evidence__stg_optin_aceitos__{source_id}.json"

    base_fields = [
        "source_id",
        "sheet_name",
        "row_number",
        "evento",
        "sessao_start_at",
        "dt_hr_compra",
        "opt_in",
        "opt_in_id",
        "opt_in_status",
        "canal_venda",
        "metodo_entrega",
        "ingresso",
        "qtd_ingresso",
        "cpf_hash",
        "email_hash",
    ]
    pii_fields = [
        "cpf_raw",
        "email_raw",
        "nome",
        "sobrenome",
        "cidade",
        "estado",
        "cep",
        "telefone",
        "genero",
        "dt_nascimento",
        "endereco_rua",
        "endereco_numero",
        "bairro",
    ]
    fields = base_fields + (pii_fields if include_pii else [])

    write_csv(staging_csv, canonical_rows, fieldnames=fields)
    write_jsonl(staging_jsonl, raw_rows)

    env = EvidenceEnvelope(
        extractor="extract_xlsx_optin_aceitos",
        source_id=source_id,
        source_path=str(xlsx_path),
        started_at=started,
        finished_at=finished,
        status="OK",
        notes=notes,
        stats={
            "file_snapshot": {"sha256": snap.sha256, "size_bytes": snap.size_bytes, "mtime_utc": snap.mtime_utc},
            "sheet": ws.title,
            "header_row": hdr.header_row,
            "max_col": hdr.max_col,
            "rows_extracted": len(canonical_rows),
            "include_pii": include_pii,
            "outputs": {
                "csv": str(staging_csv),
                "jsonl": str(staging_jsonl),
            },
            "header": hdr.headers,
        },
    )
    write_json(evidence_json, env.to_dict())

    return {"csv": staging_csv, "jsonl": staging_jsonl, "evidence": evidence_json}


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source-id", required=True)
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out-dir", required=True)
    ap.add_argument("--sheet", default="")
    ap.add_argument("--include-pii", action="store_true")
    args = ap.parse_args(argv)

    extract_optin_xlsx(
        source_id=args.source_id,
        xlsx_path=Path(args.xlsx),
        out_dir=Path(args.out_dir),
        sheet_name=args.sheet or None,
        include_pii=bool(args.include_pii),
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
