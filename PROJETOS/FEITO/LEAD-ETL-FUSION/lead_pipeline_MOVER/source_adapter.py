from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook

from .constants import (
    CITY_TO_EVENT_DATE,
    CITY_TO_LOCAL_UF,
    EVENTO_PADRAO,
    MAPPING_VERSION,
    REQUIRED_COLUMNS,
    TIPO_EVENTO_PADRAO,
    WARNING_ARQUIVO_ENCRIPTADO,
    WARNING_BATUKE_DATA_SEM_PRACA_CONFIRMADA,
)
from .normalization import canonicalize_header, city_key, normalize_header, parse_date, strip_accents


PROFILE_CANONICO = "modelo_canonico"
PROFILE_BB_V1 = "bb_semestre_v1"
PROFILE_BB_V2 = "bb_semestre_v2"
PROFILE_PARK = "park_challenge_participantes"
PROFILE_TICKETING = "sls_ticketing_multi_sheet"
PROFILE_FORMS_PUBLICO = "sls_forms_publico"
PROFILE_FORMS_SKATISTAS = "sls_forms_skatistas"
PROFILE_NPS = "sls_nps_multi_sheet"
PROFILE_LANDING = "sls_landing_page"
PROFILE_VERT_BATTLE = "vert_battle_entidades_v1"
PROFILE_BATUKE = "batuke_multiedicao_v1"
PROFILE_DESCONHECIDO = "desconhecido"

SHEET_COL_SOURCE = "__source_sheet"
ROW_COL_SOURCE = "__source_row"
CITY_OUT_COL = "__cidade_fora_mapeamento"
REJECT_REASON_COL = "__reject_reason"


VERT_BATTLE_EVENT_MAP = {
    city_key("VERT BATTLE :: NITERÓI"): ("Niterói-RJ", "2025-05-30"),
    city_key("VERT BATTLE :: GOIÂNIA"): ("Goiânia-GO", "2025-05-02"),
    city_key("VERT BATTLE :: GYN"): ("Goiânia-GO", "2025-05-02"),
}


BATUKE_EVENT_MAP = {
    "2025-11-13": (
        "Batuke do Pretinho - Rio de Janeiro - 2025-11-13",
        "Rio de Janeiro-RJ",
    ),
    "2025-11-20": (
        "Batuke do Pretinho - Rio de Janeiro - 2025-11-20",
        "Rio de Janeiro-RJ",
    ),
    "2025-11-23": (
        "Batuke do Pretinho - São Paulo - 2025-11-23",
        "São Paulo-SP",
    ),
    "2025-11-27": (
        "Batuke do Pretinho - Rio de Janeiro - 2025-11-27",
        "Rio de Janeiro-RJ",
    ),
    "2025-12-04": (
        "Batuke do Pretinho - Rio de Janeiro - 2025-12-04",
        "Rio de Janeiro-RJ",
    ),
    "2025-12-07": (
        "Batuke do Pretinho - São Paulo - 2025-12-07",
        "São Paulo-SP",
    ),
    "2025-12-13": (
        "Batuke do Pretinho - Belo Horizonte - 2025-12-13",
        "Belo Horizonte-MG",
    ),
    "2025-12-14": (
        "Batuke do Pretinho - Rio de Janeiro - 2025-12-14",
        "Rio de Janeiro-RJ",
    ),
    "2025-12-21": (
        "Batuke do Pretinho - Rio de Janeiro - 2025-12-21",
        "Rio de Janeiro-RJ",
    ),
    "2025-12-28": (
        "Batuke do Pretinho - Rio de Janeiro - 2025-12-28",
        "Rio de Janeiro-RJ",
    ),
}


@dataclass
class AdaptedInput:
    dataframe: pd.DataFrame
    source_profiles: list[str]
    fail_reasons: list[str]
    warnings: list[str]
    skipped: bool = False
    skip_reason: str | None = None


def _empty_dataframe() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            *REQUIRED_COLUMNS,
            CITY_OUT_COL,
            SHEET_COL_SOURCE,
            ROW_COL_SOURCE,
            REJECT_REASON_COL,
        ]
    )


def _clean_series(series: pd.Series) -> pd.Series:
    cleaned = series.fillna("").astype(str).str.strip()
    return cleaned.mask(cleaned.str.lower().isin({"nan", "none", "null", "nat"}), "")


def _normalized_header_map(df: pd.DataFrame) -> tuple[list[str], dict[str, list[str]]]:
    normalized_list = [canonicalize_header(column) for column in df.columns]
    header_map: dict[str, list[str]] = {}
    for original, normalized in zip(df.columns, normalized_list):
        header_map.setdefault(normalized, []).append(original)
    return normalized_list, header_map


def _series_from_candidates(
    df: pd.DataFrame,
    header_map: dict[str, list[str]],
    candidates: list[str],
    default: str = "",
) -> pd.Series:
    for candidate in candidates:
        originals = header_map.get(candidate, [])
        if originals:
            return _clean_series(df[originals[0]])
    return pd.Series([default] * len(df), index=df.index, dtype="object")


def _first_column_with_tokens(columns: list[str], tokens: list[str]) -> str | None:
    for column in columns:
        canonical = canonicalize_header(column)
        if all(token in canonical for token in tokens):
            return column
    return None


def _drop_empty_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    mask = df.apply(lambda row: any(str(v).strip() for v in row.values), axis=1)
    return df[mask]


def _extract_day_month_to_iso(text: str, year: int = 2025) -> str:
    match = re.search(r"(\d{1,2})/(\d{1,2})", str(text or ""))
    if not match:
        return ""
    day = int(match.group(1))
    month = int(match.group(2))
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return ""


def _parse_city_source(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = text.replace("–", "-").replace("—", "-")
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"\s*[-/]\s*[A-Za-z]{2}$", "", text).strip()
    return text


def _map_city_and_event_date(raw_city: str) -> tuple[str, str, bool]:
    cleaned_city = _parse_city_source(raw_city)
    if not cleaned_city:
        return "", "", True

    key = city_key(cleaned_city)
    mapped_local = CITY_TO_LOCAL_UF.get(key, "")
    mapped_date = CITY_TO_EVENT_DATE.get(key, "")
    city_out = not (mapped_local and mapped_date)
    if city_out:
        return "", "", True
    return mapped_local, mapped_date, False


def _is_encrypted_ole_office(path: Path) -> bool:
    try:
        with path.open("rb") as file_handle:
            return file_handle.read(8) == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
    except OSError:
        return False


def _normalize_filename(value: str) -> str:
    return canonicalize_header(strip_accents(str(value or "").lower()))


def detect_source_profile(df: pd.DataFrame) -> str:
    normalized_list, _ = _normalized_header_map(df)
    normalized_set = set(normalized_list)

    normalized_with_synonyms = [normalize_header(column) for column in df.columns]
    normalized_with_synonyms_set = set(normalized_with_synonyms)
    if len(normalized_with_synonyms) == len(set(normalized_with_synonyms)):
        if set(REQUIRED_COLUMNS).issubset(normalized_with_synonyms_set):
            return PROFILE_CANONICO

    base_keys = {"nome_comprador", "documento", "evento"}
    if base_keys.issubset(normalized_set):
        if "soma_de_id_inscricao" in normalized_set or "e_mail" in normalized_set:
            return PROFILE_BB_V2
        if "id_inscricao" in normalized_set:
            return PROFILE_BB_V1

    if normalized_with_synonyms_set.intersection(REQUIRED_COLUMNS):
        return PROFILE_CANONICO

    return PROFILE_DESCONHECIDO


def _adapt_canonical(df: pd.DataFrame, input_file: Path) -> tuple[pd.DataFrame, list[str]]:
    fail_reasons: list[str] = []

    normalized = [normalize_header(column) for column in df.columns]
    duplicated = sorted({name for name in normalized if normalized.count(name) > 1})
    if duplicated:
        fail_reasons.append(
            f"COLUNAS_DUPLICADAS_APOS_NORMALIZACAO [{input_file.name}]: {duplicated}"
        )
        return _empty_dataframe(), fail_reasons

    canonical_df = df.copy()
    canonical_df.columns = normalized
    missing = [column for column in REQUIRED_COLUMNS if column not in canonical_df.columns]
    if missing:
        fail_reasons.append(
            f"COLUNAS_OBRIGATORIAS_AUSENTES [{input_file.name}]: {missing}"
        )
        return _empty_dataframe(), fail_reasons

    canonical_df = canonical_df[REQUIRED_COLUMNS].copy()
    canonical_df[CITY_OUT_COL] = False
    canonical_df[SHEET_COL_SOURCE] = ""
    canonical_df[ROW_COL_SOURCE] = canonical_df.index + 2
    canonical_df[REJECT_REASON_COL] = ""
    return canonical_df, fail_reasons


def _adapt_bb(df: pd.DataFrame, input_file: Path, profile: str) -> tuple[pd.DataFrame, list[str]]:
    fail_reasons: list[str] = []
    _, header_map = _normalized_header_map(df)

    has_name = "nome_comprador" in header_map
    has_document = "documento" in header_map
    has_local = "local" in header_map or "cidade" in header_map
    if not (has_name and has_document and has_local):
        fail_reasons.append(
            "COLUNAS_MINIMAS_AUSENTES "
            f"[{input_file.name}] para perfil {profile}: "
            "requer Nome Comprador, Documento e Local/Cidade."
        )
        return _empty_dataframe(), fail_reasons

    local_primary = _series_from_candidates(df, header_map, ["local"])
    local_fallback = _series_from_candidates(df, header_map, ["cidade"])
    local_source = local_primary.where(local_primary.str.strip() != "", local_fallback)

    mapped_local: list[str] = []
    mapped_date: list[str] = []
    city_out_of_mapping: list[bool] = []
    for city in local_source.tolist():
        local_value, event_date, city_out = _map_city_and_event_date(city)
        mapped_local.append(local_value)
        mapped_date.append(event_date)
        city_out_of_mapping.append(city_out)

    adapted = pd.DataFrame(
        {
            "nome": _series_from_candidates(df, header_map, ["nome_comprador"]),
            "cpf": _series_from_candidates(df, header_map, ["documento"]),
            "data_nascimento": "",
            "email": _series_from_candidates(df, header_map, ["e_mail", "email"]),
            "telefone": "",
            "evento": EVENTO_PADRAO,
            "tipo_evento": TIPO_EVENTO_PADRAO,
            "local": mapped_local,
            "data_evento": mapped_date,
            CITY_OUT_COL: city_out_of_mapping,
            SHEET_COL_SOURCE: "",
            ROW_COL_SOURCE: df.index + 2,
            REJECT_REASON_COL: "",
        }
    )
    return adapted, fail_reasons


def adapt_input_dataframe(df: pd.DataFrame, input_file: Path) -> AdaptedInput:
    profile = detect_source_profile(df)
    if profile == PROFILE_CANONICO:
        adapted_df, fail_reasons = _adapt_canonical(df, input_file)
    elif profile in {PROFILE_BB_V1, PROFILE_BB_V2}:
        adapted_df, fail_reasons = _adapt_bb(df, input_file, profile)
    else:
        fail_reasons = [
            f"PERFIL_FONTE_NAO_SUPORTADO [{input_file.name}] para adaptacao automatica."
        ]
        adapted_df = _empty_dataframe()

    return AdaptedInput(
        dataframe=adapted_df,
        source_profiles=[profile],
        fail_reasons=fail_reasons,
        warnings=[],
    )


def _build_df_with_meta(
    *,
    source_df: pd.DataFrame,
    source_sheet: str,
    source_row_offset: int,
    nome: pd.Series,
    cpf: pd.Series,
    data_nascimento: pd.Series | str,
    email: pd.Series | str,
    telefone: pd.Series | str,
    evento: pd.Series | str,
    tipo_evento: pd.Series | str,
    local: pd.Series | str,
    data_evento: pd.Series | str,
    city_out: pd.Series | bool = False,
    reject_reason: pd.Series | str = "",
) -> pd.DataFrame:
    base_len = len(source_df)

    def series_or_constant(value: pd.Series | str | bool) -> pd.Series:
        if isinstance(value, pd.Series):
            return _clean_series(value).reindex(source_df.index)
        return pd.Series([str(value)] * base_len, index=source_df.index, dtype="object")

    if isinstance(city_out, pd.Series):
        city_out_series = city_out.reindex(source_df.index).fillna(False).astype(bool)
    else:
        city_out_series = pd.Series([bool(city_out)] * base_len, index=source_df.index, dtype=bool)

    adapted = pd.DataFrame(
        {
            "nome": series_or_constant(nome),
            "cpf": series_or_constant(cpf),
            "data_nascimento": series_or_constant(data_nascimento),
            "email": series_or_constant(email),
            "telefone": series_or_constant(telefone),
            "evento": series_or_constant(evento),
            "tipo_evento": series_or_constant(tipo_evento),
            "local": series_or_constant(local),
            "data_evento": series_or_constant(data_evento),
            CITY_OUT_COL: city_out_series,
            SHEET_COL_SOURCE: source_sheet,
            ROW_COL_SOURCE: source_df.index + source_row_offset,
            REJECT_REASON_COL: series_or_constant(reject_reason),
        }
    )
    return adapted


def _adapt_park(input_file: Path) -> AdaptedInput:
    sheet_name = "Lista de participantes"
    df = pd.read_excel(input_file, sheet_name=sheet_name, header=7, dtype=str)
    df = _drop_empty_rows(df)

    _, header_map = _normalized_header_map(df)
    nome = _series_from_candidates(df, header_map, ["nome"])
    sobrenome = _series_from_candidates(df, header_map, ["sobrenome"])
    full_name = (nome + " " + sobrenome).str.strip()
    cpf = _series_from_candidates(df, header_map, ["cpf"])
    email = _series_from_candidates(df, header_map, ["email"])
    telefone = _series_from_candidates(df, header_map, ["celular"])
    birth = _series_from_candidates(df, header_map, ["data_de_nascimento"])

    wb = load_workbook(input_file, read_only=True, data_only=True)
    ws = wb[sheet_name]
    raw_event_date = ws["A3"].value
    event_date = parse_date(raw_event_date)
    if not event_date:
        event_date = "2025-11-02"

    adapted = _build_df_with_meta(
        source_df=df,
        source_sheet=sheet_name,
        source_row_offset=9,
        nome=full_name,
        cpf=cpf,
        data_nascimento=birth,
        email=email,
        telefone=telefone,
        evento="Park Challenge 2025",
        tipo_evento=TIPO_EVENTO_PADRAO,
        local="Florianópolis-SC",
        data_evento=event_date,
        city_out=False,
    )
    return AdaptedInput(dataframe=adapted, source_profiles=[PROFILE_PARK], fail_reasons=[], warnings=[])


def _adapt_ticketing(input_file: Path) -> AdaptedInput:
    sheet_metadata = {
        "Brasilia 2025": ("SLS Select Series Brasília 2025", "Brasília-DF", "2025-07-13"),
        "Super Crown 2025": ("SLS Super Crown 2025", "São Paulo-SP", "2025-12-07"),
    }
    xls = pd.ExcelFile(input_file)
    frames: list[pd.DataFrame] = []
    for sheet_name, (evento, local, event_date) in sheet_metadata.items():
        if sheet_name not in xls.sheet_names:
            continue
        df = pd.read_excel(input_file, sheet_name=sheet_name, dtype=str)
        df = _drop_empty_rows(df)
        _, header_map = _normalized_header_map(df)
        first_name = _series_from_candidates(df, header_map, ["first_name"])
        last_name = _series_from_candidates(df, header_map, ["last_name"])
        full_name = (first_name + " " + last_name).str.strip()
        cpf = _series_from_candidates(df, header_map, ["document_number"])
        email = _series_from_candidates(df, header_map, ["email"])

        adapted = _build_df_with_meta(
            source_df=df,
            source_sheet=sheet_name,
            source_row_offset=2,
            nome=full_name,
            cpf=cpf,
            data_nascimento="",
            email=email,
            telefone="",
            evento=evento,
            tipo_evento=TIPO_EVENTO_PADRAO,
            local=local,
            data_evento=event_date,
            city_out=False,
        )
        frames.append(adapted)

    if not frames:
        return AdaptedInput(
            dataframe=_empty_dataframe(),
            source_profiles=[PROFILE_TICKETING],
            fail_reasons=[f"ABAS_VALIDAS_AUSENTES [{input_file.name}] para ticketing."],
            warnings=[],
        )
    return AdaptedInput(
        dataframe=pd.concat(frames, ignore_index=True),
        source_profiles=[PROFILE_TICKETING],
        fail_reasons=[],
        warnings=[],
    )


def _adapt_ssf_publico(input_file: Path) -> AdaptedInput:
    sheet_name = "Respostas ao formulário 1"
    df = pd.read_excel(input_file, sheet_name=sheet_name, dtype=str)
    df = _drop_empty_rows(df)
    columns = list(df.columns)

    nome_col = _first_column_with_tokens(columns, ["nome", "completo"])
    cpf_col = _first_column_with_tokens(columns, ["cpf"])
    email_col = _first_column_with_tokens(columns, ["e_mail"])
    birth_col = _first_column_with_tokens(columns, ["data", "nascimento"])
    day_col = _first_column_with_tokens(columns, ["qual", "dia", "evento"])

    if not (nome_col and cpf_col and email_col and day_col):
        return AdaptedInput(
            dataframe=_empty_dataframe(),
            source_profiles=[PROFILE_FORMS_PUBLICO],
            fail_reasons=[f"COLUNAS_MINIMAS_AUSENTES [{input_file.name}] em {sheet_name}."],
            warnings=[],
        )

    event_date = _clean_series(df[day_col]).map(_extract_day_month_to_iso)
    adapted = _build_df_with_meta(
        source_df=df,
        source_sheet=sheet_name,
        source_row_offset=2,
        nome=_clean_series(df[nome_col]),
        cpf=_clean_series(df[cpf_col]),
        data_nascimento=_clean_series(df[birth_col]) if birth_col else "",
        email=_clean_series(df[email_col]),
        telefone="",
        evento="SLS Select Series Florianópolis 2025",
        tipo_evento=TIPO_EVENTO_PADRAO,
        local="Florianópolis-SC",
        data_evento=event_date,
        city_out=False,
    )
    return AdaptedInput(dataframe=adapted, source_profiles=[PROFILE_FORMS_PUBLICO], fail_reasons=[], warnings=[])


def _adapt_ssf_skatistas(input_file: Path) -> AdaptedInput:
    sheet_name = "Respostas ao formulário 1"
    df = pd.read_excel(input_file, sheet_name=sheet_name, dtype=str)
    df = _drop_empty_rows(df)
    columns = list(df.columns)

    nome_col = _first_column_with_tokens(columns, ["nome", "completo"])
    cpf_col = _first_column_with_tokens(columns, ["cpf"])
    email_col = _first_column_with_tokens(columns, ["e_mail"])
    comp_col = _first_column_with_tokens(columns, ["de", "qual", "competicao"])

    if not (nome_col and cpf_col and email_col and comp_col):
        return AdaptedInput(
            dataframe=_empty_dataframe(),
            source_profiles=[PROFILE_FORMS_SKATISTAS],
            fail_reasons=[f"COLUNAS_MINIMAS_AUSENTES [{input_file.name}] em {sheet_name}."],
            warnings=[],
        )

    event_date = _clean_series(df[comp_col]).map(_extract_day_month_to_iso)
    adapted = _build_df_with_meta(
        source_df=df,
        source_sheet=sheet_name,
        source_row_offset=2,
        nome=_clean_series(df[nome_col]),
        cpf=_clean_series(df[cpf_col]),
        data_nascimento="",
        email=_clean_series(df[email_col]),
        telefone="",
        evento="SLS Select Series Florianópolis 2025",
        tipo_evento=TIPO_EVENTO_PADRAO,
        local="Florianópolis-SC",
        data_evento=event_date,
        city_out=False,
    )
    return AdaptedInput(dataframe=adapted, source_profiles=[PROFILE_FORMS_SKATISTAS], fail_reasons=[], warnings=[])


def _coalesce_series(values: list[pd.Series], index: pd.Index) -> pd.Series:
    result = pd.Series([""] * len(index), index=index, dtype="object")
    for series in values:
        cleaned = _clean_series(series).reindex(index)
        result = result.where(result != "", cleaned)
    return result


def _adapt_sss_nps(input_file: Path) -> AdaptedInput:
    xls = pd.ExcelFile(input_file)
    frames: list[pd.DataFrame] = []
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(input_file, sheet_name=sheet_name, dtype=str)
        df = _drop_empty_rows(df)
        columns = list(df.columns)
        nome_col = _first_column_with_tokens(columns, ["nome", "completo"])
        cpf_col = _first_column_with_tokens(columns, ["cpf"])
        if not (nome_col and cpf_col):
            continue

        email_columns = [
            column
            for column in columns
            if "e_mail" in canonicalize_header(column)
        ]
        email_series = (
            _coalesce_series([df[column] for column in email_columns], df.index)
            if email_columns
            else pd.Series([""] * len(df), index=df.index, dtype="object")
        )
        birth_col = _first_column_with_tokens(columns, ["data", "nascimento"])
        if canonicalize_header(sheet_name) == canonicalize_header("NPS PÚBLICO"):
            day_col = _first_column_with_tokens(columns, ["qual", "dia", "evento"])
            event_date = (
                _clean_series(df[day_col]).map(_extract_day_month_to_iso)
                if day_col
                else pd.Series([""] * len(df), index=df.index, dtype="object")
            )
        else:
            event_date = "2025-09-13"

        adapted = _build_df_with_meta(
            source_df=df,
            source_sheet=sheet_name,
            source_row_offset=2,
            nome=_clean_series(df[nome_col]),
            cpf=_clean_series(df[cpf_col]),
            data_nascimento=_clean_series(df[birth_col]) if birth_col else "",
            email=email_series,
            telefone="",
            evento="SLS Select Series Saquarema 2025",
            tipo_evento=TIPO_EVENTO_PADRAO,
            local="Saquarema-RJ",
            data_evento=event_date,
            city_out=False,
        )
        frames.append(adapted)

    if not frames:
        return AdaptedInput(
            dataframe=_empty_dataframe(),
            source_profiles=[PROFILE_NPS],
            fail_reasons=[f"ABAS_VALIDAS_AUSENTES [{input_file.name}] para NPS."],
            warnings=[],
        )
    return AdaptedInput(
        dataframe=pd.concat(frames, ignore_index=True),
        source_profiles=[PROFILE_NPS],
        fail_reasons=[],
        warnings=[],
    )


def _adapt_sls_landing(input_file: Path) -> AdaptedInput:
    xls = pd.ExcelFile(input_file)
    sheet_name = xls.sheet_names[0]
    df = pd.read_excel(input_file, sheet_name=sheet_name, dtype=str)
    df = _drop_empty_rows(df)
    _, header_map = _normalized_header_map(df)

    nome = _series_from_candidates(df, header_map, ["nome"])
    cpf = _series_from_candidates(df, header_map, ["cpf"])
    email = _series_from_candidates(df, header_map, ["email"])
    telefone = _series_from_candidates(df, header_map, ["telefone"])
    birth = _series_from_candidates(df, header_map, ["data_nasc", "data_de_nascimento"])

    adapted = _build_df_with_meta(
        source_df=df,
        source_sheet=sheet_name,
        source_row_offset=2,
        nome=nome,
        cpf=cpf,
        data_nascimento=birth,
        email=email,
        telefone=telefone,
        evento="SLS Super Crown 2025",
        tipo_evento=TIPO_EVENTO_PADRAO,
        local="São Paulo-SP",
        data_evento="2025-12-07",
        city_out=False,
    )
    return AdaptedInput(dataframe=adapted, source_profiles=[PROFILE_LANDING], fail_reasons=[], warnings=[])


def _adapt_batuke(input_file: Path) -> AdaptedInput:
    xls = pd.ExcelFile(input_file)
    sheet_name = xls.sheet_names[0]
    df = pd.read_excel(input_file, sheet_name=sheet_name, dtype=str)
    df = _drop_empty_rows(df)
    _, header_map = _normalized_header_map(df)

    required_headers = {"nome", "cpf", "email", "telefone", "dt_ref_data_hora"}
    if not required_headers.issubset(header_map):
        return AdaptedInput(
            dataframe=_empty_dataframe(),
            source_profiles=[PROFILE_BATUKE],
            fail_reasons=[
                "COLUNAS_MINIMAS_AUSENTES "
                f"[{input_file.name}] para perfil {PROFILE_BATUKE}: "
                "requer nome, cpf, email, telefone e dt_ref_data_hora."
            ],
            warnings=[],
        )

    event_dates = _series_from_candidates(df, header_map, ["dt_ref_data_hora"])
    mapped_events: list[str] = []
    mapped_locals: list[str] = []
    reject_reasons: list[str] = []

    for raw_date in event_dates.tolist():
        normalized_date = parse_date(raw_date)
        mapped = BATUKE_EVENT_MAP.get(normalized_date or "")
        if mapped is None:
            mapped_events.append("Batuke do Pretinho")
            mapped_locals.append("")
            reject_reasons.append(WARNING_BATUKE_DATA_SEM_PRACA_CONFIRMADA)
            continue

        event_name, local = mapped
        mapped_events.append(event_name)
        mapped_locals.append(local)
        reject_reasons.append("")

    adapted = _build_df_with_meta(
        source_df=df,
        source_sheet=sheet_name,
        source_row_offset=2,
        nome=_series_from_candidates(df, header_map, ["nome"]),
        cpf=_series_from_candidates(df, header_map, ["cpf"]),
        data_nascimento="",
        email=_series_from_candidates(df, header_map, ["email"]),
        telefone=_series_from_candidates(df, header_map, ["telefone"]),
        evento=pd.Series(mapped_events, index=df.index, dtype="object"),
        tipo_evento="Show",
        local=pd.Series(mapped_locals, index=df.index, dtype="object"),
        data_evento=event_dates,
        city_out=False,
        reject_reason=pd.Series(reject_reasons, index=df.index, dtype="object"),
    )
    return AdaptedInput(
        dataframe=adapted,
        source_profiles=[PROFILE_BATUKE],
        fail_reasons=[],
        warnings=[],
    )


def _map_vert_battle_event(event_name: str) -> tuple[str, str, bool]:
    key = city_key(event_name)
    if not key:
        return "", "", True
    mapped = VERT_BATTLE_EVENT_MAP.get(key)
    if not mapped:
        return "", "", True
    local, event_date = mapped
    return local, event_date, False


def _adapt_vert_battle(input_file: Path) -> AdaptedInput:
    sheet_name = "Entidades"
    xls = pd.ExcelFile(input_file)
    if sheet_name not in xls.sheet_names:
        return AdaptedInput(
            dataframe=_empty_dataframe(),
            source_profiles=[PROFILE_VERT_BATTLE],
            fail_reasons=[f"ABA_OBRIGATORIA_AUSENTE [{input_file.name}] em {sheet_name}."],
            warnings=[],
        )

    df = pd.read_excel(input_file, sheet_name=sheet_name, dtype=str)
    df = _drop_empty_rows(df)
    _, header_map = _normalized_header_map(df)
    required_headers = {"nome", "sobrenome", "cpf", "email", "telefone", "evento"}
    if not required_headers.issubset(header_map):
        return AdaptedInput(
            dataframe=_empty_dataframe(),
            source_profiles=[PROFILE_VERT_BATTLE],
            fail_reasons=[
                "COLUNAS_MINIMAS_AUSENTES "
                f"[{input_file.name}] para perfil {PROFILE_VERT_BATTLE}: "
                "requer Nome, Sobrenome, CPF, Email, Telefone e Evento."
            ],
            warnings=[],
        )

    nome = _series_from_candidates(df, header_map, ["nome"])
    sobrenome = _series_from_candidates(df, header_map, ["sobrenome"])
    full_name = (nome + " " + sobrenome).str.strip()
    evento = _series_from_candidates(df, header_map, ["evento"])

    mapped_local: list[str] = []
    mapped_event_date: list[str] = []
    city_out_of_mapping: list[bool] = []
    for event_name in evento.tolist():
        local, event_date, city_out = _map_vert_battle_event(event_name)
        mapped_local.append(local)
        mapped_event_date.append(event_date)
        city_out_of_mapping.append(city_out)

    adapted = _build_df_with_meta(
        source_df=df,
        source_sheet=sheet_name,
        source_row_offset=2,
        nome=full_name,
        cpf=_series_from_candidates(df, header_map, ["cpf"]),
        data_nascimento=_series_from_candidates(df, header_map, ["datanascimento"]),
        email=_series_from_candidates(df, header_map, ["email"]),
        telefone=_series_from_candidates(df, header_map, ["telefone"]),
        evento=evento,
        tipo_evento=TIPO_EVENTO_PADRAO,
        local=pd.Series(mapped_local, index=df.index, dtype="object"),
        data_evento=pd.Series(mapped_event_date, index=df.index, dtype="object"),
        city_out=pd.Series(city_out_of_mapping, index=df.index, dtype=bool),
    )
    return AdaptedInput(
        dataframe=adapted,
        source_profiles=[PROFILE_VERT_BATTLE],
        fail_reasons=[],
        warnings=[],
    )


def _adapt_xlsx(input_file: Path) -> AdaptedInput:
    if _is_encrypted_ole_office(input_file):
        return AdaptedInput(
            dataframe=_empty_dataframe(),
            source_profiles=[PROFILE_DESCONHECIDO],
            fail_reasons=[],
            warnings=[WARNING_ARQUIVO_ENCRIPTADO],
            skipped=True,
            skip_reason=WARNING_ARQUIVO_ENCRIPTADO,
        )

    normalized_name = _normalize_filename(input_file.name)
    if "park_challenge_base_de_leads_banco_do_brasil" in normalized_name:
        return _adapt_park(input_file)
    if "base_de_leads_brasilia_super_crown_ticketing" in normalized_name:
        return _adapt_ticketing(input_file)
    if "ssf_25_base_leads_publico" in normalized_name:
        return _adapt_ssf_publico(input_file)
    if "ssf_25_base_leads_skatistas" in normalized_name:
        return _adapt_ssf_skatistas(input_file)
    if "sss_25_base_de_leads_nps" in normalized_name:
        return _adapt_sss_nps(input_file)
    if "sls_sc_2025_leads_landingpage" in normalized_name:
        return _adapt_sls_landing(input_file)
    if "batuke_do_pretinho" in normalized_name:
        return _adapt_batuke(input_file)
    if "vert_battle" in normalized_name and "leads" in normalized_name:
        return _adapt_vert_battle(input_file)

    return AdaptedInput(
        dataframe=_empty_dataframe(),
        source_profiles=[PROFILE_DESCONHECIDO],
        fail_reasons=[
            f"PERFIL_FONTE_NAO_SUPORTADO [{input_file.name}] para adaptacao automatica."
        ],
        warnings=[],
    )


def adapt_source_file(input_file: Path) -> AdaptedInput:
    suffix = input_file.suffix.lower()
    if suffix == ".csv":
        try:
            df = pd.read_csv(input_file, dtype=str, keep_default_na=False)
            return adapt_input_dataframe(df, input_file)
        except UnicodeDecodeError:
            # Some sources are delivered as XLSX with .csv extension.
            return _adapt_xlsx(input_file)
    if suffix == ".xlsx":
        return _adapt_xlsx(input_file)
    return AdaptedInput(
        dataframe=_empty_dataframe(),
        source_profiles=[PROFILE_DESCONHECIDO],
        fail_reasons=[f"ARQUIVO_NAO_SUPORTADO [{input_file.name}]"],
        warnings=[],
    )


def get_mapping_version() -> str:
    return MAPPING_VERSION
