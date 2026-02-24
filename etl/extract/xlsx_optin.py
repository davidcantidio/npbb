"""XLSX opt-in extractor with staging load and lineage registration.

This module provides:
- `extract_optin_xlsx(path) -> iterator[dict]` for deterministic row extraction,
- `load_optin_xlsx_to_staging(...)` to persist staging rows with ingestion and lineage.
"""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
import json
from pathlib import Path
import sys
from typing import Any, Iterator, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlmodel import Session

try:  # package execution (repo root at `npbb`)
    from etl.extract.common import hash_pii
    from etl.extract.xlsx_utils import build_columns_with_metadata, find_header_row
except ModuleNotFoundError:  # pragma: no cover - fallback for `npbb.etl.*` execution
    from npbb.etl.extract.common import hash_pii
    from npbb.etl.extract.xlsx_utils import build_columns_with_metadata, find_header_row

try:  # pragma: no cover - import style depends on caller cwd
    from core.registry.location_ref import format_location
except ModuleNotFoundError:  # pragma: no cover
    from npbb.core.registry.location_ref import format_location

try:  # pragma: no cover
    from app.db.database import engine
    from app.models.etl_registry import IngestionStatus
    from app.models.stg_optin import StgOptinTransaction
    from app.services.etl_lineage_service import create_lineage_ref
    from app.services.etl_registry_service import (
        enforce_lineage_policy,
        finish_ingestion_run,
        register_source_from_path,
        start_ingestion_run,
    )
    from app.services.session_resolver import resolve_session_reference
except ModuleNotFoundError:  # pragma: no cover
    BACKEND_ROOT = Path(__file__).resolve().parents[2] / "backend"
    if str(BACKEND_ROOT) not in sys.path:
        sys.path.insert(0, str(BACKEND_ROOT))
    from app.db.database import engine
    from app.models.etl_registry import IngestionStatus
    from app.models.stg_optin import StgOptinTransaction
    from app.services.etl_lineage_service import create_lineage_ref
    from app.services.etl_registry_service import (
        enforce_lineage_policy,
        finish_ingestion_run,
        register_source_from_path,
        start_ingestion_run,
    )
    from app.services.session_resolver import resolve_session_reference


_DEFAULT_REQUIRED_TERMS: tuple[str, ...] = ("CPF", "Opt", "Evento")
_DEFAULT_ALIASES: dict[str, str] = {
    "cliente_nome": "nome",
    "cliente_sobrenome": "sobrenome",
    "cliente_email": "email",
    "cliente_cpf": "cpf",
    "venda_data_compra": "dt_hr_compra",
    "venda_opt_in": "opt_in",
    "venda_opt_in_id": "opt_in_id",
    "venda_opt_in_status": "opt_in_status",
    "venda_qtd_ingresso": "qtd_ingresso",
    "venda_ingresso": "ingresso",
    "venda_canal_venda": "canal_venda",
    "venda_metodo_entrega": "metodo_entrega",
    "dt_e_hr_compra": "dt_hr_compra",
    "dt_hr_compra": "dt_hr_compra",
    "opt_in": "opt_in",
    "opt_in_id": "opt_in_id",
    "opt_in_status": "opt_in_status",
    "qtd_ingresso": "qtd_ingresso",
    "canal_venda": "canal_venda",
    "metodo_entrega": "metodo_entrega",
    "evento": "evento",
    "sessao": "sessao",
    "session": "sessao",
    "cpf": "cpf",
    "email": "email",
    "nome": "nome",
    "sobrenome": "sobrenome",
}


def _to_text(value: Any) -> str | None:
    """Return trimmed text value or `None` for empty cells."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_int(value: Any) -> int | None:
    """Parse numeric-like values to integer when possible."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(".", "").replace(",", "")
    if text.isdigit():
        return int(text)
    return None


def _to_datetime(value: Any) -> datetime | None:
    """Parse date/time values to timezone-aware datetime when possible."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(" ", "T")
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)


def _find_value(row: Mapping[str, Any], keys: Sequence[str]) -> Any:
    """Return first non-empty value from candidate keys in one extracted row."""
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return value
    return None


def extract_optin_xlsx(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    aliases: Mapping[str, str] | None = None,
) -> Iterator[dict[str, Any]]:
    """Extract opt-in rows from XLSX with header normalization metadata.

    Args:
        path: XLSX file path.
        sheet_name: Optional single-sheet selection.
        aliases: Optional normalized-column aliases.

    Yields:
        Extracted row dictionaries with canonical fields plus lineage metadata keys:
        `__sheet_name`, `__header_row`, `__header_range`, `__used_range`,
        `__source_range`.
    """

    xlsx_path = Path(path).expanduser()
    workbook = load_workbook(xlsx_path, data_only=True)
    sheets = [workbook[sheet_name]] if sheet_name else [workbook[name] for name in workbook.sheetnames]
    alias_map = dict(_DEFAULT_ALIASES)
    if aliases:
        alias_map.update(aliases)

    for ws in sheets:
        header_row = find_header_row(ws, required_terms=_DEFAULT_REQUIRED_TERMS, max_scan_rows=40)
        columns_result = build_columns_with_metadata(
            ws,
            header_row=header_row,
            header_depth=2,
            aliases=alias_map,
        )
        column_names = columns_result.columns
        data_start_row = header_row + columns_result.lineage.header_depth
        last_col_letter = get_column_letter(len(column_names))

        for row_idx in range(data_start_row, ws.max_row + 1):
            values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(column_names) + 1)]
            if all(value is None or not str(value).strip() for value in values):
                continue

            raw_by_column = {column_names[idx]: values[idx] for idx in range(len(column_names))}
            source_range = f"A{row_idx}:{last_col_letter}{row_idx}"
            cpf_value = _to_text(_find_value(raw_by_column, ("cpf",)))
            email_value = _to_text(_find_value(raw_by_column, ("email",)))
            yield {
                "evento": _to_text(_find_value(raw_by_column, ("evento",))),
                "sessao": _to_text(_find_value(raw_by_column, ("sessao", "session"))),
                "dt_hr_compra": _to_datetime(_find_value(raw_by_column, ("dt_hr_compra",))),
                "opt_in": _to_text(_find_value(raw_by_column, ("opt_in",))),
                "opt_in_id": _to_text(_find_value(raw_by_column, ("opt_in_id",))),
                "opt_in_status": _to_text(_find_value(raw_by_column, ("opt_in_status",))),
                "canal_venda": _to_text(_find_value(raw_by_column, ("canal_venda",))),
                "metodo_entrega": _to_text(_find_value(raw_by_column, ("metodo_entrega",))),
                "ingresso": _to_text(_find_value(raw_by_column, ("ingresso",))),
                "qtd_ingresso": _to_int(_find_value(raw_by_column, ("qtd_ingresso",))),
                "cpf_hash": hash_pii(cpf_value) if cpf_value else None,
                "email_hash": hash_pii(email_value) if email_value else None,
                "__sheet_name": ws.title,
                "__header_row": columns_result.lineage.header_row,
                "__header_range": columns_result.lineage.header_range,
                "__used_range": columns_result.lineage.used_range,
                "__source_range": source_range,
                "__row_number": row_idx,
                "__raw_payload": {
                    key: (value.isoformat() if isinstance(value, datetime) else value)
                    for key, value in raw_by_column.items()
                },
            }


def load_optin_xlsx_to_staging(
    *,
    source_id: str,
    xlsx_path: str | Path,
    event_id: int | None = None,
    sheet_name: str | None = None,
    lineage_policy: str = "required",
    severity_on_missing: str = "partial",
) -> dict[str, Any]:
    """Load extracted opt-in rows into staging table with ingestion + lineage.

    Args:
        source_id: Stable source identifier.
        xlsx_path: XLSX file path.
        event_id: Optional event identifier used during session resolution.
        sheet_name: Optional single-sheet selection.
        lineage_policy: Lineage policy (`required` or `optional`).
        severity_on_missing: Status on missing lineage (`partial` or `failed`).

    Returns:
        Summary with source and ingestion identifiers and loaded row count.

    Raises:
        ValueError: If extraction/load validation fails.
        OSError: If file access fails.
    """

    with Session(engine) as session:
        source = register_source_from_path(session, source_id, Path(xlsx_path))
        run = start_ingestion_run(session, source.source_id, "extract_xlsx_optin")
        loaded_rows: list[StgOptinTransaction] = []
        session_resolution_findings: list[str] = []
        try:
            for extracted in extract_optin_xlsx(xlsx_path, sheet_name=sheet_name):
                range_start, range_end = str(extracted["__source_range"]).split(":")
                location_value = format_location("range", {"start": range_start, "end": range_end})
                evidence_text = (
                    f"sheet={extracted['__sheet_name']}; "
                    f"header_row={extracted['__header_row']}; "
                    f"header_range={extracted['__header_range']}"
                )
                lineage_ref = create_lineage_ref(
                    session,
                    source_id=source.source_id,
                    ingestion_id=run.id,
                    location_type="range",
                    location_value=location_value,
                    evidence_text=evidence_text,
                    is_aggregated_metric=False,
                )
                resolution = resolve_session_reference(
                    session,
                    event_id=event_id,
                    raw_session_fields={
                        "source_id": source.source_id,
                        "session_name": extracted.get("sessao"),
                        "sessao": extracted.get("sessao"),
                        "dt_hr_compra": extracted.get("dt_hr_compra"),
                        "evento": extracted.get("evento"),
                    },
                )
                row_model = StgOptinTransaction(
                    source_id=source.source_id,
                    ingestion_id=int(run.id),
                    lineage_ref_id=int(lineage_ref.id),
                    event_id=resolution.event_id,
                    session_id=resolution.session_id,
                    sheet_name=str(extracted["__sheet_name"]),
                    header_row=int(extracted["__header_row"]),
                    row_number=int(extracted["__row_number"]),
                    source_range=str(extracted["__source_range"]),
                    evento=extracted.get("evento"),
                    sessao=extracted.get("sessao"),
                    dt_hr_compra=extracted.get("dt_hr_compra"),
                    opt_in=extracted.get("opt_in"),
                    opt_in_id=extracted.get("opt_in_id"),
                    opt_in_status=extracted.get("opt_in_status"),
                    canal_venda=extracted.get("canal_venda"),
                    metodo_entrega=extracted.get("metodo_entrega"),
                    ingresso=extracted.get("ingresso"),
                    qtd_ingresso=extracted.get("qtd_ingresso"),
                    cpf_hash=extracted.get("cpf_hash"),
                    email_hash=extracted.get("email_hash"),
                    session_resolution_finding=(
                        resolution.finding.message if resolution.finding else None
                    ),
                    raw_payload_json=json.dumps(extracted.get("__raw_payload", {}), ensure_ascii=False),
                )
                loaded_rows.append(row_model)
                if resolution.finding:
                    session_resolution_findings.append(
                        (
                            f"row={row_model.row_number}; "
                            f"code={resolution.finding.code}; "
                            f"message={resolution.finding.message}"
                        )
                    )

            if loaded_rows:
                session.add_all(loaded_rows)
                session.commit()

            enforce_lineage_policy(
                session,
                ingestion_id=int(run.id),
                dataset_name="stg_optin_transactions",
                records=loaded_rows,
                policy=lineage_policy,
                severity_on_missing=severity_on_missing,
            )

            if session_resolution_findings:
                finding_summary = " | ".join(session_resolution_findings[:10])
                finish_ingestion_run(
                    session,
                    ingestion_id=int(run.id),
                    status=IngestionStatus.PARTIAL,
                    notes=(
                        f"{len(loaded_rows)} registros carregados em stg_optin_transactions; "
                        f"session_resolution_findings={len(session_resolution_findings)}; "
                        f"detalhes={finding_summary}"
                    ),
                )
                return {
                    "source_id": source.source_id,
                    "ingestion_id": int(run.id),
                    "rows_loaded": len(loaded_rows),
                    "findings_count": len(session_resolution_findings),
                    "status": IngestionStatus.PARTIAL.value,
                }

            finish_ingestion_run(
                session,
                ingestion_id=int(run.id),
                status=IngestionStatus.SUCCESS,
                notes=f"{len(loaded_rows)} registros carregados em stg_optin_transactions.",
            )
            return {
                "source_id": source.source_id,
                "ingestion_id": int(run.id),
                "rows_loaded": len(loaded_rows),
                "findings_count": 0,
                "status": IngestionStatus.SUCCESS.value,
            }
        except Exception as exc:
            persisted_run = session.get(type(run), run.id)
            if persisted_run is not None and persisted_run.finished_at is None:
                finish_ingestion_run(
                    session,
                    ingestion_id=int(run.id),
                    status=IngestionStatus.FAILED,
                    notes=f"Falha no load stg_optin_transactions: {exc}",
                )
            raise


def main(argv: list[str] | None = None) -> int:
    """CLI entrypoint for XLSX opt-in extraction and staging load."""

    parser = argparse.ArgumentParser(prog="xlsx-optin")
    parser.add_argument("--source-id", required=True, help="Stable source identifier.")
    parser.add_argument("--xlsx", required=True, help="Path to source XLSX file.")
    parser.add_argument(
        "--event-id",
        type=int,
        default=None,
        help="Optional event identifier used for session resolution.",
    )
    parser.add_argument("--sheet", default="", help="Optional sheet name.")
    parser.add_argument(
        "--lineage-policy",
        default="required",
        choices=("required", "optional"),
        help="Lineage enforcement policy for staging load.",
    )
    parser.add_argument(
        "--severity-on-missing",
        default="partial",
        choices=("partial", "failed"),
        help="Final ingestion status when required lineage is missing.",
    )
    args = parser.parse_args(argv)

    result = load_optin_xlsx_to_staging(
        source_id=args.source_id,
        xlsx_path=args.xlsx,
        event_id=args.event_id,
        sheet_name=args.sheet or None,
        lineage_policy=args.lineage_policy,
        severity_on_missing=args.severity_on_missing,
    )
    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
