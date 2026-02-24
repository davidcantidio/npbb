"""XLSX leads extractor with normalized actions and staging load.

This module provides:
- `extract_leads_xlsx(path) -> iterator[dict]` for deterministic extraction,
- `parse_actions(value)` for normalized actions parsing,
- `load_leads_xlsx_to_staging(...)` for staging persistence with lineage.
"""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
import json
from pathlib import Path
import re
import sys
from typing import Any, Iterator, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlmodel import Session

try:  # package execution (repo root at `npbb`)
    from etl.extract.common import hash_pii
    from etl.extract.xlsx_utils import build_columns_with_metadata, find_header_row
except ModuleNotFoundError:  # pragma: no cover - fallback for `npbb.etl.*` execution
    from npbb.etl.extract.common import hash_pii
    from npbb.etl.extract.xlsx_utils import build_columns_with_metadata, find_header_row

try:  # pragma: no cover - import style depends on caller cwd
    from core.registry.location_ref import format_location
except ModuleNotFoundError:  # pragma: no cover
    from npbb.core.registry.location_ref import format_location

try:  # pragma: no cover
    from app.db.database import engine
    from app.models.etl_registry import IngestionStatus
    from app.models.stg_leads import StgLead, StgLeadAction
    from app.services.etl_lineage_service import create_lineage_ref
    from app.services.etl_registry_service import (
        enforce_lineage_policy,
        finish_ingestion_run,
        register_source_from_path,
        start_ingestion_run,
    )
except ModuleNotFoundError:  # pragma: no cover
    BACKEND_ROOT = Path(__file__).resolve().parents[2] / "backend"
    if str(BACKEND_ROOT) not in sys.path:
        sys.path.insert(0, str(BACKEND_ROOT))
    from app.db.database import engine
    from app.models.etl_registry import IngestionStatus
    from app.models.stg_leads import StgLead, StgLeadAction
    from app.services.etl_lineage_service import create_lineage_ref
    from app.services.etl_registry_service import (
        enforce_lineage_policy,
        finish_ingestion_run,
        register_source_from_path,
        start_ingestion_run,
    )


_DEFAULT_REQUIRED_TERMS: tuple[str, ...] = ("CPF", "Email", "Evento")
_DEFAULT_ALIASES: dict[str, str] = {
    "nome": "nome",
    "sobrenome": "sobrenome",
    "cpf": "cpf",
    "email": "email",
    "lead_nome": "nome",
    "lead_sobrenome": "sobrenome",
    "lead_cpf": "cpf",
    "lead_email": "email",
    "sexo": "sexo",
    "estado": "estado",
    "cidade": "cidade",
    "evento": "evento",
    "evento_evento": "evento",
    "datacriacao": "data_criacao",
    "data_criacao": "data_criacao",
    "evento_datacriacao": "data_criacao",
    "evento_data_criacao": "data_criacao",
    "acoes": "acoes",
    "evento_acoes": "acoes",
    "interesses": "interesses",
    "evento_interesses": "interesses",
    "areaatuacao": "area_atuacao",
    "area_atuacao": "area_atuacao",
    "evento_areaatuacao": "area_atuacao",
    "evento_area_atuacao": "area_atuacao",
    "evento_estado": "estado",
    "evento_cidade": "cidade",
    "evento_sexo": "sexo",
    "cpfpromotor": "cpf_promotor",
    "cpf_promotor": "cpf_promotor",
    "evento_cpfpromotor": "cpf_promotor",
    "evento_cpf_promotor": "cpf_promotor",
    "nomepromotor": "nome_promotor",
    "nome_promotor": "nome_promotor",
    "evento_nomepromotor": "nome_promotor",
    "evento_nome_promotor": "nome_promotor",
}

_ACTIONS_SPLIT_RE = re.compile(r"[;|,\n•]+")
_ACTIONS_IGNORE = {"-", "n/a", "na", "nenhuma", "none"}


def _to_text(value: Any) -> str | None:
    """Return trimmed text value or `None` for empty cells."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_datetime(value: Any) -> datetime | None:
    """Parse date/time values to timezone-aware datetime when possible."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(" ", "T")
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)


def _find_value(row: Mapping[str, Any], keys: Sequence[str]) -> Any:
    """Return first non-empty value from candidate keys in one extracted row."""
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return value
    return None


def parse_actions(value: Any) -> list[str]:
    """Parse and normalize actions string into deterministic ordered list.

    Supported delimiters:
    - semicolon (`;`)
    - pipe (`|`)
    - comma (`,`)
    - line break (`\\n`)
    - bullet (`•`)

    Normalization rules:
    - trim and collapse whitespace,
    - ignore empty tokens and known placeholders (`N/A`, `nenhuma`, `-`),
    - deduplicate preserving first occurrence (case-insensitive).

    Args:
        value: Raw actions cell value.

    Returns:
        Deterministic list of normalized action labels.
    """

    text = _to_text(value)
    if not text:
        return []

    tokens = _ACTIONS_SPLIT_RE.split(text)
    seen: set[str] = set()
    out: list[str] = []
    for token in tokens:
        cleaned = re.sub(r"\s+", " ", token).strip()
        if not cleaned:
            continue
        if cleaned.casefold() in _ACTIONS_IGNORE:
            continue
        key = cleaned.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(cleaned)
    return out


def extract_leads_xlsx(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    aliases: Mapping[str, str] | None = None,
) -> Iterator[dict[str, Any]]:
    """Extract lead rows from XLSX with parsed actions and lineage metadata.

    Args:
        path: XLSX file path.
        sheet_name: Optional single-sheet selection.
        aliases: Optional normalized-column aliases.

    Yields:
        Extracted row dictionaries with canonical fields plus lineage metadata keys:
        `__sheet_name`, `__header_row`, `__header_range`, `__used_range`,
        `__source_range`.
    """

    xlsx_path = Path(path).expanduser()
    workbook = load_workbook(xlsx_path, data_only=True)
    sheets = [workbook[sheet_name]] if sheet_name else [workbook[name] for name in workbook.sheetnames]
    alias_map = dict(_DEFAULT_ALIASES)
    if aliases:
        alias_map.update(aliases)

    for ws in sheets:
        header_row = find_header_row(ws, required_terms=_DEFAULT_REQUIRED_TERMS, max_scan_rows=40)
        columns_result = build_columns_with_metadata(
            ws,
            header_row=header_row,
            header_depth=2,
            aliases=alias_map,
        )
        column_names = columns_result.columns
        data_start_row = header_row + columns_result.lineage.header_depth
        last_col_letter = get_column_letter(len(column_names))

        for row_idx in range(data_start_row, ws.max_row + 1):
            values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(column_names) + 1)]
            if all(value is None or not str(value).strip() for value in values):
                continue

            raw_by_column = {column_names[idx]: values[idx] for idx in range(len(column_names))}
            source_range = f"A{row_idx}:{last_col_letter}{row_idx}"

            cpf_value = _to_text(_find_value(raw_by_column, ("cpf",)))
            email_value = _to_text(_find_value(raw_by_column, ("email",)))
            cpf_promotor = _to_text(_find_value(raw_by_column, ("cpf_promotor",)))
            actions_raw = _to_text(_find_value(raw_by_column, ("acoes",)))
            actions_list = parse_actions(actions_raw)

            yield {
                "evento": _to_text(_find_value(raw_by_column, ("evento",))),
                "data_criacao": _to_datetime(_find_value(raw_by_column, ("data_criacao",))),
                "sexo": _to_text(_find_value(raw_by_column, ("sexo",))),
                "estado": _to_text(_find_value(raw_by_column, ("estado",))),
                "cidade": _to_text(_find_value(raw_by_column, ("cidade",))),
                "interesses": _to_text(_find_value(raw_by_column, ("interesses",))),
                "area_atuacao": _to_text(_find_value(raw_by_column, ("area_atuacao",))),
                "cpf_hash": hash_pii(cpf_value) if cpf_value else None,
                "email_hash": hash_pii(email_value) if email_value else None,
                "person_key_hash": hash_pii(cpf_value) if cpf_value else (hash_pii(email_value) if email_value else None),
                "cpf_promotor_hash": hash_pii(cpf_promotor) if cpf_promotor else None,
                "nome_promotor": _to_text(_find_value(raw_by_column, ("nome_promotor",))),
                "acoes_raw": actions_raw,
                "acoes_list": actions_list,
                "__sheet_name": ws.title,
                "__header_row": columns_result.lineage.header_row,
                "__header_range": columns_result.lineage.header_range,
                "__used_range": columns_result.lineage.used_range,
                "__source_range": source_range,
                "__row_number": row_idx,
                "__raw_payload": {
                    key: (value.isoformat() if isinstance(value, datetime) else value)
                    for key, value in raw_by_column.items()
                },
            }


def load_leads_xlsx_to_staging(
    *,
    source_id: str,
    xlsx_path: str | Path,
    sheet_name: str | None = None,
    lineage_policy: str = "required",
    severity_on_missing: str = "partial",
) -> dict[str, Any]:
    """Load extracted leads and parsed actions into staging with lineage.

    Args:
        source_id: Stable source identifier.
        xlsx_path: XLSX file path.
        sheet_name: Optional single-sheet selection.
        lineage_policy: Lineage policy (`required` or `optional`).
        severity_on_missing: Status on missing lineage (`partial` or `failed`).

    Returns:
        Summary with source and ingestion identifiers and loaded counts.

    Raises:
        ValueError: If extraction/load validation fails.
        OSError: If file access fails.
    """

    with Session(engine) as session:
        source = register_source_from_path(session, source_id, Path(xlsx_path))
        run = start_ingestion_run(session, source.source_id, "extract_xlsx_leads")
        lead_models: list[StgLead] = []
        action_models: list[StgLeadAction] = []
        action_count = 0
        dedupe_keys: set[tuple[str, str, str, str]] = set()

        try:
            for extracted in extract_leads_xlsx(xlsx_path, sheet_name=sheet_name):
                dedupe_key = (
                    extracted.get("cpf_hash") or "",
                    extracted.get("email_hash") or "",
                    extracted.get("evento") or "",
                    extracted.get("data_criacao").isoformat() if extracted.get("data_criacao") else "",
                )
                if dedupe_key in dedupe_keys and any(dedupe_key[:2]):
                    continue
                dedupe_keys.add(dedupe_key)

                range_start, range_end = str(extracted["__source_range"]).split(":")
                location_value = format_location("range", {"start": range_start, "end": range_end})
                evidence_text = (
                    f"sheet={extracted['__sheet_name']}; "
                    f"header_row={extracted['__header_row']}; "
                    f"header_range={extracted['__header_range']}"
                )
                lineage_ref = create_lineage_ref(
                    session,
                    source_id=source.source_id,
                    ingestion_id=run.id,
                    location_type="range",
                    location_value=location_value,
                    evidence_text=evidence_text,
                    is_aggregated_metric=False,
                )

                lead = StgLead(
                    source_id=source.source_id,
                    ingestion_id=int(run.id),
                    lineage_ref_id=int(lineage_ref.id),
                    sheet_name=str(extracted["__sheet_name"]),
                    header_row=int(extracted["__header_row"]),
                    row_number=int(extracted["__row_number"]),
                    source_range=str(extracted["__source_range"]),
                    evento=extracted.get("evento"),
                    data_criacao=extracted.get("data_criacao"),
                    sexo=extracted.get("sexo"),
                    estado=extracted.get("estado"),
                    cidade=extracted.get("cidade"),
                    interesses=extracted.get("interesses"),
                    area_atuacao=extracted.get("area_atuacao"),
                    cpf_hash=extracted.get("cpf_hash"),
                    email_hash=extracted.get("email_hash"),
                    person_key_hash=extracted.get("person_key_hash"),
                    cpf_promotor_hash=extracted.get("cpf_promotor_hash"),
                    nome_promotor=extracted.get("nome_promotor"),
                    acoes_raw=extracted.get("acoes_raw"),
                    raw_payload_json=json.dumps(extracted.get("__raw_payload", {}), ensure_ascii=False),
                )
                session.add(lead)
                session.flush()
                lead_models.append(lead)

                for index, action in enumerate(extracted.get("acoes_list", []), start=1):
                    action_model = StgLeadAction(
                        lead_id=int(lead.id),
                        source_id=source.source_id,
                        ingestion_id=int(run.id),
                        lineage_ref_id=int(lineage_ref.id),
                        action_order=index,
                        action_raw=action,
                        action_norm=action.casefold(),
                    )
                    session.add(action_model)
                    action_models.append(action_model)
                    action_count += 1

            session.commit()

            enforce_lineage_policy(
                session,
                ingestion_id=int(run.id),
                dataset_name="stg_leads",
                records=lead_models,
                policy=lineage_policy,
                severity_on_missing=severity_on_missing,
            )
            enforce_lineage_policy(
                session,
                ingestion_id=int(run.id),
                dataset_name="stg_lead_actions",
                records=action_models,
                policy=lineage_policy,
                severity_on_missing=severity_on_missing,
            )

            finish_ingestion_run(
                session,
                ingestion_id=int(run.id),
                status=IngestionStatus.SUCCESS,
                notes=(
                    f"{len(lead_models)} leads e {action_count} acoes carregadas "
                    "em staging."
                ),
            )
            return {
                "source_id": source.source_id,
                "ingestion_id": int(run.id),
                "leads_loaded": len(lead_models),
                "actions_loaded": action_count,
            }
        except Exception as exc:
            persisted_run = session.get(type(run), run.id)
            if persisted_run is not None and persisted_run.finished_at is None:
                finish_ingestion_run(
                    session,
                    ingestion_id=int(run.id),
                    status=IngestionStatus.FAILED,
                    notes=f"Falha no load stg_leads: {exc}",
                )
            raise


def main(argv: list[str] | None = None) -> int:
    """CLI entrypoint for leads extraction and staging load."""

    parser = argparse.ArgumentParser(prog="xlsx-leads")
    parser.add_argument("--source-id", required=True, help="Stable source identifier.")
    parser.add_argument("--xlsx", required=True, help="Path to source XLSX file.")
    parser.add_argument("--sheet", default="", help="Optional sheet name.")
    parser.add_argument(
        "--lineage-policy",
        default="required",
        choices=("required", "optional"),
        help="Lineage enforcement policy for staging load.",
    )
    parser.add_argument(
        "--severity-on-missing",
        default="partial",
        choices=("partial", "failed"),
        help="Final ingestion status when required lineage is missing.",
    )
    args = parser.parse_args(argv)

    result = load_leads_xlsx_to_staging(
        source_id=args.source_id,
        xlsx_path=args.xlsx,
        sheet_name=args.sheet or None,
        lineage_policy=args.lineage_policy,
        severity_on_missing=args.severity_on_missing,
    )
    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
