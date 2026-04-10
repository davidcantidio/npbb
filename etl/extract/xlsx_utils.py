"""Generic XLSX utilities for header detection and column canonicalization."""

from __future__ import annotations

from dataclasses import dataclass
import re
from typing import Mapping, Sequence

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

try:  # package execution (repo root at `npbb`)
    from etl.transform.column_normalize import (
        ColumnNormalizationConfig,
        normalize_column_names,
    )
except ModuleNotFoundError:  # pragma: no cover - fallback for `npbb.etl.*` imports
    from npbb.etl.transform.column_normalize import (
        ColumnNormalizationConfig,
        normalize_column_names,
    )


_SPACE_RE = re.compile(r"\s+")


@dataclass(frozen=True)
class XlsxLineageMetadata:
    """Lineage metadata captured during XLSX header extraction.

    Attributes:
        sheet_name: Workbook sheet title.
        header_row: First row used as header anchor.
        header_depth: Number of rows consumed for header composition.
        header_range: A1 range that contains header cells.
        used_range: A1 range from header row to sheet max row within used columns.
    """

    sheet_name: str
    header_row: int
    header_depth: int
    header_range: str
    used_range: str


@dataclass(frozen=True)
class XlsxColumnsResult:
    """Canonical columns and lineage metadata for one sheet header extraction."""

    columns: list[str]
    lineage: XlsxLineageMetadata


@dataclass(frozen=True)
class XlsxHeaderColumn:
    """Cleaned value found in one candidate XLSX header cell."""

    column_index: int
    column_letter: str
    source_value: str


@dataclass(frozen=True)
class HeaderNotFound:
    """Soft-failure result returned when header matching is inconclusive."""

    message: str
    required_terms: tuple[str, ...]
    scanned_rows: int
    forced_row: int | None = None
    columns: tuple[XlsxHeaderColumn, ...] = ()


def _clean_text(value: object) -> str:
    """Convert any cell value into a compact one-line string."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    return _SPACE_RE.sub(" ", text)


def _cell_value_with_merged_fallback(ws: Worksheet, row: int, col: int) -> object:
    """Return cell value, falling back to merged-cell anchor when needed."""
    value = ws.cell(row=row, column=col).value
    if _clean_text(value):
        return value
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return ws.cell(row=merged.min_row, column=merged.min_col).value
    return value


def _row_has_horizontal_merge(ws: Worksheet, row: int) -> bool:
    """Return True when row intersects at least one horizontal merged range."""
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.max_col > merged.min_col:
            return True
    return False


def _row_non_empty_values(ws: Worksheet, row: int, max_col: int) -> list[str]:
    """Return non-empty cleaned values for one row (with merged fallback)."""
    values: list[str] = []
    for col in range(1, max_col + 1):
        cleaned = _clean_text(_cell_value_with_merged_fallback(ws, row, col))
        if cleaned:
            values.append(cleaned)
    return values


def _row_header_columns(ws: Worksheet, row: int, max_col: int) -> tuple[XlsxHeaderColumn, ...]:
    """Return cleaned non-empty header cells for one row with coordinates."""
    columns: list[XlsxHeaderColumn] = []
    for col in range(1, max_col + 1):
        cleaned = _clean_text(_cell_value_with_merged_fallback(ws, row, col))
        if not cleaned:
            continue
        columns.append(
            XlsxHeaderColumn(
                column_index=col,
                column_letter=get_column_letter(col),
                source_value=cleaned,
            )
        )
    return tuple(columns)


def _required_term_candidates(
    required_terms: Sequence[str],
    term_aliases: Mapping[str, Sequence[str]] | None,
) -> list[tuple[str, ...]]:
    """Build per-term match candidates without mutating the existing alias sources."""
    alias_map = term_aliases or {}
    candidates: list[tuple[str, ...]] = []
    for term in required_terms:
        canonical = term.strip()
        if not canonical:
            continue
        values = [canonical]
        for key in (canonical, canonical.lower(), canonical.upper()):
            values.extend(alias_map.get(key, ()))
        normalized = tuple(
            dict.fromkeys(value.strip().lower() for value in values if value and value.strip())
        )
        if normalized:
            candidates.append(normalized)
    return candidates


def _row_matches_required_terms(values: Sequence[str], term_candidates: Sequence[Sequence[str]]) -> bool:
    """Return True when each required term has at least one matching cell."""
    if not term_candidates:
        return False
    lowered = [value.lower() for value in values]
    return all(any(candidate in cell for cell in lowered for candidate in candidates) for candidates in term_candidates)


def _infer_last_used_column(ws: Worksheet, *, header_row: int, header_depth: int) -> int:
    """Infer last used column based on header rows, falling back to sheet max column."""
    last_col = 0
    max_col = ws.max_column
    header_last_row = min(ws.max_row, header_row + max(header_depth - 1, 0))
    for row in range(header_row, header_last_row + 1):
        for col in range(1, max_col + 1):
            if _clean_text(_cell_value_with_merged_fallback(ws, row, col)):
                last_col = max(last_col, col)
    if last_col <= 0:
        return max_col
    return last_col


def find_header_row(
    ws: Worksheet,
    *,
    required_terms: Sequence[str] | None = None,
    term_aliases: Mapping[str, Sequence[str]] | None = None,
    forced_row: int | None = None,
    soft_fail: bool = False,
    promote_merged_header: bool = True,
    max_scan_rows: int = 30,
    min_non_empty_cells: int = 2,
) -> int | HeaderNotFound:
    """Find the most likely header row in one worksheet.

    Heuristics:
    - if `required_terms` are provided, pick the first row where all terms appear;
    - otherwise choose the highest-scoring text-like row in the scan window;
    - if the previous row contains merged group headers, promote to that row.

    Args:
        ws: OpenPyXL worksheet object.
        required_terms: Optional terms that must appear in header row cells.
        term_aliases: Optional alias candidates keyed by required term.
        forced_row: Optional 1-based row selected by the caller for validation.
        soft_fail: If True, return HeaderNotFound instead of raising ValueError.
        promote_merged_header: If True, keep legacy promotion to previous merged row.
        max_scan_rows: Maximum number of top rows scanned.
        min_non_empty_cells: Minimum non-empty cells to consider a row a header candidate.

    Returns:
        1-based header row index, or HeaderNotFound when `soft_fail=True`.

    Raises:
        ValueError: If no candidate header row is found.
    """

    scan_limit = min(max_scan_rows, ws.max_row)
    terms = tuple(term.strip() for term in (required_terms or []) if term and term.strip())
    term_candidates = _required_term_candidates(terms, term_aliases)

    if forced_row is not None:
        if forced_row < 1 or forced_row > ws.max_row:
            not_found = HeaderNotFound(
                message=f"Forced header row is outside worksheet bounds: {forced_row}.",
                required_terms=terms,
                scanned_rows=scan_limit,
                forced_row=forced_row,
            )
            if soft_fail:
                return not_found
            raise ValueError(not_found.message)
        values = _row_non_empty_values(ws, forced_row, ws.max_column)
        if term_candidates and not _row_matches_required_terms(values, term_candidates):
            not_found = HeaderNotFound(
                message="Forced header row does not contain the required terms.",
                required_terms=terms,
                scanned_rows=scan_limit,
                forced_row=forced_row,
                columns=_row_header_columns(ws, forced_row, ws.max_column),
            )
            if soft_fail:
                return not_found
            raise ValueError(not_found.message)
        if len(values) < min_non_empty_cells:
            not_found = HeaderNotFound(
                message="Forced header row does not contain enough non-empty cells.",
                required_terms=terms,
                scanned_rows=scan_limit,
                forced_row=forced_row,
                columns=_row_header_columns(ws, forced_row, ws.max_column),
            )
            if soft_fail:
                return not_found
            raise ValueError(not_found.message)
        return forced_row

    best_row = 0
    best_score = -10_000
    required_match_row = 0

    for row in range(1, scan_limit + 1):
        values = _row_non_empty_values(ws, row, ws.max_column)
        if len(values) < min_non_empty_cells:
            continue

        if term_candidates and _row_matches_required_terms(values, term_candidates):
            required_match_row = row
            break

        text_like = sum(1 for value in values if any(ch.isalpha() for ch in value))
        numeric_like = len(values) - text_like
        score = len(values) + (2 * text_like) - numeric_like
        if score > best_score:
            best_score = score
            best_row = row

    if term_candidates and required_match_row <= 0 and soft_fail:
        return HeaderNotFound(
            message="Header row could not be identified with the required terms.",
            required_terms=terms,
            scanned_rows=scan_limit,
        )

    header_row = required_match_row or best_row
    if header_row <= 0:
        if soft_fail:
            return HeaderNotFound(
                message="Header row could not be identified in worksheet.",
                required_terms=terms,
                scanned_rows=scan_limit,
            )
        raise ValueError("Header row could not be identified in worksheet.")

    previous_row = header_row - 1
    if promote_merged_header and previous_row >= 1 and _row_has_horizontal_merge(ws, previous_row):
        previous_values = _row_non_empty_values(ws, previous_row, ws.max_column)
        if previous_values:
            return previous_row

    return header_row


def build_columns(
    ws: Worksheet,
    header_row: int,
    *,
    header_depth: int = 2,
    normalization_config: ColumnNormalizationConfig | None = None,
    aliases: Mapping[str, str] | None = None,
) -> list[str]:
    """Build canonical column names from one XLSX header block.

    Args:
        ws: OpenPyXL worksheet object.
        header_row: First row of header block.
        header_depth: Number of rows combined to build one header token.
        normalization_config: Optional normalization configuration.
        aliases: Optional normalized-name alias mapping.

    Returns:
        Canonical column list.
    """

    if header_row <= 0:
        raise ValueError("header_row must be a positive 1-based row index.")
    if header_depth <= 0:
        raise ValueError("header_depth must be a positive integer.")

    last_col = _infer_last_used_column(ws, header_row=header_row, header_depth=header_depth)
    raw_columns: list[str] = []

    header_last_row = min(ws.max_row, header_row + header_depth - 1)
    for col in range(1, last_col + 1):
        parts: list[str] = []
        for row in range(header_row, header_last_row + 1):
            text = _clean_text(_cell_value_with_merged_fallback(ws, row, col))
            if text and (not parts or parts[-1].lower() != text.lower()):
                parts.append(text)
        raw_columns.append(" ".join(parts) if parts else f"col_{col}")

    return normalize_column_names(raw_columns, config=normalization_config, aliases=aliases)


def build_columns_with_metadata(
    ws: Worksheet,
    header_row: int,
    *,
    header_depth: int = 2,
    normalization_config: ColumnNormalizationConfig | None = None,
    aliases: Mapping[str, str] | None = None,
) -> XlsxColumnsResult:
    """Build canonical columns and include lineage metadata for auditing.

    Args:
        ws: OpenPyXL worksheet object.
        header_row: First row of header block.
        header_depth: Number of rows combined to build one header token.
        normalization_config: Optional normalization configuration.
        aliases: Optional normalized-name alias mapping.

    Returns:
        Canonical columns plus lineage metadata (sheet, header row and range used).
    """

    columns = build_columns(
        ws,
        header_row,
        header_depth=header_depth,
        normalization_config=normalization_config,
        aliases=aliases,
    )
    last_col = len(columns)
    last_col_letter = get_column_letter(last_col)
    header_last_row = min(ws.max_row, header_row + max(header_depth - 1, 0))
    lineage = XlsxLineageMetadata(
        sheet_name=ws.title,
        header_row=header_row,
        header_depth=header_depth,
        header_range=f"A{header_row}:{last_col_letter}{header_last_row}",
        used_range=f"A{header_row}:{last_col_letter}{ws.max_row}",
    )
    return XlsxColumnsResult(columns=columns, lineage=lineage)
