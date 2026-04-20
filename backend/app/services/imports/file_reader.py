"""Leitura padrao de arquivos de importacao (CSV/XLSX)."""

from __future__ import annotations

import codecs
import csv
import io
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO, Iterator

from fastapi import UploadFile
from openpyxl import load_workbook

from app.services.imports.contracts import ImportPreviewResult

ALLOWED_IMPORT_EXTENSIONS = {".csv", ".xlsx"}
DEFAULT_IMPORT_MAX_BYTES = 50 * 1024 * 1024
BYTES_PER_MEGABYTE = 1024 * 1024
HEADER_SCAN_LIMIT = 50
CSV_STREAM_SNIFF_BYTES = 8192
CSV_STREAM_UTF8_VALIDATION_CHUNK_BYTES = 64 * 1024
XLSX_REQUIRED_ARCHIVE_MEMBERS = frozenset({"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"})


@dataclass(frozen=True)
class ImportFileError(Exception):
    code: str
    message: str
    field: str = "file"
    extra: dict | None = None


def _probe_starts_with_zip_local_header(probe: bytes) -> bool:
    """True if bytes start like a ZIP local header (OOXML .xlsx is ZIP-based)."""
    if len(probe) < 4:
        return False
    return probe[:2] == b"PK" and probe[2:4] in (b"\x03\x04", b"\x05\x06", b"\x07\x08")


def _validate_import_content_probe(ext: str, probe: bytes) -> None:
    """Reject extension/content mismatch early (magic bytes, light CSV checks)."""
    if ext == ".xlsx":
        if len(probe) < 4:
            raise ImportFileError(
                code="INVALID_FILE_CONTENT",
                message="Arquivo muito curto ou corrompido para ser XLSX.",
                field="file",
            )
        if not _probe_starts_with_zip_local_header(probe):
            raise ImportFileError(
                code="INVALID_FILE_CONTENT",
                message="O conteudo do arquivo nao corresponde a uma planilha XLSX (Office Open XML).",
                field="file",
            )
        return

    if ext == ".csv":
        if b"\x00" in probe:
            raise ImportFileError(
                code="INVALID_FILE_CONTENT",
                message="O conteudo do arquivo nao corresponde a um CSV de texto (dados binarios detectados).",
                field="file",
            )
        if _probe_starts_with_zip_local_header(probe):
            raise ImportFileError(
                code="INVALID_FILE_CONTENT",
                message="O conteudo parece planilha compactada (ZIP/XLSX), mas a extensao e .csv.",
                field="file",
            )


def _validate_xlsx_archive_structure(file_obj: BinaryIO) -> None:
    """Ensure the ZIP structure matches the minimum OOXML workbook layout."""
    start_pos = file_obj.tell()
    try:
        file_obj.seek(0)
        try:
            with zipfile.ZipFile(file_obj) as archive:
                names = set(archive.namelist())
        except zipfile.BadZipFile as exc:
            raise ImportFileError(
                code="INVALID_FILE_CONTENT",
                message="O conteudo do arquivo nao corresponde a uma planilha XLSX valida.",
                field="file",
            ) from exc

        if not XLSX_REQUIRED_ARCHIVE_MEMBERS.issubset(names):
            raise ImportFileError(
                code="INVALID_FILE_CONTENT",
                message="O conteudo do arquivo nao corresponde a uma planilha XLSX valida.",
                field="file",
            )
    finally:
        file_obj.seek(start_pos)


def _detect_csv_delimiter(sample_text: str) -> str:
    first_line = sample_text.splitlines()[0] if sample_text else ""
    comma_count = first_line.count(",")
    semicolon_count = first_line.count(";")
    return ";" if semicolon_count > comma_count else ","


def _score_row_tabular(row: list[str]) -> int:
    if not row:
        return 0
    non_empty = [cell for cell in row if cell.strip()]
    fill_ratio = len(non_empty) / max(len(row), 1)
    if len(non_empty) >= 2 and fill_ratio >= 0.5:
        return int(fill_ratio * 100)
    return 0


def detect_data_start_index(rows: list[list[str]]) -> int:
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows[:50]):
        score = _score_row_tabular(row)
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx


def _decode_payload(raw: bytes) -> str:
    try:
        return raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        return raw.decode("latin-1")


def _detect_csv_stream_encoding(file_obj: io.BufferedIOBase) -> tuple[str, str]:
    """Valida UTF-8 incrementalmente e faz fallback para latin-1 sem materializar o upload."""
    start_pos = file_obj.tell()
    utf8_decoder = codecs.getincrementaldecoder("utf-8-sig")()
    sniff_bytes = bytearray()
    try:
        while True:
            chunk = file_obj.read(CSV_STREAM_UTF8_VALIDATION_CHUNK_BYTES)
            if not chunk:
                break
            if len(sniff_bytes) < CSV_STREAM_SNIFF_BYTES:
                missing = CSV_STREAM_SNIFF_BYTES - len(sniff_bytes)
                sniff_bytes.extend(chunk[:missing])
            try:
                utf8_decoder.decode(chunk, final=False)
            except UnicodeDecodeError:
                return "latin-1", bytes(sniff_bytes).decode("latin-1")
        utf8_decoder.decode(b"", final=True)
        return "utf-8-sig", bytes(sniff_bytes).decode("utf-8-sig")
    finally:
        file_obj.seek(start_pos)


def _normalize_row(row: list[object] | tuple[object, ...]) -> list[str]:
    return [("" if cell is None else str(cell)).strip() for cell in row]


def _validate_extension(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_IMPORT_EXTENSIONS:
        accepted_extensions = ", ".join(sorted(ALLOWED_IMPORT_EXTENSIONS))
        raise ImportFileError(
            code="INVALID_FILE_TYPE",
            message=f"Arquivo '{ext}' nao e suportado. Extensoes aceitas: {accepted_extensions}",
            field="file",
        )
    return ext


def _build_preview_result(
    *,
    filename: str,
    rows: list[list[str]],
    delimiter: str | None,
    sheet_name: str | None = None,
) -> ImportPreviewResult:
    start_index = detect_data_start_index(rows)
    headers = rows[start_index] if rows else []
    data_rows: list[list[str]] = []
    physical_line_numbers: list[int] = []
    for idx, row in enumerate(rows[start_index + 1 :], start=start_index + 2):
        if any(cell.strip() for cell in row):
            data_rows.append(row)
            physical_line_numbers.append(idx)
    return ImportPreviewResult(
        filename=filename,
        headers=headers,
        rows=data_rows,
        delimiter=delimiter,
        start_index=start_index,
        sheet_name=sheet_name,
        physical_line_numbers=physical_line_numbers,
    )


def _read_preview_window_from_rows(
    rows_iter: Iterator[list[str]],
    *,
    filename: str,
    delimiter: str | None,
    sample_rows: int,
    sheet_name: str | None = None,
    scan_rows: int = HEADER_SCAN_LIMIT,
) -> ImportPreviewResult:
    buffered_rows: list[list[str]] = []
    for row in rows_iter:
        buffered_rows.append(row)
        if len(buffered_rows) >= scan_rows:
            break

    if not buffered_rows:
        return ImportPreviewResult(
            filename=filename,
            headers=[],
            rows=[],
            delimiter=delimiter,
            start_index=0,
            physical_line_numbers=[],
        )

    start_index = detect_data_start_index(buffered_rows)
    headers = buffered_rows[start_index] if start_index < len(buffered_rows) else []
    data_rows: list[list[str]] = []
    physical_line_numbers: list[int] = []
    for idx in range(start_index + 1, len(buffered_rows)):
        row = buffered_rows[idx]
        if any(cell.strip() for cell in row):
            data_rows.append(row)
            physical_line_numbers.append(idx + 1)

    next_row_idx = len(buffered_rows)
    while len(data_rows) < sample_rows:
        try:
            row = next(rows_iter)
        except StopIteration:
            break
        physical_here = next_row_idx + 1
        next_row_idx += 1
        if any(cell.strip() for cell in row):
            data_rows.append(row)
            physical_line_numbers.append(physical_here)

    return ImportPreviewResult(
        filename=filename,
        headers=headers,
        rows=data_rows[:sample_rows],
        delimiter=delimiter,
        start_index=start_index,
        sheet_name=sheet_name,
        physical_line_numbers=physical_line_numbers[:sample_rows],
    )


def _read_csv_rows_from_raw(raw: bytes) -> tuple[list[list[str]], str]:
    text = _decode_payload(raw)
    delimiter = _detect_csv_delimiter(text[:4096])
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [_normalize_row(row) for row in reader]
    return rows, delimiter


def _read_xlsx_rows_from_raw(raw: bytes) -> tuple[list[list[str]], str | None]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        return [_normalize_row(row) for row in ws.iter_rows(values_only=True)], ws.title
    finally:
        wb.close()


def _read_csv_preview_from_raw(raw: bytes, *, filename: str, sample_rows: int) -> ImportPreviewResult:
    text = _decode_payload(raw)
    delimiter = _detect_csv_delimiter(text[:4096])
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    normalized_rows = (_normalize_row(row) for row in reader)
    return _read_preview_window_from_rows(
        normalized_rows,
        filename=filename,
        delimiter=delimiter,
        sample_rows=sample_rows,
    )


def _read_xlsx_preview_from_raw(raw: bytes, *, filename: str, sample_rows: int) -> ImportPreviewResult:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        normalized_rows = (_normalize_row(row) for row in ws.iter_rows(values_only=True))
        return _read_preview_window_from_rows(
            normalized_rows,
            filename=filename,
            delimiter=None,
            sample_rows=sample_rows,
            sheet_name=ws.title,
        )
    finally:
        wb.close()


def read_fileobj_preview(
    file_obj: BinaryIO,
    *,
    filename: str,
    sample_rows: int,
) -> ImportPreviewResult:
    """Read a small preview window from an already-spooled upload."""
    ext = _validate_extension(filename)
    file_obj.seek(0)
    if ext == ".xlsx":
        wb = load_workbook(file_obj, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            normalized_rows = (_normalize_row(row) for row in ws.iter_rows(values_only=True))
            return _read_preview_window_from_rows(
                normalized_rows,
                filename=filename,
                delimiter=None,
                sample_rows=sample_rows,
                sheet_name=ws.title,
            )
        finally:
            wb.close()
            file_obj.seek(0)

    encoding, sample_text = _detect_csv_stream_encoding(file_obj)
    text_io = io.TextIOWrapper(file_obj, encoding=encoding, newline="")
    try:
        delimiter = _detect_csv_delimiter(sample_text[:4096] if sample_text else "")
        text_io.seek(0)
        reader = csv.reader(text_io, delimiter=delimiter)
        normalized_rows = (_normalize_row(row) for row in reader)
        return _read_preview_window_from_rows(
            normalized_rows,
            filename=filename,
            delimiter=delimiter,
            sample_rows=sample_rows,
        )
    finally:
        text_io.detach()
        file_obj.seek(0)


def inspect_upload(
    file: UploadFile,
    *,
    max_bytes: int = DEFAULT_IMPORT_MAX_BYTES,
) -> tuple[str, str, int]:
    filename = file.filename or ""
    ext = _validate_extension(filename)

    file.file.seek(0, 2)
    size = file.file.tell()
    file.file.seek(0)
    if size > max_bytes:
        raise ImportFileError(
            code="FILE_TOO_LARGE",
            message=(
                f"Arquivo muito grande: {size / BYTES_PER_MEGABYTE:.1f} MB. "
                f"Limite permitido: {max_bytes / BYTES_PER_MEGABYTE:.1f} MB"
            ),
            field="file",
            extra={"max_bytes": max_bytes},
        )
    if size == 0:
        raise ImportFileError(
            code="EMPTY_FILE",
            message="Arquivo vazio",
            field="file",
        )

    probe_len = min(size, CSV_STREAM_SNIFF_BYTES)
    file.file.seek(0)
    probe = file.file.read(probe_len)
    _validate_import_content_probe(ext, probe)
    if ext == ".xlsx":
        _validate_xlsx_archive_structure(file.file)
    file.file.seek(0)
    return filename, ext, size


def read_raw_file_rows(raw: bytes, *, filename: str) -> ImportPreviewResult:
    ext = _validate_extension(filename)
    if ext == ".xlsx":
        rows, sheet_name = _read_xlsx_rows_from_raw(raw)
        return _build_preview_result(
            filename=filename,
            rows=rows,
            delimiter=None,
            sheet_name=sheet_name,
        )

    rows, delimiter = _read_csv_rows_from_raw(raw)
    return _build_preview_result(filename=filename, rows=rows, delimiter=delimiter)


def read_raw_file_headers(raw: bytes, *, filename: str) -> ImportPreviewResult:
    return read_raw_file_preview(raw, filename=filename, sample_rows=0)


def read_raw_file_preview(raw: bytes, *, filename: str, sample_rows: int) -> ImportPreviewResult:
    ext = _validate_extension(filename)
    if ext == ".xlsx":
        return _read_xlsx_preview_from_raw(raw, filename=filename, sample_rows=sample_rows)
    return _read_csv_preview_from_raw(raw, filename=filename, sample_rows=sample_rows)


def _read_csv_sample(file: UploadFile, *, max_rows: int) -> ImportPreviewResult:
    file.file.seek(0)
    raw = file.file.read()
    return read_raw_file_preview(raw, filename=file.filename or "", sample_rows=max_rows)


def _read_xlsx_sample(file: UploadFile, *, max_rows: int) -> ImportPreviewResult:
    file.file.seek(0)
    raw = file.file.read()
    return read_raw_file_preview(raw, filename=file.filename or "", sample_rows=max_rows)


def read_file_sample(
    file: UploadFile,
    *,
    sample_rows: int,
    max_bytes: int = DEFAULT_IMPORT_MAX_BYTES,
) -> tuple[ImportPreviewResult, str]:
    if sample_rows < 1 or sample_rows > 50:
        raise ImportFileError(
            code="INVALID_SAMPLE_SIZE",
            message="sample_rows deve ser entre 1 e 50",
            field="sample_rows",
        )

    _, ext, _ = inspect_upload(file, max_bytes=max_bytes)
    preview = _read_xlsx_sample(file, max_rows=sample_rows) if ext == ".xlsx" else _read_csv_sample(
        file, max_rows=sample_rows
    )
    if not preview.headers and not preview.rows:
        raise ImportFileError(code="EMPTY_FILE", message="Arquivo vazio", field="file")
    return preview, ext


def iter_data_rows(
    file: UploadFile,
    *,
    ext: str,
    start_index: int,
    delimiter: str | None = None,
) -> Iterator[list[str]]:
    if ext == ".xlsx":
        file.file.seek(0)
        wb = load_workbook(file.file, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx <= start_index:
                continue
            yield [("" if v is None else str(v)).strip() for v in row]
        return

    file.file.seek(0)
    encoding, sample_text = _detect_csv_stream_encoding(file.file)
    text_io = io.TextIOWrapper(file.file, encoding=encoding, newline="")
    try:
        active_delimiter = delimiter or _detect_csv_delimiter(sample_text[:4096] if sample_text else "")
        text_io.seek(0)
        reader = csv.reader(text_io, delimiter=active_delimiter)
        for idx, row in enumerate(reader):
            if idx <= start_index:
                continue
            yield [str(cell).strip() for cell in row]
    finally:
        text_io.detach()
