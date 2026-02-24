"""Tests for XLSX opt-in extractor and staging loader with lineage."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
import sys

import pytest
from openpyxl import Workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select


BACKEND_ROOT = Path(__file__).resolve().parents[1] / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.models.etl_lineage import LineageRef  # noqa: E402
from app.models.etl_registry import IngestionRun, IngestionStatus, Source  # noqa: E402
from app.models.models import EventSession  # noqa: E402
from app.models.stg_optin import StgOptinTransaction  # noqa: E402
from etl.extract import xlsx_optin  # noqa: E402


def _make_engine():
    """Create isolated in-memory SQLite engine for opt-in staging tests."""
    return create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )


def _create_required_tables(engine) -> None:  # noqa: ANN001
    """Create minimal registry + lineage + staging tables for tests."""
    SQLModel.metadata.create_all(
        engine,
        tables=[
            Source.__table__,
            IngestionRun.__table__,
            LineageRef.__table__,
            EventSession.__table__,
            StgOptinTransaction.__table__,
        ],
    )


def _write_optin_fixture(path: Path) -> None:
    """Write a minimal XLSX fixture with merged headers and one data row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "OptIn"

    ws["A1"] = "Relatorio Opt-In"
    ws["A2"] = "Gerado em 2025-12-12"

    ws.merge_cells("A4:C4")
    ws["A4"] = "Cliente"
    ws.merge_cells("D4:E4")
    ws["D4"] = "Venda"

    ws["A5"] = "Nome"
    ws["B5"] = "CPF"
    ws["C5"] = "Email"
    ws["D5"] = "Data Compra"
    ws["E5"] = "Opt In"
    ws["F5"] = "Opt In ID"
    ws["G5"] = "Evento"
    ws["H5"] = "Sessao"
    ws["I5"] = "Qtd Ingresso"

    ws["A6"] = "Alice"
    ws["B6"] = "11111111111"
    ws["C6"] = "alice@example.com"
    ws["D6"] = datetime(2025, 12, 12, 20, 30, 0)
    ws["E6"] = "Sim"
    ws["F6"] = "OID-1"
    ws["G6"] = "Tamo Junto"
    ws["H6"] = "Show 12/12"
    ws["I6"] = 2

    wb.save(path)


def test_extract_optin_xlsx_returns_canonical_rows_with_lineage_metadata(tmp_path: Path) -> None:
    """Extractor should output canonical opt-in fields and lineage metadata keys."""
    xlsx_path = tmp_path / "optin.xlsx"
    _write_optin_fixture(xlsx_path)

    rows = list(xlsx_optin.extract_optin_xlsx(xlsx_path))

    assert len(rows) == 1
    row = rows[0]
    assert row["evento"] == "Tamo Junto"
    assert row["sessao"] == "Show 12/12"
    assert row["opt_in"] == "Sim"
    assert row["opt_in_id"] == "OID-1"
    assert row["qtd_ingresso"] == 2
    assert row["cpf_hash"]
    assert row["email_hash"]
    assert row["__sheet_name"] == "OptIn"
    assert row["__header_row"] == 4
    assert row["__header_range"] == "A4:I5"
    assert row["__source_range"] == "A6:I6"


def test_load_optin_xlsx_to_staging_persists_ingestion_and_lineage(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Loader should persist staging rows with `ingestion_id` and `lineage_ref_id`."""
    xlsx_path = tmp_path / "optin.xlsx"
    _write_optin_fixture(xlsx_path)

    engine = _make_engine()
    _create_required_tables(engine)
    monkeypatch.setattr(xlsx_optin, "engine", engine)

    summary = xlsx_optin.load_optin_xlsx_to_staging(
        source_id="src_optin_doze",
        xlsx_path=xlsx_path,
        lineage_policy="required",
        severity_on_missing="partial",
    )

    assert summary["source_id"] == "SRC_OPTIN_DOZE"
    assert summary["rows_loaded"] == 1
    assert summary["ingestion_id"] > 0
    assert summary["status"] == IngestionStatus.SUCCESS.value

    with Session(engine) as session:
        run = session.get(IngestionRun, summary["ingestion_id"])
        assert run is not None
        assert run.status == IngestionStatus.SUCCESS
        staged = session.exec(select(StgOptinTransaction)).all()
        assert len(staged) == 1
        stg_row = staged[0]
        assert stg_row.ingestion_id == run.id
        assert stg_row.lineage_ref_id > 0
        assert stg_row.session_id is not None
        assert stg_row.sheet_name == "OptIn"
        assert stg_row.source_range == "A6:I6"
        lineage = session.get(LineageRef, stg_row.lineage_ref_id)
        assert lineage is not None
        assert lineage.location_type.value == "range"
        assert lineage.location_value == "range:A6:I6"


def test_load_optin_xlsx_to_staging_marks_failed_when_lineage_missing(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Required lineage policy should fail ingestion when lineage IDs are missing."""
    xlsx_path = tmp_path / "optin.xlsx"
    _write_optin_fixture(xlsx_path)

    engine = _make_engine()
    _create_required_tables(engine)
    monkeypatch.setattr(xlsx_optin, "engine", engine)

    class _DummyLineage:
        id = 0

    monkeypatch.setattr(xlsx_optin, "create_lineage_ref", lambda *args, **kwargs: _DummyLineage())

    with pytest.raises(ValueError, match="exige linhagem"):
        xlsx_optin.load_optin_xlsx_to_staging(
            source_id="SRC_OPTIN_MISSING_LINEAGE",
            xlsx_path=xlsx_path,
            lineage_policy="required",
            severity_on_missing="failed",
        )

    with Session(engine) as session:
        runs = session.exec(select(IngestionRun).order_by(IngestionRun.id.desc())).all()
        assert runs
        assert runs[0].status == IngestionStatus.FAILED
