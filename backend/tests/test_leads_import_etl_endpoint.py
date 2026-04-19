from __future__ import annotations

from datetime import date
from io import BytesIO
import json
from pathlib import Path

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select

from app.db.database import get_session
from app.main import app
from app.models.lead_public_models import LeadEventoSourceKind
from app.models.models import Evento, ImportAlias, Lead, LeadEvento, LeadImportEtlPreviewSession, StatusEvento, Usuario
from app.routers.leads import _map_etl_import_error
from app.schemas.lead_import_etl import ImportEtlPreviewResponse, ImportEtlResult
from app.utils.security import hash_password


def make_engine():
    return create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )


@pytest.fixture
def engine(monkeypatch):
    monkeypatch.setenv("SECRET_KEY", "secret-test")
    engine = make_engine()
    SQLModel.metadata.create_all(engine)
    return engine


@pytest.fixture
def client(engine):
    def override_get_session():
        with Session(engine) as session:
            yield session

    app.dependency_overrides[get_session] = override_get_session
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


def seed_user(engine, email="etl@example.com", password="Senha123!") -> Usuario:
    with Session(engine) as session:
        user = Usuario(
            email=email,
            password_hash=hash_password(password),
            tipo_usuario="npbb",
            ativo=True,
        )
        session.add(user)
        session.commit()
        session.refresh(user)
        return user


def seed_statuses(engine) -> None:
    with Session(engine) as session:
        for nome in ["Previsto", "A Confirmar", "Confirmado", "Realizado", "Cancelado"]:
            exists = session.exec(select(StatusEvento).where(StatusEvento.nome == nome)).first()
            if not exists:
                session.add(StatusEvento(nome=nome))
        session.commit()


def seed_event(engine, nome="ETL-EVENTO-TESTE") -> Evento:
    with Session(engine) as session:
        status = session.exec(select(StatusEvento).where(StatusEvento.nome == "Previsto")).first()
        assert status is not None
        evento = Evento(
            nome=nome,
            cidade="Brasilia",
            estado="DF",
            concorrencia=False,
            data_inicio_prevista=date(2099, 1, 1),
            status_id=status.id,
        )
        session.add(evento)
        session.commit()
        session.refresh(evento)
        return evento


def rename_event(engine, *, evento_id: int, nome: str) -> None:
    with Session(engine) as session:
        evento = session.get(Evento, evento_id)
        assert evento is not None
        evento.nome = nome
        session.add(evento)
        session.commit()


def seed_homonymous_event(engine, *, nome: str) -> Evento:
    return seed_event(engine, nome=nome)


def login_and_get_token(client: TestClient, email: str, password: str) -> str:
    resp = client.post("/auth/login", json={"email": email, "password": password})
    assert resp.status_code == 200
    return resp.json()["access_token"]


def make_xlsx_payload(rows: list[list[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def test_preview_etl_returns_session_token_and_dq_report(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    fixture_path = Path(__file__).parent / "fixtures" / "lead_import_sample.xlsx"
    with fixture_path.open("rb") as fh:
        res = client.post(
            "/leads/import/etl/preview",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("lead_import_sample.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"evento_id": str(evento.id), "strict": "false"},
        )

    assert res.status_code == 200
    payload = ImportEtlPreviewResponse.model_validate(res.json())
    assert payload.session_token
    assert payload.total_rows >= 1
    assert payload.valid_rows >= 1
    assert payload.invalid_rows >= 0
    assert {item.severity for item in payload.dq_report} <= {"info", "warning", "error"}


def test_preview_etl_accepts_csv_with_semicolon_delimiter(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    res = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={
            "file": (
                "lead_import_sample.csv",
                b"Email;CPF;Nome;Sessao\ncsv@example.com;529.982.247-25;Lead CSV;Show 1\n",
                "text/csv",
            )
        },
        data={"evento_id": str(evento.id), "strict": "false"},
    )

    assert res.status_code == 200
    payload = ImportEtlPreviewResponse.model_validate(res.json())
    assert payload.status == "previewed"
    assert payload.total_rows == 1
    assert payload.valid_rows == 1


def test_preview_etl_import_file_error_maps_to_400_with_exception_message(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    res = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("leads.txt", b"email\nfoo@example.com\n", "text/plain")},
        data={"evento_id": str(evento.id), "strict": "false"},
    )

    assert res.status_code == 400
    detail = res.json()["detail"]
    assert detail["code"] == "INVALID_FILE_TYPE"
    assert detail["field"] == "file"
    assert ".txt" in detail["message"]


def test_map_etl_import_error_value_error_maps_to_400_with_exception_message() -> None:
    with pytest.raises(HTTPException) as exc_info:
        _map_etl_import_error(ValueError("campo invalido"))

    err = exc_info.value
    assert err.status_code == 400
    assert err.detail == {
        "code": "ETL_INVALID_INPUT",
        "message": "campo invalido",
    }


def test_preview_etl_requests_header_row_when_cpf_header_is_not_detected(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    xlsx_payload = make_xlsx_payload(
        [
            ["Relatorio de Leads"],
            ["Nome", "Documento", "Sessao"],
            ["Alice", "529.982.247-25", "Show 1"],
        ]
    )

    res = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("sem-cpf.xlsx", xlsx_payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false"},
    )

    assert res.status_code == 200
    payload = res.json()
    assert payload["status"] == "header_required"
    assert "session_token" not in payload
    assert payload["required_fields"] == ["cpf"]
    assert payload.get("available_sheets") == ["Sheet"]
    assert payload.get("active_sheet") == "Sheet"


def test_preview_etl_accepts_forced_header_row_with_cpf(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    xlsx_payload = make_xlsx_payload(
        [
            ["Relatorio de Leads"],
            ["Nome", "CPF", "Sessao"],
            ["Alice", "529.982.247-25", "Show 1"],
        ]
    )

    res = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("cpf-forcado.xlsx", xlsx_payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false", "header_row": "2"},
    )

    assert res.status_code == 200
    payload = ImportEtlPreviewResponse.model_validate(res.json())
    assert payload.status == "previewed"
    assert payload.session_token
    assert payload.total_rows == 1


def test_preview_etl_forced_header_row_without_cpf_returns_columns(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    xlsx_payload = make_xlsx_payload(
        [
            ["Relatorio de Leads"],
            ["Nome", "Documento", "Sessao"],
            ["Alice", "529.982.247-25", "Show 1"],
        ]
    )

    res = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("cpf-a-mapear.xlsx", xlsx_payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false", "header_row": "2"},
    )

    assert res.status_code == 200
    payload = res.json()
    assert payload["status"] == "cpf_column_required"
    assert payload["header_row"] == 2
    assert payload["columns"][1] == {
        "column_index": 2,
        "column_letter": "B",
        "source_value": "Documento",
    }


def test_preview_etl_persists_cpf_header_alias_and_reuses_it(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    xlsx_payload = make_xlsx_payload(
        [
            ["Relatorio de Leads"],
            ["Nome", "Documento", "Sessao"],
            ["Alice", "529.982.247-25", "Show 1"],
        ]
    )

    first = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("alias-cpf.xlsx", xlsx_payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "evento_id": str(evento.id),
            "strict": "false",
            "header_row": "2",
            "field_aliases_json": json.dumps({"cpf": {"column_index": 2, "source_value": "Documento"}}),
        },
    )

    assert first.status_code == 200
    first_payload = ImportEtlPreviewResponse.model_validate(first.json())
    assert first_payload.status == "previewed"

    with Session(engine) as session:
        alias = session.exec(
            select(ImportAlias).where(
                ImportAlias.domain == "lead_import_etl_header",
                ImportAlias.field_name == "cpf",
                ImportAlias.source_value == "Documento",
            )
        ).first()
        assert alias is not None
        assert alias.canonical_value == "cpf"

    second = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("alias-cpf.xlsx", xlsx_payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false"},
    )

    assert second.status_code == 200
    second_payload = ImportEtlPreviewResponse.model_validate(second.json())
    assert second_payload.status == "previewed"

    commit = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={"session_token": second_payload.session_token, "evento_id": evento.id, "force_warnings": True},
    )

    assert commit.status_code == 200
    assert ImportEtlResult.model_validate(commit.json()).status == "committed"


def test_commit_etl_blocks_warnings_until_force_confirmation(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    warning_xlsx = make_xlsx_payload(
        [
            ["Email", "CPF", "Data Nascimento"],
            ["warning@example.com", "52998224725", "01/01/1990"],
            ["warning@example.com", "52998224725", "01/01/1990"],
        ]
    )
    preview = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("warning.xlsx", warning_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false"},
    )

    assert preview.status_code == 200
    session_token = preview.json()["session_token"]

    blocked = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={"session_token": session_token, "evento_id": evento.id, "force_warnings": False},
    )

    assert blocked.status_code == 409
    assert blocked.json()["detail"]["code"] == "ETL_COMMIT_BLOCKED"

    forced = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={"session_token": session_token, "evento_id": evento.id, "force_warnings": True},
    )

    assert forced.status_code == 200
    payload = ImportEtlResult.model_validate(forced.json())
    assert payload.status == "committed"
    assert payload.created + payload.updated >= 1


def test_commit_etl_serializes_public_schema(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    fixture_path = Path(__file__).parent / "fixtures" / "lead_import_sample.xlsx"
    with fixture_path.open("rb") as fh:
        preview = client.post(
            "/leads/import/etl/preview",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("lead_import_sample.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"evento_id": str(evento.id), "strict": "false"},
        )

    assert preview.status_code == 200
    session_token = preview.json()["session_token"]

    commit = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={"session_token": session_token, "evento_id": evento.id, "force_warnings": True},
    )
    assert commit.status_code == 200
    payload = ImportEtlResult.model_validate(commit.json())
    assert payload.status == "committed"
    assert payload.created + payload.updated >= 1
    assert {item.severity for item in payload.dq_report} <= {"info", "warning", "error"}

    with Session(engine) as session:
        assert len(session.exec(select(Lead)).all()) >= 1


def test_commit_etl_reuses_committed_result_idempotently(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    fixture_path = Path(__file__).parent / "fixtures" / "lead_import_sample.xlsx"
    with fixture_path.open("rb") as fh:
        preview = client.post(
            "/leads/import/etl/preview",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("lead_import_sample.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"evento_id": str(evento.id), "strict": "false"},
        )

    session_token = preview.json()["session_token"]
    first = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={"session_token": session_token, "evento_id": evento.id, "force_warnings": True},
    )
    second = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={"session_token": session_token, "evento_id": evento.id, "force_warnings": True},
    )

    assert first.status_code == 200
    assert second.status_code == 200
    assert second.json() == first.json()


def test_commit_etl_partial_failure_allows_retry_then_idempotent_committed(
    client: TestClient,
    engine,
    monkeypatch,
) -> None:
    from app.modules.leads_publicidade.application.etl_import import commit_service
    from app.modules.leads_publicidade.application.etl_import.persistence import (
        LeadBatchPersistenceResult,
        LeadPersistenceFailedRow,
        persist_lead_batch as real_persist_lead_batch,
    )

    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    xlsx_payload = make_xlsx_payload(
        [
            ["Email", "CPF", "Nome", "Sessao"],
            ["partial-retry-a@example.com", "52998224725", "Lead A", "Show 1"],
            ["partial-retry-b@example.com", "39053344705", "Lead B", "Show 1"],
        ]
    )
    preview = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("partial-retry.xlsx", xlsx_payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false"},
    )
    assert preview.status_code == 200
    session_token = preview.json()["session_token"]

    attempt = {"n": 0}

    def persist_stub(db, batch, canonical_evento_id=None):
        attempt["n"] += 1
        if attempt["n"] == 1:
            return LeadBatchPersistenceResult(
                created=1,
                updated=0,
                failed_rows=[
                    LeadPersistenceFailedRow(row_index=3, reason="injected persistence failure"),
                ],
            )
        return real_persist_lead_batch(db, batch, canonical_evento_id=canonical_evento_id)

    monkeypatch.setattr(commit_service, "persist_lead_batch", persist_stub)

    first = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={"session_token": session_token, "evento_id": evento.id, "force_warnings": True},
    )
    assert first.status_code == 200
    body1 = ImportEtlResult.model_validate(first.json())
    assert body1.status == "partial_failure"
    assert len(body1.persistence_failures) == 1
    assert body1.persistence_failures[0].row_number == 3
    assert "injected" in body1.persistence_failures[0].reason

    with Session(engine) as session:
        row = session.get(LeadImportEtlPreviewSession, session_token)
        assert row is not None
        assert row.status == "partial_failure"
        stored = json.loads(row.commit_result_json or "{}")
        assert stored.get("status") == "partial_failure"

    second = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={"session_token": session_token, "evento_id": evento.id, "force_warnings": True},
    )
    assert second.status_code == 200
    body2 = ImportEtlResult.model_validate(second.json())
    assert body2.status == "committed"
    assert attempt["n"] == 2

    third = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={"session_token": session_token, "evento_id": evento.id, "force_warnings": True},
    )
    assert third.status_code == 200
    assert third.json() == second.json()


def test_commit_etl_keeps_canonical_event_link_when_event_is_renamed_after_preview(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine, nome="ETL-EVENTO-RENAME-ORIGINAL")
    token = login_and_get_token(client, user.email, "Senha123!")

    xlsx_payload = make_xlsx_payload(
        [
            ["Email", "CPF", "Nome", "Sessao"],
            ["rename@example.com", "52998224725", "Lead Rename", "Show 1"],
        ]
    )
    preview = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("rename.xlsx", xlsx_payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false"},
    )

    assert preview.status_code == 200
    session_token = preview.json()["session_token"]

    rename_event(engine, evento_id=evento.id, nome="ETL-EVENTO-RENAMED")

    commit = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={"session_token": session_token, "evento_id": evento.id, "force_warnings": True},
    )

    assert commit.status_code == 200
    payload = ImportEtlResult.model_validate(commit.json())
    assert payload.status == "committed"

    with Session(engine) as session:
        lead = session.exec(select(Lead).where(Lead.email == "rename@example.com")).first()
        assert lead is not None
        lead_eventos = session.exec(select(LeadEvento).where(LeadEvento.lead_id == lead.id)).all()
        assert len(lead_eventos) == 1
        assert lead_eventos[0].evento_id == evento.id
        assert lead_eventos[0].source_kind == LeadEventoSourceKind.EVENT_DIRECT


def test_commit_etl_links_canonical_event_by_snapshot_evento_id_even_with_homonymous_events(
    client: TestClient,
    engine,
) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine, nome="ETL-EVENTO-HOMONIMO")
    outro_evento = seed_homonymous_event(engine, nome="ETL-EVENTO-HOMONIMO")
    token = login_and_get_token(client, user.email, "Senha123!")

    xlsx_payload = make_xlsx_payload(
        [
            ["Email", "CPF", "Nome", "Sessao"],
            ["homonimo@example.com", "52998224725", "Lead Homonimo", "Show 1"],
        ]
    )
    preview = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("homonimo.xlsx", xlsx_payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false"},
    )

    assert preview.status_code == 200
    session_token = preview.json()["session_token"]
    assert outro_evento.id != evento.id

    commit = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={"session_token": session_token, "evento_id": evento.id, "force_warnings": True},
    )

    assert commit.status_code == 200

    with Session(engine) as session:
        lead = session.exec(select(Lead).where(Lead.email == "homonimo@example.com")).first()
        assert lead is not None
        lead_eventos = session.exec(select(LeadEvento).where(LeadEvento.lead_id == lead.id)).all()
        assert len(lead_eventos) == 1
        assert lead_eventos[0].evento_id == evento.id
        assert lead_eventos[0].evento_id != outro_evento.id
        assert lead_eventos[0].source_kind == LeadEventoSourceKind.EVENT_DIRECT


def test_commit_etl_preserves_existing_lead_fields_after_event_rename(
    client: TestClient,
    engine,
) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine, nome="ETL-EVENTO-RENOMEAR-UPDATE")
    token = login_and_get_token(client, user.email, "Senha123!")

    with Session(engine) as session:
        existing = Lead(
            email="rename-existing@example.com",
            cpf="52998224725",
            nome="Lead Original",
            evento_nome="ETL-EVENTO-LEGADO",
            sessao="Show 1",
        )
        session.add(existing)
        session.commit()
        session.refresh(existing)
        existing_id = existing.id
        session.add(
            LeadEvento(
                lead_id=existing_id,
                evento_id=evento.id,
                source_kind=LeadEventoSourceKind.EVENT_DIRECT,
            )
        )
        session.commit()

    xlsx_payload = make_xlsx_payload(
        [
            ["Email", "CPF", "Nome", "Sessao"],
            ["rename-existing@example.com", "52998224725", "Lead Atualizado", "Show 1"],
        ]
    )
    preview = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("rename-update.xlsx", xlsx_payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false"},
    )

    assert preview.status_code == 200
    session_token = preview.json()["session_token"]

    rename_event(engine, evento_id=evento.id, nome="ETL-EVENTO-RENOMEADO-DEPOIS-DO-PREVIEW")

    commit = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={"session_token": session_token, "evento_id": evento.id, "force_warnings": True},
    )

    assert commit.status_code == 200
    payload = ImportEtlResult.model_validate(commit.json())
    assert payload.created == 0
    assert payload.updated == 1
    assert payload.status == "committed"

    with Session(engine) as session:
        leads = session.exec(select(Lead).where(Lead.email == "rename-existing@example.com")).all()
        assert len(leads) == 1
        assert leads[0].id == existing_id
        assert leads[0].nome == "Lead Original"

        lead_eventos = session.exec(select(LeadEvento).where(LeadEvento.lead_id == existing_id)).all()
        assert len(lead_eventos) == 1
        assert lead_eventos[0].evento_id == evento.id
        assert lead_eventos[0].source_kind == LeadEventoSourceKind.EVENT_DIRECT


def test_commit_etl_rejects_unknown_session_token(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    res = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={"session_token": "missing-token", "evento_id": evento.id, "force_warnings": True},
    )

    assert res.status_code == 404
    assert res.json()["detail"]["code"] == "ETL_SESSION_NOT_FOUND"


def test_commit_etl_strict_blocks_preview_with_validation_errors(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine, nome="ETL-EVENTO-STRICT")
    token = login_and_get_token(client, user.email, "Senha123!")

    invalid_xlsx = make_xlsx_payload(
        [
            ["Email", "CPF", "Nome Completo", "Sessao"],
            ["strict@example.com", "12345678901", "Lead Invalido", "Show 1"],
        ]
    )
    preview = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("invalid.xlsx", invalid_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "true"},
    )

    assert preview.status_code == 200
    payload = ImportEtlPreviewResponse.model_validate(preview.json())
    assert payload.invalid_rows == 1
    session_token = payload.session_token

    res = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={"session_token": session_token, "evento_id": evento.id, "force_warnings": True},
    )

    assert res.status_code == 409
    assert res.json()["detail"]["code"] == "ETL_COMMIT_BLOCKED"


def test_preview_etl_reports_invalid_email_and_cpf_validation_errors(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine, nome="ETL-EVENTO-VALIDATORS")
    token = login_and_get_token(client, user.email, "Senha123!")

    invalid_xlsx = make_xlsx_payload(
        [
            ["Email", "CPF", "Nome", "Sessao"],
            ["mal formado", "11111111111", "Lead Invalido", "Show 1"],
        ]
    )
    preview = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("invalid-fields.xlsx", invalid_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "true"},
    )

    assert preview.status_code == 200
    payload = ImportEtlPreviewResponse.model_validate(preview.json())
    assert payload.valid_rows == 0
    assert payload.invalid_rows == 1
    assert sum(payload.rejection_reason_counts.values()) == payload.invalid_rows

    rejected_rows = [
        item for item in payload.dq_report if item.check_id == "dq.preview.rejected_rows"
    ]
    assert rejected_rows
    assert rejected_rows[0].sample == [
        {
            "row_number": 2,
            "errors": ["email malformado", "CPF inválido"],
        }
    ]


def test_preview_etl_rejects_known_placeholder_cpf_sequence(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine, nome="ETL-EVENTO-CPF-PLACEHOLDER")
    token = login_and_get_token(client, user.email, "Senha123!")

    invalid_xlsx = make_xlsx_payload(
        [
            ["Email", "CPF", "Nome", "Sessao"],
            ["placeholder@example.com", "12345678909", "Lead Placeholder", "Show 1"],
        ]
    )
    preview = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("invalid-placeholder-cpf.xlsx", invalid_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "true"},
    )

    assert preview.status_code == 200
    payload = ImportEtlPreviewResponse.model_validate(preview.json())
    assert payload.valid_rows == 0
    assert payload.invalid_rows == 1

    rejected_rows = [
        item for item in payload.dq_report if item.check_id == "dq.preview.rejected_rows"
    ]
    assert rejected_rows
    assert rejected_rows[0].sample == [
        {
            "row_number": 2,
            "errors": ["CPF inválido"],
        }
    ]


def test_commit_etl_blocks_validation_errors_even_with_force_warnings(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine, nome="ETL-EVENTO-VALIDATION-BLOCK")
    token = login_and_get_token(client, user.email, "Senha123!")

    invalid_xlsx = make_xlsx_payload(
        [
            ["Email", "CPF", "Nome Completo", "Sessao"],
            ["strict@example.com", "12345678901", "Lead Invalido", "Show 1"],
        ]
    )
    preview = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("invalid.xlsx", invalid_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "true"},
    )

    assert preview.status_code == 200
    payload = ImportEtlPreviewResponse.model_validate(preview.json())
    assert payload.invalid_rows == 1

    commit = client.post(
        "/leads/import/etl/commit",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "session_token": payload.session_token,
            "evento_id": evento.id,
            "force_warnings": True,
        },
    )

    assert commit.status_code == 409
    assert commit.json()["detail"]["code"] == "ETL_COMMIT_BLOCKED"


def _make_two_sheet_xlsx() -> bytes:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Indice"
    ws0.append(["Resumo"])
    ws1 = wb.create_sheet("Dados")
    ws1.append(["CPF", "Email"])
    ws1.append(["52998224725", "segunda.aba@example.com"])
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def test_preview_etl_multi_sheet_requires_sheet_selection_then_succeeds(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")
    payload_bytes = _make_two_sheet_xlsx()

    res = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("multi.xlsx", payload_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false"},
    )
    assert res.status_code == 200
    first = res.json()
    assert first["status"] == "header_required"
    assert set(first["available_sheets"]) == {"Indice", "Dados"}

    res2 = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("multi.xlsx", payload_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false", "sheet_name": "Dados"},
    )
    assert res2.status_code == 200
    body = ImportEtlPreviewResponse.model_validate(res2.json())
    assert body.status == "previewed"
    assert body.valid_rows >= 1
    assert body.sheet_name == "Dados"
    assert "Dados" in (body.available_sheets or [])


def test_preview_etl_sheet_name_idempotency_preserves_sheet_metadata(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")
    payload_bytes = _make_two_sheet_xlsx()

    res1 = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("multi.xlsx", payload_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false", "sheet_name": "Dados"},
    )
    assert res1.status_code == 200
    body1 = ImportEtlPreviewResponse.model_validate(res1.json())
    assert body1.status == "previewed"
    assert body1.sheet_name == "Dados"
    assert body1.available_sheets == ["Indice", "Dados"]

    res2 = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("multi.xlsx", payload_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false", "sheet_name": "Dados"},
    )
    assert res2.status_code == 200
    body2 = ImportEtlPreviewResponse.model_validate(res2.json())
    assert body2.session_token == body1.session_token
    assert body2.sheet_name == "Dados"
    assert body2.available_sheets == ["Indice", "Dados"]

    with Session(engine) as session:
        rows = session.exec(select(LeadImportEtlPreviewSession)).all()

    assert len(rows) == 1


def test_preview_etl_invalid_sheet_name_returns_400(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    res = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={
            "file": (
                "multi.xlsx",
                _make_two_sheet_xlsx(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"evento_id": str(evento.id), "strict": "false", "sheet_name": "NaoExiste"},
    )
    assert res.status_code == 400
    assert res.json()["detail"]["code"] == "ETL_INVALID_INPUT"


def test_preview_etl_max_scan_rows_out_of_range_returns_422(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    res = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={
            "file": (
                "x.xlsx",
                make_xlsx_payload([["CPF"], ["52998224725"]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"evento_id": str(evento.id), "strict": "false", "max_scan_rows": "9999"},
    )
    assert res.status_code == 422
    assert res.json()["detail"]["field"] == "max_scan_rows"


def _make_late_header_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    for _ in range(49):
        ws.append(["preambulo", "x"])
    ws.append(["CPF", "Email"])
    ws.append(["52998224725", "late.header@example.com"])
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def test_preview_etl_late_header_row_requires_extended_scan(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")
    raw = _make_late_header_xlsx()

    res_fail = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("late.xlsx", raw, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false"},
    )
    assert res_fail.status_code == 200
    assert res_fail.json()["status"] == "header_required"

    res_ok = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("late.xlsx", raw, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"evento_id": str(evento.id), "strict": "false", "max_scan_rows": "55"},
    )
    assert res_ok.status_code == 200
    body = ImportEtlPreviewResponse.model_validate(res_ok.json())
    assert body.status == "previewed"
    assert body.valid_rows >= 1


def _make_merged_title_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A1:C1")
    ws["A1"] = "Relatorio Participantes"
    ws.append(["CPF", "Email", "Nome"])
    ws.append(["52998224725", "merged.title@example.com", "Ana"])
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def test_preview_etl_merged_title_row_above_header(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    res = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={
            "file": (
                "merged.xlsx",
                _make_merged_title_xlsx(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"evento_id": str(evento.id), "strict": "false"},
    )
    assert res.status_code == 200
    body = ImportEtlPreviewResponse.model_validate(res.json())
    assert body.status == "previewed"
    assert body.valid_rows >= 1


def test_preview_etl_tab_delimited_csv(client: TestClient, engine) -> None:
    seed_statuses(engine)
    user = seed_user(engine)
    evento = seed_event(engine)
    token = login_and_get_token(client, user.email, "Senha123!")

    csv_bytes = b"CPF\tEmail\n52998224725\ttab.csv@example.com\n"
    res = client.post(
        "/leads/import/etl/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("tab.csv", csv_bytes, "text/csv")},
        data={"evento_id": str(evento.id), "strict": "false"},
    )
    assert res.status_code == 200
    body = ImportEtlPreviewResponse.model_validate(res.json())
    assert body.status == "previewed"
    assert body.valid_rows == 1
