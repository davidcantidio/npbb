"""Unit tests for XLSX header detection and column normalization utilities."""

from __future__ import annotations

from openpyxl import Workbook

from etl.extract.xlsx_utils import HeaderNotFound, build_columns_with_metadata, find_header_row
from etl.transform.column_normalize import ColumnNormalizationConfig, normalize_column_names


def _build_merged_header_workbook() -> Workbook:
    """Create one workbook fixture with merged/grouped header rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "OptIn"

    ws["A1"] = "Relatorio Eventim"
    ws["A2"] = "Gerado em 2025-12-12"
    ws["A3"] = ""

    ws.merge_cells("A4:B4")
    ws["A4"] = "Cliente"
    ws.merge_cells("C4:D4")
    ws["C4"] = "Venda"

    ws["A5"] = "Nome"
    ws["B5"] = "CPF"
    ws["C5"] = "Data Compra"
    ws["D5"] = "Opt In"

    ws["A6"] = "Alice"
    ws["B6"] = "123"
    ws["C6"] = "2025-12-12"
    ws["D6"] = "Sim"
    return wb


def test_find_header_row_detects_group_header_above_required_terms() -> None:
    """Header detection should promote to merged group row when applicable."""
    wb = _build_merged_header_workbook()
    ws = wb["OptIn"]

    header_row = find_header_row(ws, required_terms=["CPF", "Opt In"], max_scan_rows=10)

    assert header_row == 4


def test_find_header_row_soft_fail_returns_header_not_found_for_missing_terms() -> None:
    """Soft-fail mode should return details instead of raising for missing required terms."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Relatorio"])
    ws.append(["Nome", "Documento", "Sessao"])

    result = find_header_row(ws, required_terms=["CPF"], max_scan_rows=10, soft_fail=True)

    assert isinstance(result, HeaderNotFound)
    assert result.required_terms == ("CPF",)
    assert result.scanned_rows == 2


def test_find_header_row_forced_row_validates_required_aliases() -> None:
    """Forced header row should honor explicit term aliases."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Relatorio"])
    ws.append(["Nome", "Documento", "Sessao"])

    result = find_header_row(
        ws,
        required_terms=["CPF"],
        term_aliases={"CPF": ["Documento"]},
        forced_row=2,
        soft_fail=True,
    )

    assert result == 2


def test_find_header_row_forced_row_without_cpf_returns_columns() -> None:
    """Forced row soft-fail should include the columns found in that row."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Relatorio"])
    ws.append(["Nome", "Documento", "Sessao"])

    result = find_header_row(ws, required_terms=["CPF"], forced_row=2, soft_fail=True)

    assert isinstance(result, HeaderNotFound)
    assert result.forced_row == 2
    assert [column.source_value for column in result.columns] == ["Nome", "Documento", "Sessao"]
    assert [column.column_letter for column in result.columns] == ["A", "B", "C"]


def test_find_header_row_default_still_raises_when_no_candidate_exists() -> None:
    """Legacy hard-fail behavior remains the default when no candidate row exists."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = ""

    try:
        find_header_row(ws, required_terms=["CPF"], max_scan_rows=10)
    except ValueError as exc:
        assert "Header row could not be identified" in str(exc)
    else:  # pragma: no cover - defensive assertion
        raise AssertionError("find_header_row should raise when no candidate exists")


def test_build_columns_with_metadata_handles_merged_headers_and_lineage() -> None:
    """Column builder should generate canonical columns and lineage metadata."""
    wb = _build_merged_header_workbook()
    ws = wb["OptIn"]
    aliases = {
        "venda_data_compra": "dt_hr_compra",
        "venda_opt_in": "opt_in_status",
    }

    result = build_columns_with_metadata(
        ws,
        header_row=4,
        header_depth=2,
        aliases=aliases,
    )

    assert result.columns == [
        "cliente_nome",
        "cliente_cpf",
        "dt_hr_compra",
        "opt_in_status",
    ]
    assert result.lineage.sheet_name == "OptIn"
    assert result.lineage.header_row == 4
    assert result.lineage.header_range == "A4:D5"
    assert result.lineage.used_range == "A4:D6"


def test_normalize_column_names_applies_alias_and_deduplicates() -> None:
    """Column normalization should apply aliases, fallbacks and stable dedupe suffixes."""
    cfg = ColumnNormalizationConfig(case="lower")
    aliases = {
        "dt_hr_compra": "data_hora_compra",
        "opt_in": "optin_status",
    }
    raw = ["  Dt Hr Compra  ", "DT-HR-COMPRA", "Opt-In", "", None, "Opt In"]

    normalized = normalize_column_names(raw, config=cfg, aliases=aliases)

    assert normalized == [
        "data_hora_compra",
        "data_hora_compra_2",
        "optin_status",
        "col_4",
        "col_5",
        "optin_status_2",
    ]

