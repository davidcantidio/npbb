
"""Rotas do fluxo classico de importacao e exportacao Gold."""

from __future__ import annotations

import csv
import io
import re
import tracemalloc
from collections.abc import Iterator

import httpx
from fastapi import APIRouter, Depends, File, Form, UploadFile, status
from fastapi import Response as FastAPIResponse
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlmodel import Session

from core.leads_etl.models import coerce_lead_field
from app.core.auth import get_current_user
from app.db.database import get_session
from app.models.models import Lead, Usuario
from app.modules.lead_imports.application.etl_import.persistence import (
    build_dedupe_key,
    find_existing_lead,
    merge_lead,
    persist_lead_batch,
)
from app.modules.lead_imports.application.leads_import_usecases import (
    importar_leads_usecase,
    preview_import_sample_usecase,
    validar_mapeamento_usecase,
)
from app.schemas.lead_import import LeadImportMapping
from app.services.leads_export import generate_gold_export
from app.utils.http_errors import raise_http_error

from ._shared import (
    ALLOWED_IMPORT_EXTENSIONS,
    BATCH_SIZE,
    BATCH_SUMMARY_LIMIT,
    LOG_MEMORY,
    MAX_IMPORT_FILE_BYTES,
    logger,
)

router = APIRouter()

CEP_CACHE: dict[str, dict[str, str]] = {}
CEP_CACHE_MAX = 500


def _fetch_cep_data(cep: str) -> dict[str, str] | None:
    if not cep or len(cep) != 8:
        return None
    cached = CEP_CACHE.get(cep)
    if cached is not None:
        return cached or None
    try:
        resp = httpx.get(f"https://viacep.com.br/ws/{cep}/json/", timeout=2.0)
        if resp.status_code != 200:
            CEP_CACHE[cep] = {}
            return None
        data = resp.json()
        if data.get("erro"):
            CEP_CACHE[cep] = {}
            return None
        result = {
            "endereco_rua": (data.get("logradouro") or "").strip(),
            "bairro": (data.get("bairro") or "").strip(),
            "cidade": (data.get("localidade") or "").strip(),
            "estado": (data.get("uf") or "").strip(),
        }
        CEP_CACHE[cep] = result
        if len(CEP_CACHE) > CEP_CACHE_MAX:
            CEP_CACHE.clear()
        return result
    except Exception:
        return None


def _find_existing_lead(session: Session, payload: dict[str, object]) -> Lead | None:
    return find_existing_lead(session, payload)


def _dedupe_key(payload: dict[str, object]) -> str | None:
    return build_dedupe_key(payload)


def _merge_lead(existing: Lead, payload: dict[str, object]) -> None:
    merge_lead(existing, payload)


def _process_batch(
    session: Session,
    batch: list[tuple[dict[str, object], int] | None],
) -> tuple[int, int, int, bool]:
    return persist_lead_batch(session, batch)


def _log_memory_usage(stage: str, batch_index: int | None = None) -> None:
    if not LOG_MEMORY:
        return
    if not tracemalloc.is_tracing():
        tracemalloc.start()
    current, peak = tracemalloc.get_traced_memory()
    logger.info(
        "Lead import memory stage=%s batch=%s current_mb=%.2f peak_mb=%.2f",
        stage,
        batch_index,
        current / (1024 * 1024),
        peak / (1024 * 1024),
    )


def _detect_csv_delimiter(sample_text: str) -> str:
    first_line = sample_text.splitlines()[0] if sample_text else ""
    comma_count = first_line.count(",")
    semi_count = first_line.count(";")
    return ";" if semi_count > comma_count else ","


def _score_row_tabular(row: list[str]) -> int:
    non_empty = [cell for cell in row if cell.strip()]
    if not row:
        return 0
    fill_ratio = len(non_empty) / max(len(row), 1)
    if len(non_empty) >= 2 and fill_ratio >= 0.5:
        return int(fill_ratio * 100)
    return 0


def _detect_data_start_index(rows: list[list[str]]) -> int:
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows[:50]):
        score = _score_row_tabular(row)
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx


def _read_csv_sample(file: UploadFile, max_rows: int) -> dict:
    file.file.seek(0)
    sample_bytes = file.file.read(4096)
    try:
        sample_text = sample_bytes.decode("utf-8-sig", errors="ignore")
    except Exception:
        sample_text = ""
    delimiter = _detect_csv_delimiter(sample_text)

    file.file.seek(0)
    text_stream = io.TextIOWrapper(file.file, encoding="utf-8-sig", errors="ignore", newline="")
    reader = csv.reader(text_stream, delimiter=delimiter)
    rows: list[list[str]] = []
    try:
        for row in reader:
            if row is None:
                continue
            rows.append([cell.strip() for cell in row])
            if len(rows) >= max_rows + 1:
                break
    finally:
        text_stream.detach()

    start_index = _detect_data_start_index(rows)
    headers = rows[start_index] if rows else []
    data_rows = rows[start_index + 1 :] if len(rows) > start_index + 1 else []
    return {
        "headers": headers,
        "rows": data_rows,
        "delimiter": delimiter,
        "start_index": start_index,
    }


def _read_xlsx_sample(file: UploadFile, max_rows: int) -> dict:
    file.file.seek(0)
    wb = load_workbook(file.file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows: list[list[str]] = []
    for row in ws.iter_rows(max_row=max_rows + 1, values_only=True):
        rows.append([("" if v is None else str(v)).strip() for v in row])
        if len(rows) >= max_rows + 1:
            break
    start_index = _detect_data_start_index(rows)
    headers = rows[start_index] if rows else []
    data_rows = rows[start_index + 1 :] if len(rows) > start_index + 1 else []
    return {
        "headers": headers,
        "rows": data_rows,
        "delimiter": None,
        "start_index": start_index,
    }


def _read_xlsx_rows(file: UploadFile) -> list[list[str]]:
    file.file.seek(0)
    wb = load_workbook(file.file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([("" if v is None else str(v)).strip() for v in row])
    return rows


def _iter_csv_data_rows(file: UploadFile, delimiter: str, start_index: int) -> Iterator[list[str]]:
    file.file.seek(0)
    text_stream = io.TextIOWrapper(file.file, encoding="utf-8-sig", errors="ignore", newline="")
    reader = csv.reader(text_stream, delimiter=delimiter)
    try:
        for idx, row in enumerate(reader):
            if idx <= start_index:
                continue
            yield [cell.strip() for cell in row]
    finally:
        text_stream.detach()


def _iter_xlsx_data_rows(file: UploadFile, start_index: int) -> Iterator[list[str]]:
    file.file.seek(0)
    wb = load_workbook(file.file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx <= start_index:
            continue
        yield [("" if v is None else str(v)).strip() for v in row]


def _ensure_mapping_has_essential(mappings: list[LeadImportMapping]) -> None:
    mapped_fields = {m.campo for m in mappings if m.campo}
    if "cpf" not in mapped_fields:
        raise_http_error(
            status.HTTP_400_BAD_REQUEST,
            code="MAPPING_MISSING_ESSENTIAL",
            message="O mapeamento deve incluir a coluna CPF.",
        )


def _column_samples(rows: list[list[str]], col_index: int, max_samples: int = 5) -> list[str]:
    samples: list[str] = []
    for row in rows:
        if col_index >= len(row):
            continue
        value = row[col_index].strip()
        if not value:
            continue
        samples.append(value)
        if len(samples) >= max_samples:
            break
    return samples


def _score_email(values: list[str]) -> float:
    if not values:
        return 0.0
    pattern = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
    hits = sum(1 for v in values if pattern.match(v.lower()))
    return hits / len(values)


def _score_cpf(values: list[str]) -> float:
    if not values:
        return 0.0
    pattern = re.compile(r"^\d{11}$")
    hits = 0
    for v in values:
        digits = "".join(ch for ch in v if ch.isdigit())
        if pattern.match(digits):
            hits += 1
    return hits / len(values)


def _score_phone_br(values: list[str]) -> float:
    if not values:
        return 0.0
    # DDDs validos (Brasil)
    valid_ddds = {
        "11", "12", "13", "14", "15", "16", "17", "18", "19",
        "21", "22", "24", "27", "28",
        "31", "32", "33", "34", "35", "37", "38",
        "41", "42", "43", "44", "45", "46",
        "47", "48", "49",
        "51", "53", "54", "55",
        "61", "62", "63", "64", "65", "66", "67", "68", "69",
        "71", "73", "74", "75", "77",
        "79",
        "81", "82", "83", "84", "85", "86", "87", "88", "89",
        "91", "92", "93", "94", "95", "96", "97", "98", "99",
    }
    hits = 0
    for v in values:
        digits = "".join(ch for ch in v if ch.isdigit())
        if digits.startswith("55"):
            digits = digits[2:]
        if len(digits) != 11:
            continue
        ddd = digits[:2]
        numero = digits[2:]
        if ddd not in valid_ddds:
            continue
        if not numero.startswith("9"):
            continue
        hits += 1
    return hits / len(values)


def _score_cep(values: list[str]) -> float:
    if not values:
        return 0.0
    hits = 0
    for v in values:
        digits = "".join(ch for ch in v if ch.isdigit())
        if len(digits) == 8:
            hits += 1
    return hits / len(values)


def _score_uf(values: list[str]) -> float:
    if not values:
        return 0.0
    ufs = {
        "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA",
        "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN",
        "RS", "RO", "RR", "SC", "SP", "SE", "TO",
    }
    hits = sum(1 for v in values if v.strip().upper() in ufs)
    return hits / len(values)


def _score_date(values: list[str]) -> float:
    if not values:
        return 0.0
    patterns = [
        re.compile(r"^\d{2}/\d{2}/\d{4}$"),
        re.compile(r"^\d{4}-\d{2}-\d{2}$"),
    ]
    hits = sum(1 for v in values if any(p.match(v.strip()) for p in patterns))
    return hits / len(values)


def _score_number(values: list[str]) -> float:
    if not values:
        return 0.0
    hits = 0
    for v in values:
        text = v.strip().replace(".", "").replace(",", ".")
        try:
            float(text)
            hits += 1
        except ValueError:
            continue
    return hits / len(values)


def _score_datetime(values: list[str]) -> float:
    if not values:
        return 0.0
    patterns = [
        re.compile(r"^\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}(:\d{2})?$"),
        re.compile(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}(:\d{2})?$"),
    ]
    hits = sum(1 for v in values if any(p.match(v.strip()) for p in patterns))
    return hits / len(values)


def _score_address(values: list[str]) -> float:
    if not values:
        return 0.0
    keywords = ("rua", "avenida", "av", "travessa", "alameda", "rodovia", "praca", "praça")
    hits = 0
    for v in values:
        text = v.strip().lower()
        if any(k in text for k in keywords):
            hits += 1
    return hits / len(values)


def _score_genero(values: list[str]) -> float:
    if not values:
        return 0.0
    vocab = {
        "masculino",
        "feminino",
        "homem",
        "mulher",
        "m",
        "f",
    }
    hits = 0
    for v in values:
        text = v.strip().lower()
        if text in vocab:
            hits += 1
    return hits / len(values)


def _apply_sample_penalty(score: float, sample_size: int) -> float:
    if sample_size <= 0:
        return 0.0
    if sample_size < 3:
        return score * (sample_size / 3)
    return score


def _infer_column_mapping(header: str, samples: list[str]) -> LeadImportMapping:
    header_lc = header.lower().strip()
    sample_size = len(samples)
    thresholds = {
        "email": 0.8,
        "cpf": 0.8,
        "telefone": 0.85,
        "cep": 0.85,
        "uf": 0.85,
        "datetime": 0.85,
        "data": 0.75,
        "numero": 0.85,
        "endereco": 0.7,
        "genero": 0.7,
    }
    scores = {
        "email": _apply_sample_penalty(_score_email(samples), sample_size),
        "cpf": _apply_sample_penalty(_score_cpf(samples), sample_size),
        "telefone": _apply_sample_penalty(_score_phone_br(samples), sample_size),
        "cep": _apply_sample_penalty(_score_cep(samples), sample_size),
        "uf": _apply_sample_penalty(_score_uf(samples), sample_size),
        "data": _apply_sample_penalty(_score_date(samples), sample_size),
        "datetime": _apply_sample_penalty(_score_datetime(samples), sample_size),
        "numero": _apply_sample_penalty(_score_number(samples), sample_size),
        "endereco": _apply_sample_penalty(_score_address(samples), sample_size),
        "genero": _apply_sample_penalty(_score_genero(samples), sample_size),
    }

    campo = None
    confianca = None
    allow_auto = sample_size >= 3

    if allow_auto and scores["email"] >= thresholds["email"]:
        campo = "email"
        confianca = scores["email"]
    elif allow_auto and scores["cpf"] >= thresholds["cpf"]:
        campo = "cpf"
        confianca = scores["cpf"]
    elif allow_auto and scores["telefone"] >= thresholds["telefone"]:
        campo = "telefone"
        confianca = scores["telefone"]
    elif allow_auto and scores["cep"] >= thresholds["cep"]:
        campo = "cep"
        confianca = scores["cep"]
    elif allow_auto and scores["uf"] >= thresholds["uf"]:
        campo = "estado"
        confianca = scores["uf"]
    elif allow_auto and scores["datetime"] >= thresholds["datetime"]:
        campo = "data_compra"
        confianca = scores["datetime"]
    elif "email" in header_lc:
        campo = "email"
        confianca = max(scores["email"], 0.6)
    elif "cpf" in header_lc:
        campo = "cpf"
        confianca = max(scores["cpf"], 0.6)
    elif "telefone" in header_lc or "fone" in header_lc or "cel" in header_lc:
        campo = "telefone"
        confianca = max(scores["telefone"], 0.6)
    elif "cep" in header_lc:
        campo = "cep"
        confianca = max(scores["cep"], 0.6)
    elif "uf" in header_lc or "estado" in header_lc:
        campo = "estado"
        confianca = max(scores["uf"], 0.6)
    elif "compra" in header_lc and "hora" in header_lc:
        campo = "data_compra"
        confianca = max(scores["datetime"], 0.6)
    elif "endereco" in header_lc or "rua" in header_lc or header_lc.startswith("av"):
        campo = "endereco_rua"
        confianca = max(scores["endereco"], 0.55)
    elif "genero" in header_lc or "gênero" in header_lc:
        campo = "genero"
        confianca = max(scores["genero"], 0.55)
    elif "data" in header_lc or "nasc" in header_lc:
        campo = "data_nascimento"
        confianca = max(scores["data"], 0.5)
    elif allow_auto and scores["data"] >= thresholds["data"]:
        campo = "data_nascimento"
        confianca = scores["data"]
    elif allow_auto and scores["numero"] >= thresholds["numero"]:
        campo = "ingresso_qtd"
        confianca = scores["numero"]
    elif allow_auto and scores["endereco"] >= thresholds["endereco"]:
        campo = "endereco_rua"
        confianca = scores["endereco"]
    elif allow_auto and scores["genero"] >= thresholds["genero"]:
        campo = "genero"
        confianca = scores["genero"]

    if confianca is not None:
        confianca = min(confianca, 0.9)
    return LeadImportMapping(coluna=header or "", campo=campo, confianca=confianca)


@router.post("/import/sample")
@router.post("/import/sample/")
def preview_import_sample(
    file: UploadFile = File(...),
    sample_rows: int = 10,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    return preview_import_sample_usecase(
        file=file,
        sample_rows=sample_rows,
        session=session,
        current_user=current_user,
        allowed_import_extensions=ALLOWED_IMPORT_EXTENSIONS,
        max_import_file_bytes=MAX_IMPORT_FILE_BYTES,
        read_xlsx_sample=_read_xlsx_sample,
        read_csv_sample=_read_csv_sample,
        column_samples=_column_samples,
        infer_column_mapping=_infer_column_mapping,
        normalize_alias_value=_normalize_alias_value,
        raise_http_error=raise_http_error,
    )


@router.post("/import/validate")
@router.post("/import/validate/")
def validar_mapeamento(
    mappings: list[LeadImportMapping],
    current_user: Usuario = Depends(get_current_user),
):
    _ = current_user
    return validar_mapeamento_usecase(
        mappings=mappings,
        ensure_mapping_has_essential=_ensure_mapping_has_essential,
    )


@router.post("/import")
@router.post("/import/")
def importar_leads(
    file: UploadFile = File(...),
    mappings_json: str = Form(...),
    fonte_origem: str | None = Form(None),
    enriquecer_cep: bool = Form(False),
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    return importar_leads_usecase(
        file=file,
        mappings_json=mappings_json,
        fonte_origem=fonte_origem,
        enriquecer_cep=enriquecer_cep,
        session=session,
        current_user=current_user,
        allowed_import_extensions=ALLOWED_IMPORT_EXTENSIONS,
        read_xlsx_sample=_read_xlsx_sample,
        read_csv_sample=_read_csv_sample,
        iter_xlsx_data_rows=_iter_xlsx_data_rows,
        iter_csv_data_rows=_iter_csv_data_rows,
        ensure_mapping_has_essential=_ensure_mapping_has_essential,
        dedupe_key=_dedupe_key,
        process_batch=_process_batch,
        fetch_cep_data=_fetch_cep_data,
        coerce_field=coerce_lead_field,
        log_memory_usage=_log_memory_usage,
        batch_size=BATCH_SIZE,
        batch_summary_limit=BATCH_SUMMARY_LIMIT,
        raise_http_error=raise_http_error,
    )


@router.post("/import/preview")
@router.post("/import/preview/")
def preview_import(
    file: UploadFile = File(...),
    sample_rows: int = 10,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    return preview_import_sample(
        file=file,
        sample_rows=sample_rows,
        session=session,
        current_user=current_user,
    )


@router.get("/export/gold")
@router.get("/export/gold/")
def exportar_leads_gold(
    evento_id: int | None = None,
    formato: str = "xlsx",
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    """Exports Gold-stage leads as .xlsx or .csv.

    Returns HTTP 204 when no leads are found for the selected filters.
    """
    _ = current_user
    resultado = generate_gold_export(db=session, evento_id=evento_id, formato=formato)

    if resultado is None:
        return FastAPIResponse(status_code=204)

    file_bytes, filename = resultado

    if formato == "csv":
        media_type = "text/csv; charset=utf-8"
    else:
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
