"""Servico de mapeamento Silver: sugestao de colunas e persistencia em leads_silver.

Fluxo:
  1. suggest_column_mapping(batch_id, db) -> list[ColumnSuggestion]
     Le arquivo_bronze do lote, extrai headers e retorna sugestao por coluna
     usando HEADER_SYNONYMS + LeadColumnAlias salvos.

  2. mapear_batch(batch_id, evento_id, mapeamento, user_id, db) -> MapearResult
     Le todas as linhas de arquivo_bronze, aplica mapeamento confirmado
     (coluna -> campo_canonico), persiste LeadSilver, salva aliases novos
     e atualiza stage do lote para silver.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlmodel import Session, select

from lead_pipeline.constants import HEADER_SYNONYMS
from lead_pipeline.normalization import canonicalize_header

from app.models.lead_batch import BatchStage, LeadBatch, LeadColumnAlias, LeadSilver


CANONICAL_FIELDS = {
    "nome",
    "cpf",
    "data_nascimento",
    "email",
    "telefone",
    "evento",
    "tipo_evento",
    "local",
    "data_evento",
}

Confidence = str  # "exact_match" | "synonym_match" | "alias_match" | "none"


@dataclass
class ColumnSuggestion:
    coluna_original: str
    campo_sugerido: str | None
    confianca: Confidence


@dataclass
class MapearResult:
    batch_id: int
    silver_count: int
    stage: str


def _detect_csv_delimiter(text: str) -> str:
    first_line = text.splitlines()[0] if text else ""
    return ";" if first_line.count(";") > first_line.count(",") else ","


def _read_headers_from_raw(raw: bytes, filename: str) -> list[str]:
    ext = Path(filename).suffix.lower()
    if ext == ".xlsx":
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(max_row=1, values_only=True):
            return [("" if v is None else str(v)).strip() for v in row]
        return []
    else:
        try:
            text = raw.decode("utf-8-sig", errors="ignore")
        except Exception:
            text = raw.decode("latin-1", errors="ignore")
        delimiter = _detect_csv_delimiter(text)
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        for row in reader:
            return [cell.strip() for cell in row]
        return []


def _read_all_rows_from_raw(raw: bytes, filename: str) -> tuple[list[str], list[list[str]]]:
    """Retorna (headers, data_rows) — todas as linhas de dados (sem limite)."""
    ext = Path(filename).suffix.lower()
    if ext == ".xlsx":
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        all_rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([("" if v is None else str(v)).strip() for v in row])
        headers = all_rows[0] if all_rows else []
        return headers, all_rows[1:]
    else:
        try:
            text = raw.decode("utf-8-sig", errors="ignore")
        except Exception:
            text = raw.decode("latin-1", errors="ignore")
        delimiter = _detect_csv_delimiter(text)
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = list(reader)
        if not rows:
            return [], []
        headers = [cell.strip() for cell in rows[0]]
        data_rows = [[cell.strip() for cell in row] for row in rows[1:] if any(cell.strip() for cell in row)]
        return headers, data_rows


def _suggest_for_header(
    header: str,
    aliases_by_coluna: dict[str, str],
) -> tuple[str | None, Confidence]:
    canonical = canonicalize_header(header)

    # exact_match: coluna já é um campo canônico
    if canonical in CANONICAL_FIELDS:
        return canonical, "exact_match"

    # synonym_match: sinônimo definido em HEADER_SYNONYMS
    synonym = HEADER_SYNONYMS.get(canonical)
    if synonym and synonym in CANONICAL_FIELDS:
        return synonym, "synonym_match"

    # alias_match: alias salvo de envios anteriores para a plataforma
    saved = aliases_by_coluna.get(header)
    if saved:
        return saved, "alias_match"

    # Também tenta alias com a versão canonicalizada
    saved_canonical = aliases_by_coluna.get(canonical)
    if saved_canonical:
        return saved_canonical, "alias_match"

    return None, "none"


def suggest_column_mapping(
    batch_id: int,
    db: Session,
) -> list[ColumnSuggestion]:
    """Retorna sugestoes automaticas de mapeamento para todas as colunas do lote."""
    batch = db.get(LeadBatch, batch_id)
    if not batch:
        raise ValueError(f"Lote {batch_id} não encontrado")

    headers = _read_headers_from_raw(batch.arquivo_bronze, batch.nome_arquivo_original)

    aliases = db.exec(
        select(LeadColumnAlias).where(
            LeadColumnAlias.plataforma_origem == batch.plataforma_origem
        )
    ).all()
    aliases_by_coluna: dict[str, str] = {a.nome_coluna_original: a.campo_canonico for a in aliases}

    suggestions: list[ColumnSuggestion] = []
    for header in headers:
        if not header:
            continue
        campo, confianca = _suggest_for_header(header, aliases_by_coluna)
        suggestions.append(ColumnSuggestion(
            coluna_original=header,
            campo_sugerido=campo,
            confianca=confianca,
        ))
    return suggestions


def mapear_batch(
    batch_id: int,
    evento_id: int,
    mapeamento: dict[str, str],
    user_id: int,
    db: Session,
) -> MapearResult:
    """Aplica mapeamento confirmado, persiste linhas silver, salva aliases, atualiza stage."""
    batch = db.get(LeadBatch, batch_id)
    if not batch:
        raise ValueError(f"Lote {batch_id} não encontrado")

    headers, data_rows = _read_all_rows_from_raw(batch.arquivo_bronze, batch.nome_arquivo_original)

    # Delete existing silver rows for this batch (idempotent re-mapping)
    existing_silver = db.exec(
        select(LeadSilver).where(LeadSilver.batch_id == batch_id)
    ).all()
    for row in existing_silver:
        db.delete(row)
    db.flush()

    # Persist silver rows
    silver_count = 0
    for row_index, row in enumerate(data_rows):
        dados_brutos: dict[str, Any] = {}
        for col_idx, header in enumerate(headers):
            campo_canonico = mapeamento.get(header)
            if not campo_canonico:
                continue
            value = row[col_idx] if col_idx < len(row) else ""
            dados_brutos[campo_canonico] = value

        if not dados_brutos:
            continue

        silver = LeadSilver(
            batch_id=batch_id,
            row_index=row_index,
            dados_brutos=dados_brutos,
            evento_id=evento_id,
        )
        db.add(silver)
        silver_count += 1

    # Save / upsert new aliases
    for coluna_original, campo_canonico in mapeamento.items():
        if not campo_canonico:
            continue
        existing_alias = db.exec(
            select(LeadColumnAlias).where(
                LeadColumnAlias.nome_coluna_original == coluna_original,
                LeadColumnAlias.plataforma_origem == batch.plataforma_origem,
            )
        ).first()
        if existing_alias:
            if existing_alias.campo_canonico != campo_canonico:
                existing_alias.campo_canonico = campo_canonico
                db.add(existing_alias)
        else:
            alias = LeadColumnAlias(
                nome_coluna_original=coluna_original,
                campo_canonico=campo_canonico,
                plataforma_origem=batch.plataforma_origem,
                criado_por=user_id,
            )
            db.add(alias)

    # Update batch stage and evento_id
    batch.stage = BatchStage.SILVER
    batch.evento_id = evento_id
    db.add(batch)

    db.commit()

    return MapearResult(
        batch_id=batch_id,
        silver_count=silver_count,
        stage="silver",
    )
