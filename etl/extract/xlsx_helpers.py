"""XLSX helpers for extractor implementations."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime, date
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet


_WS_SPACE_RE = re.compile(r"\s+")


def _clean_header(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = _WS_SPACE_RE.sub(" ", s)
    return s


def find_header_row(ws: Worksheet, *, required_terms: List[str], max_scan_rows: int = 30) -> int:
    """Find the row index that likely contains headers.

    Heuristic: first row where all required_terms appear at least once across the row.
    """
    req = [t.strip().lower() for t in required_terms if t.strip()]
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        row_vals = [_clean_header(ws.cell(row=r, column=c).value).lower() for c in range(1, ws.max_column + 1)]
        if all(any(term in v for v in row_vals) for term in req):
            return r
    raise ValueError("Header row not found with required terms: " + ", ".join(required_terms))


def get_row_values(ws: Worksheet, row: int, *, max_col: Optional[int] = None) -> List[Any]:
    max_c = max_col or ws.max_column
    vals = [ws.cell(row=row, column=c).value for c in range(1, max_c + 1)]
    # trim trailing None/empty
    while vals and (vals[-1] is None or str(vals[-1]).strip() == ""):
        vals.pop()
    return vals


def infer_last_data_row(ws: Worksheet, *, start_row: int, max_col: int, max_blank_rows: int = 20) -> int:
    """Infer last data row by scanning until N consecutive blank rows."""
    blank = 0
    last = start_row
    for r in range(start_row, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if all(v is None or str(v).strip() == "" for v in row_vals):
            blank += 1
            if blank >= max_blank_rows:
                return last
            continue
        blank = 0
        last = r
    return last


def to_iso(value: Any) -> Any:
    """Normalize datetime/date values to ISO strings for CSV/JSON outputs."""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


@dataclass(frozen=True)
class HeaderSpec:
    header_row: int
    headers: List[str]
    max_col: int

    def as_mapping(self) -> Dict[int, str]:
        return {i + 1: h for i, h in enumerate(self.headers)}


def build_optin_header_spec(ws: Worksheet) -> HeaderSpec:
    """Build a robust header spec for the Opt-In Aceitos files."""
    header_row = find_header_row(ws, required_terms=["Evento", "CPF"])
    raw = get_row_values(ws, header_row)
    max_col = len(raw)

    # Known template (by position) for this specific Eventim export.
    # This handles merged/group headers like CLIENTE/ENDERECO where subheaders are blank.
    headers = [""] * max_col
    for idx, v in enumerate(raw, start=1):
        headers[idx - 1] = _clean_header(v)

    # Fill blanks by position (template-specific)
    # Column numbers are 1-based.
    overrides = {
        5: "Opt In ID",
        9: "Nome",
        10: "Sobrenome",
        11: "Email",
        13: "Endereco Rua",
        14: "Endereco Numero",
        15: "Bairro",
    }
    for col, name in overrides.items():
        if 1 <= col <= max_col:
            headers[col - 1] = name

    # Also, standardize some known headers
    renames = {
        "SESSAO": "Sessao",
        "SESSÃO": "Sessao",
        "DT E HR COMPRA": "Dt Hr Compra",
        "METODO ENTREGA": "Metodo Entrega",
        "MÉTODO ENTREGA": "Metodo Entrega",
        "COD PROMOCIONAL": "Cod Promocional",
        "QTD INGRESSO": "Qtd Ingresso",
        "DT NASCIMENTO": "Dt Nascimento",
        "GÊNERO": "Genero",
        "ENDEREÇO": "Endereco",
        "OPT-IN": "Opt-in Status",
    }
    out: List[str] = []
    for h in headers:
        hh = h.strip()
        if not hh:
            out.append("")
            continue
        key = hh.upper()
        out.append(renames.get(key, hh))

    # Ensure no empty header remains; assign fallback names
    for i, h in enumerate(out):
        if not h:
            out[i] = f"col_{i+1}"

    return HeaderSpec(header_row=header_row, headers=out, max_col=max_col)


def build_simple_header_spec(ws: Worksheet, *, required_terms: List[str]) -> HeaderSpec:
    """Build header spec for simpler XLSX exports."""
    header_row = find_header_row(ws, required_terms=required_terms)
    raw = get_row_values(ws, header_row)
    max_col = len(raw)
    headers = [_clean_header(v) or f"col_{i+1}" for i, v in enumerate(raw)]
    return HeaderSpec(header_row=header_row, headers=headers, max_col=max_col)

